from __future__ import annotations

import csv
import random

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from PIL import Image, ImageDraw, ImageFont



FONT_PATH = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"
BOLD_FONT_PATH = "/usr/share/fonts/truetype/dejavu/DejaVuSans-Bold.ttf"


def _font(path: str, size: int) -> ImageFont.FreeTypeFont:
    if Path(path).is_file():
        return ImageFont.truetype(path, size)
    return ImageFont.load_default()


REGULAR = _font(FONT_PATH, 22)
SMALL = _font(FONT_PATH, 18)
BOLD = _font(BOLD_FONT_PATH, 22)
TITLE = _font(BOLD_FONT_PATH, 32)


def _scan_text(draw: ImageDraw.ImageDraw, text: str, x: int, y: int, font: ImageFont.FreeTypeFont, fill: tuple[int, int, int] = (35, 35, 35)) -> int:
    draw.text((x, y), text, font=font, fill=fill)
    return y + font.size + 8


def write_pdf(path: Path, vendor: str, invoice_number: str, invoice_date: date, lines: list[tuple[str, str, int, float, float]]) -> None:
    """Create an image-only, receipt-like PDF scan for OCR testing."""
    width, height = 1700, 2200
    image = Image.new("RGB", (width, height), (247, 246, 239))
    draw = ImageDraw.Draw(image)
    y = 90
    y = _scan_text(draw, vendor.upper(), 90, y, TITLE)
    y = _scan_text(draw, "BARREL & FLAME BAR + GRILL", 90, y, BOLD)
    y += 20
    y = _scan_text(draw, f"INVOICE #: {invoice_number}", 90, y, REGULAR)
    y = _scan_text(draw, f"INVOICE DATE: {invoice_date.isoformat()}", 90, y, REGULAR)
    y = _scan_text(draw, "BILL TO: BARREL & FLAME BAR + GRILL", 90, y, REGULAR)
    y = _scan_text(draw, "1840 RIVERFRONT DRIVE", 90, y, REGULAR)
    y += 25
    draw.line((90, y, width - 90, y), fill=(50, 50, 50), width=3)
    y += 22
    for heading in ("SKU", "DESCRIPTION", "QTY", "UNIT", "PRICE", "TOTAL"):
        x = {"SKU": 90, "DESCRIPTION": 250, "QTY": 960, "UNIT": 1080, "PRICE": 1330, "TOTAL": 1510}[heading]
        _scan_text(draw, heading, x, y, SMALL, (20, 45, 70))
    y += 36
    subtotal = 0.0
    for sku, name, qty, unit, price, total in lines:
        y = _scan_text(draw, sku, 90, y, SMALL)
        _scan_text(draw, name[:34], 250, y - SMALL.size - 8, SMALL)
        _scan_text(draw, str(qty), 960, y - SMALL.size - 8, SMALL)
        _scan_text(draw, unit[:15], 1080, y - SMALL.size - 8, SMALL)
        _scan_text(draw, f"${price:,.2f}", 1330, y - SMALL.size - 8, SMALL)
        y = _scan_text(draw, f"${total:,.2f}", 1510, y - SMALL.size - 8, SMALL)
        subtotal += total
    tax = round(subtotal * 0.025, 2)
    y += 20
    draw.line((90, y, width - 90, y), fill=(50, 50, 50), width=3)
    y += 24
    for label, value in (("SUBTOTAL", subtotal), ("TAX", tax), ("TOTAL", subtotal + tax)):
        y = _scan_text(draw, f"{label}: ${value:,.2f}", 1180, y, BOLD if label == "TOTAL" else REGULAR)
    y += 40
    _scan_text(draw, "PAYMENT TERMS: NET 15", 90, y, REGULAR)
    _scan_text(draw, "THANK YOU FOR YOUR BUSINESS", 90, y + 45, REGULAR)
    image.save(path, "PDF", resolution=150.0)

ROOT = Path(__file__).resolve().parent / "bar-grill-month"
INVOICES = ROOT / "invoices"
ROOT.mkdir(parents=True, exist_ok=True)
INVOICES.mkdir(parents=True, exist_ok=True)

START = date(2026, 7, 1)
END = date(2026, 7, 31)
random.seed(20260731)

ITEMS = [
    ("DG-001", "Dry Goods", "All-purpose flour", "50 lb bag", 22.50),
    ("DG-002", "Dry Goods", "French fries seasoning", "5 lb tub", 18.00),
    ("DG-003", "Dry Goods", "Tortilla chips", "6 lb case", 31.00),
    ("DG-004", "Dry Goods", "Hamburger buns", "24 count case", 38.00),
    ("FR-001", "Frozen Goods", "French fries", "30 lb case", 34.00),
    ("FR-002", "Frozen Goods", "Chicken wings", "40 lb case", 92.00),
    ("FR-003", "Frozen Goods", "Mozzarella sticks", "6 lb case", 41.00),
    ("FR-004", "Frozen Goods", "Breaded shrimp", "10 lb case", 64.00),
    ("DA-001", "Dairy", "Cheddar cheese", "40 lb case", 118.00),
    ("DA-002", "Dairy", "Sour cream", "5 lb tub", 13.50),
    ("DA-003", "Dairy", "Butter", "30 lb case", 86.00),
    ("DA-004", "Dairy", "Heavy cream", "1 gallon", 19.00),
    ("PR-001", "Produce", "Romaine lettuce", "24 head case", 29.00),
    ("PR-002", "Produce", "Tomatoes", "25 lb case", 24.00),
    ("PR-003", "Produce", "Onions", "25 lb case", 19.00),
    ("PR-004", "Produce", "Avocados", "48 count case", 46.00),
    ("PR-005", "Produce", "Lemons", "115 count case", 39.00),
    ("BV-001", "Beverage", "Cola syrup", "5 gallon box", 112.00),
    ("BV-002", "Beverage", "Bottled water", "24 count case", 8.50),
    ("BV-003", "Beverage", "Iced tea bags", "100 count box", 14.00),
    ("BV-004", "Beverage", "Lime juice", "1 gallon", 18.00),
    ("LI-001", "Liquor", "House vodka", "1.75 liter bottle", 21.00),
    ("LI-002", "Liquor", "House bourbon", "1.75 liter bottle", 29.00),
    ("LI-003", "Liquor", "Silver tequila", "1 liter bottle", 24.00),
    ("LI-004", "Liquor", "Triple sec", "1 liter bottle", 17.00),
    ("LI-005", "Liquor", "House rum", "1.75 liter bottle", 20.00),
    ("MI-001", "Misc", "Takeout containers", "200 count case", 42.00),
    ("MI-002", "Misc", "Paper napkins", "3000 count case", 34.00),
    ("MI-003", "Misc", "Food gloves", "1000 count case", 28.00),
    ("MI-004", "Misc", "Dishwasher detergent", "5 gallon pail", 76.00),
]

VENDORS = {
    "Dry Goods": ("Heartland Restaurant Supply", "HRS"),
    "Frozen Goods": ("Heartland Restaurant Supply", "HRS"),
    "Dairy": ("FreshRoute Foods", "FRF"),
    "Produce": ("FreshRoute Foods", "FRF"),
    "Beverage": ("Gulf Coast Beverage", "GCB"),
    "Liquor": ("Southern Spirits Distributors", "SSD"),
    "Misc": ("Heartland Restaurant Supply", "HRS"),
}


def write_csv(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.writer(handle)
        writer.writerow(headers)
        writer.writerows(rows)


def write_inventory() -> None:
    workbook = Workbook()
    beginning = workbook.active
    beginning.title = "Beginning Inventory"
    ending = workbook.create_sheet("Ending Inventory")
    headers = ["Count Date", "Item ID", "Vendor", "Vendor SKU", "Item Name", "Category", "Count Unit", "Counted Quantity", "Unit Cost", "Inventory Value", "Notes"]
    for sheet, count_date, factor in ((beginning, START, 1.0), (ending, END, .72)):
        sheet.append(["Barrel & Flame Bar + Grill", "Inventory Count", count_date.isoformat()])
        sheet.append(headers)
        for cell in sheet[2]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="17324D")
        for item_id, category, name, unit, price in ITEMS:
            vendor = VENDORS[category][0]
            qty = round((random.uniform(1.5, 8.0) if category in {"Liquor", "Beverage"} else random.uniform(4, 28)) * factor, 2)
            if category == "Liquor":
                qty = round(random.uniform(2, 18) * factor, 2)
            value = round(qty * price, 2)
            sheet.append([count_date.isoformat(), item_id, vendor, item_id, name, category, unit, qty, price, value, "Physical count"])
        sheet.freeze_panes = "A3"
        sheet.auto_filter.ref = f"A2:K{sheet.max_row}"
        for column, width in {"A":14, "B":12, "C":28, "D":14, "E":28, "F":16, "G":22, "H":18, "I":12, "J":16, "K":18}.items():
            sheet.column_dimensions[column].width = width
    workbook.save(ROOT / "inventory_counts.xlsx")


def write_recipes() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Recipes"
    headers = ["Menu Item Name", "POS Item Key", "Menu Category", "Menu Price", "Ingredient Name", "Inventory Item ID", "Quantity Count Units", "Count Unit", "Yield Percent", "Notes"]
    sheet.append(headers)
    recipes = {
        "Flame Burger": [("Hamburger buns", .04, "each"), ("Cheddar cheese", .08, "lb"), ("Tomatoes", .04, "lb"), ("Romaine lettuce", .03, "head")],
        "Buffalo Wings": [("Chicken wings", .45, "lb"), ("Butter", .03, "lb"), ("Lime juice", .01, "gallon")],
        "Fish Tacos": [("Breaded shrimp", .20, "lb"), ("Tortilla chips", .05, "lb"), ("Romaine lettuce", .02, "head"), ("Avocados", .03, "each")],
        "Loaded Fries": [("French fries", .35, "lb"), ("Cheddar cheese", .08, "lb"), ("Sour cream", .03, "lb"), ("Onions", .02, "lb")],
        "House Margarita": [("Silver tequila", .05, "liter"), ("Triple sec", .02, "liter"), ("Lime juice", .02, "gallon"), ("Lemons", .01, "case")],
        "Draft Beer": [("Cola syrup", .01, "gallon"), ("Bottled water", .02, "case")],
        "Chicken Basket": [("Chicken wings", .30, "lb"), ("French fries", .30, "lb"), ("Hamburger buns", .04, "each")],
        "Chips and Queso": [("Tortilla chips", .18, "lb"), ("Cheddar cheese", .10, "lb"), ("Sour cream", .02, "lb")],
        "Bourbon Smash": [("House bourbon", .05, "liter"), ("Lemons", .01, "case")],
        "Shrimp Basket": [("Breaded shrimp", .25, "lb"), ("French fries", .25, "lb"), ("Lemons", .01, "case")],
        "Side Salad": [("Romaine lettuce", .08, "head"), ("Tomatoes", .04, "lb"), ("Avocados", .02, "each")],
        "Fried Mozzarella": [("Mozzarella sticks", .25, "lb"), ("French fries", .15, "lb")],
        "House Rum Punch": [("House rum", .05, "liter"), ("Lime juice", .02, "gallon")],
        "Tequila Sunrise": [("Silver tequila", .05, "liter"), ("Lemons", .01, "case")],
        "Fish and Chips": [("Breaded shrimp", .22, "lb"), ("French fries", .30, "lb"), ("Lemons", .01, "case")],
        "Steak Sandwich": [("Hamburger buns", .04, "each"), ("Cheddar cheese", .06, "lb"), ("Onions", .04, "lb"), ("Romaine lettuce", .02, "head")],
        "Nacho Platter": [("Tortilla chips", .25, "lb"), ("Cheddar cheese", .15, "lb"), ("Avocados", .03, "each"), ("Tomatoes", .04, "lb")],
        "Garden Salad": [("Romaine lettuce", .10, "head"), ("Tomatoes", .05, "lb"), ("Avocados", .02, "each")],
        "Bottled Water": [("Bottled water", .04, "case")],
        "Lemonade": [("Lemons", .01, "case"), ("Lime juice", .01, "gallon")],
        "Vodka Lime": [("House vodka", .05, "liter"), ("Lime juice", .02, "gallon")],
        "Loaded Nachos": [("Tortilla chips", .20, "lb"), ("Cheddar cheese", .12, "lb"), ("Sour cream", .03, "lb"), ("Onions", .03, "lb")],
        "Wing Sampler": [("Chicken wings", .60, "lb"), ("French fries", .25, "lb")],
        "Burger Combo": [("Hamburger buns", .04, "each"), ("Cheddar cheese", .08, "lb"), ("French fries", .25, "lb"), ("Romaine lettuce", .02, "head")],
    }
    prices = {
        "Flame Burger": 14.99, "Buffalo Wings": 16.99, "Fish Tacos": 15.99,
        "Loaded Fries": 10.99, "House Margarita": 11.99, "Draft Beer": 6.50,
        "Chicken Basket": 15.99, "Chips and Queso": 9.99, "Bourbon Smash": 12.99,
        "Shrimp Basket": 17.99, "Side Salad": 8.99, "Fried Mozzarella": 10.99,
        "House Rum Punch": 11.99, "Tequila Sunrise": 10.99, "Fish and Chips": 18.99,
        "Steak Sandwich": 17.99, "Nacho Platter": 13.99, "Garden Salad": 10.99,
        "Bottled Water": 2.50, "Lemonade": 4.50, "Vodka Lime": 10.99,
        "Loaded Nachos": 14.99, "Wing Sampler": 24.99, "Burger Combo": 18.99,
    }
    ids = {name: item_id for item_id, _, name, _, _ in ITEMS}
    for menu, ingredients in recipes.items():
        for ingredient, quantity, unit in ingredients:
            sheet.append([menu, menu.upper().replace(" ", "_")[:20], "Bar & Grill", prices[menu], ingredient, ids[ingredient], quantity, unit, 100, "Standard portion"])
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="17324D")
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:J{sheet.max_row}"
    for column, width in {"A":22, "B":18, "C":16, "D":14, "E":24, "F":16, "G":20, "H":14, "I":14, "J":20}.items():
        sheet.column_dimensions[column].width = width
    workbook.save(ROOT / "recipe_guide.xlsx")


def write_sales() -> None:
    menu_prices = {
        "Flame Burger": 14.99, "Buffalo Wings": 16.99, "Fish Tacos": 15.99,
        "Loaded Fries": 10.99, "House Margarita": 11.99, "Draft Beer": 6.50,
        "Chicken Basket": 15.99, "Chips and Queso": 9.99, "Bourbon Smash": 12.99,
        "Shrimp Basket": 17.99, "Side Salad": 8.99, "Fried Mozzarella": 10.99,
        "House Rum Punch": 11.99, "Tequila Sunrise": 10.99, "Fish and Chips": 18.99,
        "Steak Sandwich": 17.99, "Nacho Platter": 13.99, "Garden Salad": 10.99,
        "Bottled Water": 2.50, "Lemonade": 4.50, "Vodka Lime": 10.99,
        "Loaded Nachos": 14.99, "Wing Sampler": 24.99, "Burger Combo": 18.99,
    }
    menu_categories = {
        "Flame Burger": "Dry Goods", "Buffalo Wings": "Frozen Goods", "Fish Tacos": "Frozen Goods",
        "Loaded Fries": "Dairy", "House Margarita": "Liquor", "Draft Beer": "Beverage",
        "Chicken Basket": "Frozen Goods", "Chips and Queso": "Misc", "Bourbon Smash": "Liquor",
        "Shrimp Basket": "Frozen Goods", "Side Salad": "Produce", "Fried Mozzarella": "Dairy",
        "House Rum Punch": "Liquor", "Tequila Sunrise": "Liquor", "Fish and Chips": "Frozen Goods",
        "Steak Sandwich": "Dry Goods", "Nacho Platter": "Misc", "Garden Salad": "Produce",
        "Bottled Water": "Beverage", "Lemonade": "Produce", "Vodka Lime": "Liquor",
        "Loaded Nachos": "Misc", "Wing Sampler": "Frozen Goods", "Burger Combo": "Dry Goods",
    }
    rows: list[list[object]] = []
    day = START
    order_no = 1000
    while day <= END:
        for _ in range(18 + (day.weekday() in {4, 5}) * 14):
            menu = random.choice(list(menu_prices))
            qty = random.randint(1, 4)
            price = menu_prices[menu]
            gross = round(qty * price, 2)
            discount = round(gross * random.choice([0, 0, .05]), 2)
            tax = round((gross - discount) * .0825, 2)
            rows.append([day.isoformat(), f"ORD-{order_no}", "Main Dining Room", "Dine In", menu, menu_categories[menu], qty, price, gross, discount, 0, round(gross - discount, 2), tax])
            order_no += 1
        day += timedelta(days=1)
    write_csv(ROOT / "sales_detail_july_2026.csv", ["Business Date", "Order ID", "Location", "Channel", "Menu Item Name", "Category", "Quantity", "Unit Price", "Gross Sales", "Discounts", "Refunds", "Net Sales", "Sales Tax"], rows)
    daily: dict[str, float] = {}
    for row in rows:
        daily[row[0]] = daily.get(row[0], 0) + float(row[11])
    summary = [[START.isoformat(), END.isoformat(), "2026-07", sum(float(row[8]) for row in rows), sum(float(row[9]) for row in rows), 0, sum(float(row[12]) for row in rows), sum(float(row[11]) for row in rows), "One month bar-and-grill sales summary"]]
    write_csv(ROOT / "sales_summary_july_2026.csv", ["Period Start", "Period End", "Month", "Gross Sales", "Discounts", "Refunds", "Sales Tax Collected", "Net Sales", "Notes"], summary)
    write_csv(ROOT / "daily_sales_july_2026.csv", ["Business Date", "Net Sales"], [[day, round(total, 2)] for day, total in sorted(daily.items())])


def main() -> None:
    for path in INVOICES.glob("*"):
        path.unlink()
    write_recipes()
    write_inventory()
    write_sales()
    invoice_lines: dict[str, list[tuple[str, str, int, str, float, float]]] = {}
    for item_id, category, name, unit, price in ITEMS:
        qty = random.randint(8, 24) if category not in {"Liquor", "Beverage"} else random.randint(2, 8)
        total = round(qty * price, 2)
        vendor = VENDORS[category][0]
        invoice_lines.setdefault(vendor, []).append((item_id, name, qty, unit, price, total))
    for index, (vendor, lines) in enumerate(sorted(invoice_lines.items()), 1):
        write_pdf(INVOICES / f"2026-07-{index:02d}_{vendor.replace(' ', '_')}_invoice.pdf", vendor, f"BF-0726-{index:03d}", START + timedelta(days=index - 1), lines)
    (ROOT / "README.md").write_text("""# Barrel & Flame Bar + Grill — MarginMise Test Data\n\nFictional one-month test dataset for July 2026. All records are synthetic.\n\n## Files\n- `invoices/`: three image-only PDF invoice scans, covering all requested category groupings across vendors.\n- `recipe_guide.xlsx`: bar-and-grill recipes using the inventory item IDs.\n- `sales_detail_july_2026.csv`: daily item-level sales for July.\n- `daily_sales_july_2026.csv`: daily net-sales series for dashboard charts.\n- `sales_summary_july_2026.csv`: month summary.\n- `inventory_counts.xlsx`: Beginning Inventory on July 1 and Ending Inventory on July 31.\n\nCategories represented in inventory, invoices, recipes, and/or sales inputs: Dry Goods, Frozen Goods, Dairy, Produce, Beverage, Liquor, and Misc.\n""", encoding="utf-8")


if __name__ == "__main__":
    main()
