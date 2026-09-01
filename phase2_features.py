#!/usr/bin/env python3
"""Phase 2 operations for Restaurant Cost Controller v2.7.

Provides POS-neutral sales imports, menu recipe costing, mobile inventory count
sessions, waste logging, vendor-ready purchase orders, and accounting exports.
All features remain local-first and manager-reviewed.
"""
from __future__ import annotations

import csv
import hashlib
import html
import io
import json
import math
import os
import re
import secrets
import socket
import sqlite3
import threading
import uuid
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any, Iterable
from urllib.parse import parse_qs, urlparse

from inventory_planning import preferred_sales_rows

MONEY = Decimal("0.01")
QTY = Decimal("0.0001")

PHASE2_SCHEMA_SQL = """
PRAGMA foreign_keys = ON;

CREATE TABLE IF NOT EXISTS pos_import_profiles (
    profile_id TEXT PRIMARY KEY,
    profile_name TEXT NOT NULL UNIQUE COLLATE NOCASE,
    source_type TEXT NOT NULL DEFAULT 'CSV',
    mapping_json TEXT NOT NULL,
    sheet_name TEXT,
    delimiter TEXT NOT NULL DEFAULT ',',
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS pos_import_runs (
    run_id TEXT PRIMARY KEY,
    profile_id TEXT,
    source_file TEXT NOT NULL,
    source_sha256 TEXT NOT NULL UNIQUE,
    imported_at TEXT NOT NULL,
    row_count INTEGER NOT NULL DEFAULT 0,
    rejected_count INTEGER NOT NULL DEFAULT 0,
    gross_sales TEXT NOT NULL DEFAULT '0.00',
    net_sales TEXT NOT NULL DEFAULT '0.00',
    status TEXT NOT NULL,
    errors_json TEXT,
    FOREIGN KEY(profile_id) REFERENCES pos_import_profiles(profile_id)
);

CREATE TABLE IF NOT EXISTS menu_items (
    menu_item_id TEXT PRIMARY KEY,
    pos_item_key TEXT UNIQUE COLLATE NOCASE,
    menu_item_name TEXT NOT NULL,
    category TEXT NOT NULL DEFAULT 'Unclassified',
    menu_price TEXT NOT NULL DEFAULT '0.00',
    active INTEGER NOT NULL DEFAULT 1,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_menu_items_name ON menu_items(menu_item_name);

CREATE TABLE IF NOT EXISTS pos_sales_lines (
    sale_line_id INTEGER PRIMARY KEY AUTOINCREMENT,
    run_id TEXT NOT NULL REFERENCES pos_import_runs(run_id) ON DELETE CASCADE,
    business_date TEXT NOT NULL,
    order_id TEXT,
    location TEXT,
    pos_item_key TEXT,
    menu_item_id TEXT REFERENCES menu_items(menu_item_id),
    menu_item_name TEXT NOT NULL,
    quantity TEXT NOT NULL,
    unit_price TEXT NOT NULL DEFAULT '0.00',
    gross_sales TEXT NOT NULL DEFAULT '0.00',
    discounts TEXT NOT NULL DEFAULT '0.00',
    refunds TEXT NOT NULL DEFAULT '0.00',
    net_sales TEXT NOT NULL DEFAULT '0.00',
    sales_tax TEXT NOT NULL DEFAULT '0.00',
    channel TEXT,
    modifiers TEXT,
    raw_json TEXT
);
CREATE INDEX IF NOT EXISTS idx_pos_sales_date ON pos_sales_lines(business_date);
CREATE INDEX IF NOT EXISTS idx_pos_sales_menu ON pos_sales_lines(menu_item_id, business_date);

CREATE TABLE IF NOT EXISTS recipe_ingredients (
    recipe_line_id INTEGER PRIMARY KEY AUTOINCREMENT,
    menu_item_id TEXT NOT NULL REFERENCES menu_items(menu_item_id) ON DELETE CASCADE,
    item_id TEXT NOT NULL REFERENCES items(item_id),
    quantity_count_units TEXT NOT NULL,
    yield_percent TEXT NOT NULL DEFAULT '100.00',
    notes TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL,
    UNIQUE(menu_item_id, item_id)
);

CREATE TABLE IF NOT EXISTS waste_events (
    waste_id TEXT PRIMARY KEY,
    event_date TEXT NOT NULL,
    item_id TEXT NOT NULL REFERENCES items(item_id),
    quantity_count_units TEXT NOT NULL,
    reason TEXT NOT NULL,
    shift TEXT,
    estimated_cost TEXT NOT NULL DEFAULT '0.00',
    notes TEXT,
    photo_path TEXT,
    created_by TEXT NOT NULL,
    created_at TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_waste_date ON waste_events(event_date);

CREATE TABLE IF NOT EXISTS mobile_count_sessions (
    session_id TEXT PRIMARY KEY,
    count_date TEXT NOT NULL,
    count_month TEXT NOT NULL,
    status TEXT NOT NULL DEFAULT 'Open',
    token_hash TEXT NOT NULL,
    expires_at TEXT NOT NULL,
    created_by TEXT NOT NULL,
    created_at TEXT NOT NULL,
    submitted_at TEXT,
    finalized_at TEXT,
    notes TEXT
);

CREATE TABLE IF NOT EXISTS mobile_count_entries (
    entry_id INTEGER PRIMARY KEY AUTOINCREMENT,
    session_id TEXT NOT NULL REFERENCES mobile_count_sessions(session_id) ON DELETE CASCADE,
    item_id TEXT NOT NULL REFERENCES items(item_id),
    quantity_on_hand TEXT NOT NULL DEFAULT '0',
    notes TEXT,
    updated_at TEXT NOT NULL,
    UNIQUE(session_id, item_id)
);

CREATE TABLE IF NOT EXISTS vendor_order_profiles (
    vendor_name TEXT PRIMARY KEY COLLATE NOCASE,
    vendor_email TEXT,
    account_number TEXT,
    delivery_days TEXT,
    payment_terms TEXT,
    po_notes TEXT,
    updated_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS purchase_orders (
    po_id TEXT PRIMARY KEY,
    batch_id TEXT NOT NULL,
    vendor_name TEXT NOT NULL,
    po_date TEXT NOT NULL,
    expected_delivery_date TEXT,
    status TEXT NOT NULL DEFAULT 'Draft',
    subtotal TEXT NOT NULL DEFAULT '0.00',
    created_by TEXT NOT NULL,
    created_at TEXT NOT NULL,
    approved_at TEXT,
    notes TEXT,
    UNIQUE(batch_id, vendor_name)
);
CREATE INDEX IF NOT EXISTS idx_purchase_orders_status ON purchase_orders(status, po_date DESC);

CREATE TABLE IF NOT EXISTS purchase_order_lines (
    po_line_id INTEGER PRIMARY KEY AUTOINCREMENT,
    po_id TEXT NOT NULL REFERENCES purchase_orders(po_id) ON DELETE CASCADE,
    prediction_id INTEGER,
    item_id TEXT REFERENCES items(item_id),
    vendor_sku TEXT,
    item_name TEXT NOT NULL,
    quantity TEXT NOT NULL,
    purchase_unit TEXT,
    unit_price TEXT NOT NULL DEFAULT '0.00',
    line_total TEXT NOT NULL DEFAULT '0.00',
    notes TEXT
);

CREATE TABLE IF NOT EXISTS accounting_mappings (
    mapping_key TEXT PRIMARY KEY,
    debit_account TEXT NOT NULL,
    credit_account TEXT NOT NULL,
    updated_at TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS accounting_export_history (
    export_id TEXT PRIMARY KEY,
    export_type TEXT NOT NULL,
    period_start TEXT NOT NULL,
    period_end TEXT NOT NULL,
    file_path TEXT NOT NULL,
    row_count INTEGER NOT NULL,
    total_debits TEXT NOT NULL,
    total_credits TEXT NOT NULL,
    created_by TEXT NOT NULL,
    created_at TEXT NOT NULL
);
"""

DEFAULT_MAPPING_SYNONYMS: dict[str, tuple[str, ...]] = {
    "business_date": ("business date", "sale date", "date", "trading date", "order date", "transaction date"),
    "order_id": ("order id", "ticket id", "check id", "transaction id", "receipt number", "order number"),
    "location": ("location", "store", "restaurant", "site"),
    "pos_item_key": ("pos item key", "item id", "product id", "sku", "plu", "menu item id", "item code"),
    "menu_item_name": ("menu item name", "menu item", "item name", "product", "product name", "description", "item"),
    "quantity": ("quantity", "qty", "units sold", "item quantity", "count"),
    "unit_price": ("unit price", "price", "item price", "average price"),
    "gross_sales": ("gross sales", "gross amount", "gross revenue", "extended price", "sales"),
    "discounts": ("discounts", "discount amount", "discount"),
    "refunds": ("refunds", "refund amount", "returns", "void amount"),
    "net_sales": ("net sales", "net amount", "net revenue", "net"),
    "sales_tax": ("tax", "sales tax", "tax amount"),
    "channel": ("channel", "order type", "service mode", "source"),
    "modifiers": ("modifiers", "modifier", "options", "add ons", "add-ons"),
}

REQUIRED_POS_FIELDS = ("business_date", "menu_item_name", "quantity")


def now_iso() -> str:
    return datetime.now().replace(microsecond=0).isoformat()


def dec(value: Any, default: str = "0") -> Decimal:
    if value in (None, ""):
        return Decimal(default)
    try:
        return Decimal(str(value).replace(",", "").replace("$", "").strip())
    except (InvalidOperation, ValueError):
        return Decimal(default)


def money(value: Any) -> Decimal:
    return dec(value).quantize(MONEY, rounding=ROUND_HALF_UP)


def qty(value: Any) -> Decimal:
    return dec(value).quantize(QTY, rounding=ROUND_HALF_UP)


def normalize(value: Any) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip())


def safe_name(value: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip()).strip("._") or "export"


def sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def parse_date(value: Any) -> str:
    text = str(value or "").strip()
    if not text:
        raise ValueError("Missing business date")
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%m/%d/%y", "%m-%d-%y", "%Y/%m/%d"):
        try:
            return datetime.strptime(text.split()[0], fmt).date().isoformat()
        except ValueError:
            continue
    try:
        # Excel numeric date values occasionally arrive as floats in CSV exports.
        serial = float(text)
        if serial > 20000:
            return (date(1899, 12, 30) + timedelta(days=int(serial))).isoformat()
    except ValueError:
        pass
    raise ValueError(f"Unsupported date: {value!r}")


@dataclass
class POSImportResult:
    run_id: str
    imported: int
    rejected: int
    gross_sales: Decimal
    net_sales: Decimal
    errors: list[str]
    mapping: dict[str, str]


@dataclass
class MobileServerHandle:
    server: ThreadingHTTPServer
    thread: threading.Thread
    url: str
    session_id: str

    def stop(self) -> None:
        self.server.shutdown()
        self.server.server_close()
        if self.thread.is_alive():
            self.thread.join(timeout=2)


class Phase2Error(RuntimeError):
    pass


class Phase2Service:
    def __init__(self, workspace: Any, planning: Any, controls: Any):
        self.workspace = workspace
        self.planning = planning
        self.controls = controls
        self.mobile_handle: MobileServerHandle | None = None
        self.ensure_schema()

    def ensure_schema(self) -> None:
        folders = {
            "pos": self.workspace.root / "POS Imports",
            "recipes": self.workspace.root / "Recipes",
            "waste": self.workspace.root / "Waste Logs",
            "purchase_orders": self.workspace.root / "Purchase Orders",
            "accounting": self.workspace.root / "Accounting Exports",
            "mobile_counts": self.workspace.root / "Mobile Counts",
        }
        for key, path in folders.items():
            self.workspace.folders.setdefault(key, path)
            Path(path).mkdir(parents=True, exist_ok=True)
        with self.workspace.connect() as conn:
            conn.executescript(PHASE2_SCHEMA_SQL)
            stamp = now_iso()
            defaults = {
                "inventory_purchase": ("Inventory Purchases", "Accounts Payable"),
                "sales": ("Undeposited Funds", "Sales Revenue"),
                "sales_tax": ("Undeposited Funds", "Sales Tax Payable"),
                "discounts": ("Sales Discounts", "Sales Revenue"),
                "refunds": ("Sales Returns", "Undeposited Funds"),
                "operating_cost": ("Operating Expense", "Cash / Accounts Payable"),
            }
            for key, (debit, credit) in defaults.items():
                conn.execute(
                    """INSERT INTO accounting_mappings(mapping_key,debit_account,credit_account,updated_at)
                       VALUES(?,?,?,?) ON CONFLICT(mapping_key) DO NOTHING""",
                    (key, debit, credit, stamp),
                )

    # ------------------------------------------------------------------
    # POS-neutral sales import
    # ------------------------------------------------------------------
    def read_table(self, path: Path, sheet_name: str | None = None) -> tuple[list[str], list[dict[str, Any]]]:
        path = Path(path)
        if path.suffix.lower() == ".csv":
            with path.open("r", encoding="utf-8-sig", newline="") as fh:
                sample = fh.read(4096)
                fh.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
                except csv.Error:
                    dialect = csv.excel
                reader = csv.DictReader(fh, dialect=dialect)
                headers = [str(h or "").strip() for h in (reader.fieldnames or [])]
                return headers, [dict(row) for row in reader]
        if path.suffix.lower() in {".xlsx", ".xlsm"}:
            try:
                from openpyxl import load_workbook  # type: ignore
            except ImportError as exc:
                raise Phase2Error("openpyxl is required for Excel POS imports") from exc
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
            rows = ws.iter_rows(values_only=True)
            try:
                raw_headers = next(rows)
            except StopIteration:
                return [], []
            headers = [str(v or "").strip() for v in raw_headers]
            output: list[dict[str, Any]] = []
            for values in rows:
                if not any(v not in (None, "") for v in values):
                    continue
                output.append({headers[i]: values[i] if i < len(values) else "" for i in range(len(headers))})
            return headers, output
        raise Phase2Error("POS import supports CSV, XLSX, and XLSM files.")

    def suggest_mapping(self, headers: Iterable[str]) -> dict[str, str]:
        normalized_headers = {normalize(h): h for h in headers if str(h).strip()}
        mapping: dict[str, str] = {}
        used_headers: set[str] = set()
        field_order = [
            *REQUIRED_POS_FIELDS,
            *(field for field in DEFAULT_MAPPING_SYNONYMS if field not in REQUIRED_POS_FIELDS),
        ]
        for field in field_order:
            synonyms = DEFAULT_MAPPING_SYNONYMS[field]
            best: tuple[int, str] | None = None
            for norm_header, original in normalized_headers.items():
                if norm_header in used_headers:
                    continue
                for synonym in synonyms:
                    target = normalize(synonym)
                    if norm_header == target:
                        score = 1000 + len(target)
                    elif target and target in norm_header:
                        score = 100 + len(target)
                    elif norm_header and norm_header in target:
                        score = 80 + len(norm_header)
                    else:
                        score = 0
                    if score and (best is None or score > best[0]):
                        best = (score, original)
            if best:
                mapping[field] = best[1]
                used_headers.add(normalize(best[1]))
        return mapping

    def save_pos_profile(self, profile_name: str, mapping: dict[str, str], source_type: str = "CSV", sheet_name: str = "") -> str:
        profile_name = str(profile_name or "").strip()
        if not profile_name:
            raise Phase2Error("A mapping profile name is required.")
        missing = [field for field in REQUIRED_POS_FIELDS if not mapping.get(field)]
        if missing:
            raise Phase2Error("Mapping is missing required fields: " + ", ".join(missing))
        stamp = now_iso()
        profile_id = f"POSMAP-{uuid.uuid4().hex[:12].upper()}"
        with self.workspace.connect() as conn:
            existing = conn.execute("SELECT profile_id FROM pos_import_profiles WHERE profile_name=?", (profile_name,)).fetchone()
            if existing:
                profile_id = existing["profile_id"]
            conn.execute(
                """INSERT INTO pos_import_profiles(profile_id,profile_name,source_type,mapping_json,sheet_name,created_at,updated_at)
                   VALUES(?,?,?,?,?,?,?) ON CONFLICT(profile_id) DO UPDATE SET profile_name=excluded.profile_name,
                   source_type=excluded.source_type,mapping_json=excluded.mapping_json,sheet_name=excluded.sheet_name,updated_at=excluded.updated_at""",
                (profile_id, profile_name, source_type.upper(), json.dumps(mapping), sheet_name, stamp, stamp),
            )
        return profile_id

    def list_pos_profiles(self) -> list[sqlite3.Row]:
        with self.workspace.connect() as conn:
            return conn.execute("SELECT * FROM pos_import_profiles ORDER BY profile_name").fetchall()

    def get_pos_profile(self, profile_id: str) -> dict[str, Any]:
        with self.workspace.connect() as conn:
            row = conn.execute("SELECT * FROM pos_import_profiles WHERE profile_id=?", (profile_id,)).fetchone()
        if not row:
            raise Phase2Error("POS mapping profile not found.")
        data = dict(row)
        data["mapping"] = json.loads(data.pop("mapping_json") or "{}")
        return data

    def _get_or_create_menu_item(self, conn: sqlite3.Connection, key: str, name: str, unit_price: Decimal) -> str:
        key = str(key or "").strip() or normalize(name)
        row = conn.execute(
            "SELECT menu_item_id FROM menu_items WHERE pos_item_key=? COLLATE NOCASE", (key,)
        ).fetchone()
        stamp = now_iso()
        if row:
            conn.execute(
                "UPDATE menu_items SET menu_item_name=?,menu_price=CASE WHEN ?>0 THEN ? ELSE menu_price END,updated_at=? WHERE menu_item_id=?",
                (name, float(unit_price), f"{unit_price:.2f}", stamp, row["menu_item_id"]),
            )
            return row["menu_item_id"]
        menu_item_id = f"MENU-{hashlib.sha256(key.encode()).hexdigest()[:14].upper()}"
        conn.execute(
            "INSERT INTO menu_items(menu_item_id,pos_item_key,menu_item_name,menu_price,created_at,updated_at) VALUES(?,?,?,?,?,?)",
            (menu_item_id, key, name, f"{unit_price:.2f}", stamp, stamp),
        )
        return menu_item_id

    def import_pos_report(
        self,
        path: Path,
        *,
        mapping: dict[str, str] | None = None,
        profile_id: str | None = None,
        profile_name: str | None = None,
        sheet_name: str | None = None,
    ) -> POSImportResult:
        path = Path(path)
        digest = sha256_file(path)
        with self.workspace.connect() as conn:
            duplicate = conn.execute("SELECT run_id FROM pos_import_runs WHERE source_sha256=?", (digest,)).fetchone()
        if duplicate:
            raise Phase2Error(f"This POS report was already imported as {duplicate['run_id']}.")
        profile: dict[str, Any] | None = None
        if profile_id:
            profile = self.get_pos_profile(profile_id)
            mapping = profile["mapping"]
            sheet_name = sheet_name or profile.get("sheet_name")
        headers, rows = self.read_table(path, sheet_name)
        mapping = dict(mapping or self.suggest_mapping(headers))
        missing = [field for field in REQUIRED_POS_FIELDS if not mapping.get(field)]
        if missing:
            raise Phase2Error("Could not identify required POS columns: " + ", ".join(missing))
        if profile_name:
            profile_id = self.save_pos_profile(profile_name, mapping, path.suffix.lstrip("."), sheet_name or "")
        run_id = f"POS-{datetime.now().strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:6].upper()}"
        errors: list[str] = []
        imported = rejected = 0
        gross_total = net_total = Decimal("0")
        daily: dict[str, dict[str, Decimal]] = defaultdict(lambda: defaultdict(Decimal))
        stamp = now_iso()
        with self.workspace.connect() as conn:
            conn.execute(
                """INSERT INTO pos_import_runs(run_id,profile_id,source_file,source_sha256,imported_at,status)
                   VALUES(?,?,?,?,?,'Processing')""",
                (run_id, profile_id, path.name, digest, stamp),
            )
            for index, row in enumerate(rows, 2):
                try:
                    def value(field: str, default: Any = "") -> Any:
                        header = mapping.get(field)
                        return row.get(header, default) if header else default
                    business_date = parse_date(value("business_date"))
                    name = str(value("menu_item_name") or "").strip()
                    if not name:
                        raise ValueError("missing menu item name")
                    quantity_value = qty(value("quantity", "1"))
                    if quantity_value == 0:
                        continue
                    unit_price = money(value("unit_price"))
                    gross = money(value("gross_sales"))
                    if gross == 0 and unit_price:
                        gross = money(quantity_value * unit_price)
                    discounts = abs(money(value("discounts")))
                    refunds = abs(money(value("refunds")))
                    net = money(value("net_sales"))
                    if net == 0 and gross:
                        net = money(gross - discounts - refunds)
                    if gross == 0 and net:
                        gross = money(net + discounts + refunds)
                    tax = money(value("sales_tax"))
                    if unit_price == 0 and quantity_value:
                        unit_price = money(gross / quantity_value)
                    pos_key = str(value("pos_item_key") or normalize(name)).strip()
                    menu_item_id = self._get_or_create_menu_item(conn, pos_key, name, unit_price)
                    conn.execute(
                        """INSERT INTO pos_sales_lines(run_id,business_date,order_id,location,pos_item_key,menu_item_id,
                           menu_item_name,quantity,unit_price,gross_sales,discounts,refunds,net_sales,sales_tax,channel,modifiers,raw_json)
                           VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                        (
                            run_id, business_date, str(value("order_id") or ""), str(value("location") or ""), pos_key,
                            menu_item_id, name, f"{quantity_value:.4f}", f"{unit_price:.2f}", f"{gross:.2f}",
                            f"{discounts:.2f}", f"{refunds:.2f}", f"{net:.2f}", f"{tax:.2f}",
                            str(value("channel") or ""), str(value("modifiers") or ""), json.dumps(row, default=str),
                        ),
                    )
                    imported += 1
                    gross_total += gross
                    net_total += net
                    daily[business_date]["gross"] += gross
                    daily[business_date]["discounts"] += discounts
                    daily[business_date]["refunds"] += refunds
                    daily[business_date]["tax"] += tax
                    daily[business_date]["net"] += net
                except Exception as exc:
                    rejected += 1
                    if len(errors) < 100:
                        errors.append(f"Row {index}: {exc}")
            for day, values in daily.items():
                conn.execute(
                    """INSERT INTO sales(period_start,period_end,gross_sales,discounts,refunds,sales_tax,net_sales,source_file)
                       VALUES(?,?,?,?,?,?,?,?)
                       ON CONFLICT(period_start,period_end,source_file) DO UPDATE SET gross_sales=excluded.gross_sales,
                       discounts=excluded.discounts,refunds=excluded.refunds,sales_tax=excluded.sales_tax,net_sales=excluded.net_sales""",
                    (
                        day, day, f"{money(values['gross']):.2f}", f"{money(values['discounts']):.2f}",
                        f"{money(values['refunds']):.2f}", f"{money(values['tax']):.2f}",
                        f"{money(values['net']):.2f}", f"POS:{path.name}",
                    ),
                )
            conn.execute(
                """UPDATE pos_import_runs SET row_count=?,rejected_count=?,gross_sales=?,net_sales=?,status=?,errors_json=?
                   WHERE run_id=?""",
                (imported, rejected, f"{money(gross_total):.2f}", f"{money(net_total):.2f}",
                 "Imported" if imported else "Failed", json.dumps(errors), run_id),
            )
        archive = Path(self.workspace.folders["pos"]) / safe_name(path.name)
        if archive.resolve() != path.resolve():
            archive.write_bytes(path.read_bytes())
        return POSImportResult(run_id, imported, rejected, money(gross_total), money(net_total), errors, mapping)

    def list_pos_runs(self, limit: int = 200) -> list[sqlite3.Row]:
        with self.workspace.connect() as conn:
            return conn.execute("SELECT * FROM pos_import_runs ORDER BY imported_at DESC LIMIT ?", (limit,)).fetchall()

    # ------------------------------------------------------------------
    # Recipes and menu costing
    # ------------------------------------------------------------------
    def export_recipe_template(self, destination: Path | None = None) -> Path:
        destination = destination or Path(self.workspace.folders["recipes"]) / "Recipe_Import_Template.csv"
        headers = [
            "Menu Item ID", "POS Item Key", "Menu Item Name", "Menu Category", "Menu Price",
            "Inventory Item ID", "Vendor SKU", "Inventory Item Name", "Quantity Count Units",
            "Yield Percent", "Notes",
        ]
        with destination.open("w", encoding="utf-8", newline="") as fh:
            writer = csv.writer(fh)
            writer.writerow(headers)
            writer.writerow(["", "BURGER-001", "Classic Burger", "Burgers", "9.49", "", "BEEF-40", "Ground Beef", "0.33", "100", ""])
        return destination

    def _match_inventory_item(self, conn: sqlite3.Connection, row: dict[str, Any]) -> str:
        """Resolve a recipe ingredient using ID, SKU, or normalized name.

        Recipe workbooks commonly contain an Inventory Item ID. That identifier
        is authoritative and must be preferred over fuzzy name matching.
        """
        def value(*names: str) -> str:
            normalized = {normalize(k): v for k, v in row.items()}
            for name in names:
                raw = normalized.get(normalize(name))
                if raw is not None and str(raw).strip():
                    return str(raw).strip()
            return ""

        item_id = value("Inventory Item ID", "item_id", "Item ID")
        if item_id:
            found = conn.execute("SELECT item_id FROM items WHERE item_id=? COLLATE NOCASE", (item_id,)).fetchone()
            if found:
                return found["item_id"]
            # Many recipe workbooks label the distributor SKU as
            # "Inventory Item ID". Treat an ID-like value as a SKU fallback
            # before attempting fuzzy name matching.
            found = conn.execute(
                "SELECT item_id FROM items WHERE vendor_sku=? COLLATE NOCASE ORDER BY last_purchase_date DESC LIMIT 1",
                (item_id,),
            ).fetchone()
            if found:
                return found["item_id"]

        sku = value("Vendor SKU", "vendor_sku", "SKU")
        if sku:
            found = conn.execute(
                "SELECT item_id FROM items WHERE vendor_sku=? COLLATE NOCASE ORDER BY last_purchase_date DESC LIMIT 1",
                (sku,),
            ).fetchone()
            if found:
                return found["item_id"]

        name = normalize(value("Inventory Item Name", "ingredient", "Ingredient Name", "item name", "Item Name"))
        if name:
            found = conn.execute(
                "SELECT item_id FROM items WHERE normalized_description=? ORDER BY last_purchase_date DESC LIMIT 1",
                (name,),
            ).fetchone()
            if found:
                return found["item_id"]
            # Token-safe partial fallback for vendor descriptors such as
            # 'Hamburger buns - 12 pack' versus 'Hamburger buns'.
            found = conn.execute(
                "SELECT item_id FROM items WHERE normalized_description LIKE ? OR LOWER(item_name) LIKE ? "
                "ORDER BY last_purchase_date DESC LIMIT 1",
                (f"%{name}%", f"%{name}%"),
            ).fetchone()
            if found:
                return found["item_id"]

        raise Phase2Error(
            f"Inventory ingredient could not be matched: {item_id or sku or name or 'blank ingredient'}"
        )

    def import_recipes_csv(self, path: Path) -> dict[str, Any]:
        """Import CSV or Excel recipe workbooks using the same recipe schema."""
        imported = skipped = 0
        errors: list[str] = []
        stamp = now_iso()
        from excel_io import is_excel_path, read_xlsx
        source_path = Path(path)
        if is_excel_path(source_path):
            rows = read_xlsx(source_path)
            row_iter = enumerate(rows, 2)
        else:
            with source_path.open("r", encoding="utf-8-sig", newline="") as fh:
                rows = list(csv.DictReader(fh))
            row_iter = enumerate(rows, 2)
        with self.workspace.connect() as conn:
            for index, row in row_iter:
                try:
                    name = str(row.get("Menu Item Name") or row.get("menu_item_name") or "").strip()
                    if not name:
                        raise Phase2Error("Menu Item Name is required")
                    key = str(row.get("POS Item Key") or row.get("pos_item_key") or normalize(name)).strip()
                    menu_item_id = str(row.get("Menu Item ID") or row.get("menu_item_id") or "").strip()
                    if not menu_item_id:
                        found = conn.execute("SELECT menu_item_id FROM menu_items WHERE pos_item_key=? COLLATE NOCASE", (key,)).fetchone()
                        menu_item_id = found["menu_item_id"] if found else f"MENU-{hashlib.sha256(key.encode()).hexdigest()[:14].upper()}"
                    menu_price = money(row.get("Menu Price") or row.get("menu_price"))
                    conn.execute(
                        """INSERT INTO menu_items(menu_item_id,pos_item_key,menu_item_name,category,menu_price,created_at,updated_at)
                           VALUES(?,?,?,?,?,?,?) ON CONFLICT(menu_item_id) DO UPDATE SET pos_item_key=excluded.pos_item_key,
                           menu_item_name=excluded.menu_item_name,category=excluded.category,menu_price=excluded.menu_price,updated_at=excluded.updated_at""",
                        (menu_item_id, key, name, str(row.get("Menu Category") or "Unclassified"), f"{menu_price:.2f}", stamp, stamp),
                    )
                    item_id = self._match_inventory_item(conn, row)
                    amount = qty(row.get("Quantity Count Units") or row.get("quantity_count_units"))
                    if amount <= 0:
                        raise Phase2Error("Quantity Count Units must be positive")
                    yield_percent = dec(row.get("Yield Percent") or row.get("yield_percent") or "100")
                    if yield_percent <= 0 or yield_percent > 100:
                        raise Phase2Error("Yield Percent must be between 0 and 100")
                    conn.execute(
                        """INSERT INTO recipe_ingredients(menu_item_id,item_id,quantity_count_units,yield_percent,notes,created_at,updated_at)
                           VALUES(?,?,?,?,?,?,?) ON CONFLICT(menu_item_id,item_id) DO UPDATE SET
                           quantity_count_units=excluded.quantity_count_units,yield_percent=excluded.yield_percent,
                           notes=excluded.notes,updated_at=excluded.updated_at""",
                        (menu_item_id, item_id, f"{amount:.4f}", f"{yield_percent:.2f}", str(row.get("Notes") or ""), stamp, stamp),
                    )
                    imported += 1
                except Exception as exc:
                    skipped += 1
                    errors.append(f"Row {index}: {exc}")
        target = Path(self.workspace.folders["recipes"]) / safe_name(Path(path).name)
        if target.resolve() != Path(path).resolve():
            target.write_bytes(Path(path).read_bytes())
        return {"imported": imported, "skipped": skipped, "errors": errors}

    def list_menu_costs(self, start: str | None = None, end: str | None = None) -> list[dict[str, Any]]:
        start = start or f"{date.today().year}-01-01"
        end = end or date.today().isoformat()
        with self.workspace.connect() as conn:
            menus = conn.execute("SELECT * FROM menu_items WHERE active=1 ORDER BY category,menu_item_name").fetchall()
            output: list[dict[str, Any]] = []
            for menu in menus:
                ingredients = conn.execute(
                    """SELECT r.*,i.item_name,i.current_price,i.units_per_purchase_unit,i.count_unit,i.unit
                       FROM recipe_ingredients r JOIN items i ON i.item_id=r.item_id WHERE r.menu_item_id=?""",
                    (menu["menu_item_id"],),
                ).fetchall()
                recipe_cost = Decimal("0")
                ingredient_count = 0
                for ingredient in ingredients:
                    units_per = dec(ingredient["units_per_purchase_unit"], "1") or Decimal("1")
                    unit_cost = money(dec(ingredient["current_price"]) / units_per)
                    effective_qty = dec(ingredient["quantity_count_units"]) / (dec(ingredient["yield_percent"], "100") / Decimal("100"))
                    recipe_cost += money(unit_cost * effective_qty)
                    ingredient_count += 1
                sold = conn.execute(
                    """SELECT COALESCE(SUM(CAST(quantity AS REAL)),0) AS qty,
                       COALESCE(SUM(CAST(net_sales AS REAL)),0) AS sales
                       FROM pos_sales_lines WHERE menu_item_id=? AND business_date BETWEEN ? AND ?""",
                    (menu["menu_item_id"], start, end),
                ).fetchone()
                menu_price = money(menu["menu_price"])
                food_cost_pct = (recipe_cost / menu_price * Decimal("100")) if menu_price else Decimal("0")
                output.append({
                    **dict(menu),
                    "ingredient_count": ingredient_count,
                    "recipe_cost": f"{money(recipe_cost):.2f}",
                    "food_cost_percent": f"{food_cost_pct.quantize(Decimal('0.01')):.2f}",
                    "contribution_margin": f"{money(menu_price - recipe_cost):.2f}",
                    "quantity_sold": f"{qty(sold['qty']):.4f}",
                    "net_sales": f"{money(sold['sales']):.2f}",
                    "theoretical_food_cost": f"{money(recipe_cost * dec(sold['qty'])):.2f}",
                })
        return output

    def recipe_variance(self, month: str) -> list[dict[str, Any]]:
        start = date.fromisoformat(month + "-01")
        if start.month == 12:
            end = date(start.year, 12, 31)
        else:
            end = date(start.year, start.month + 1, 1) - timedelta(days=1)
        # Inventory usage is available as a read-only preview as soon as both
        # physical counts exist.  Do not make variance reporting depend on the
        # manager closing the month first.
        usage_rows = self.planning.list_month_usage(month)
        actual = {
            row["item_id"]: dec(row["estimated_usage_quantity"])
            for row in usage_rows
            if row["estimated_usage_quantity"] not in (None, "")
        }
        with self.workspace.connect() as conn:
            theoretical: dict[str, Decimal] = defaultdict(Decimal)
            rows = conn.execute(
                """SELECT s.quantity,r.item_id,r.quantity_count_units,r.yield_percent
                   FROM pos_sales_lines s JOIN recipe_ingredients r ON r.menu_item_id=s.menu_item_id
                   WHERE s.business_date BETWEEN ? AND ?""",
                (start.isoformat(), end.isoformat()),
            ).fetchall()
            for row in rows:
                theoretical[row["item_id"]] += dec(row["quantity"]) * dec(row["quantity_count_units"]) / (dec(row["yield_percent"], "100") / Decimal("100"))
            waste: dict[str, Decimal] = defaultdict(Decimal)
            for row in conn.execute("SELECT item_id,SUM(CAST(quantity_count_units AS REAL)) AS qty FROM waste_events WHERE event_date BETWEEN ? AND ? GROUP BY item_id", (start.isoformat(), end.isoformat())):
                waste[row["item_id"]] = dec(row["qty"])
            items = conn.execute("SELECT item_id,item_name,count_unit FROM items WHERE active=1 ORDER BY item_name").fetchall()
            output = []
            for item in items:
                theory = theoretical.get(item["item_id"], Decimal("0"))
                logged_waste = waste.get(item["item_id"], Decimal("0"))
                expected = theory + logged_waste
                actual_qty = actual.get(item["item_id"], Decimal("0"))
                variance = actual_qty - expected
                output.append({
                    "item_id": item["item_id"], "item_name": item["item_name"], "count_unit": item["count_unit"],
                    "theoretical_usage": f"{qty(theory):.4f}", "logged_waste": f"{qty(logged_waste):.4f}",
                    "expected_depletion": f"{qty(expected):.4f}", "actual_depletion": f"{qty(actual_qty):.4f}",
                    "unexplained_variance": f"{qty(variance):.4f}",
                })
            return output

    # ------------------------------------------------------------------
    # Waste logging
    # ------------------------------------------------------------------
    def log_waste(self, item_id: str, quantity: Any, reason: str, *, event_date: str | None = None,
                  shift: str = "", notes: str = "", photo_path: str = "", created_by: str = "system") -> str:
        amount = qty(quantity)
        if amount <= 0:
            raise Phase2Error("Waste quantity must be positive.")
        reason = str(reason or "").strip()
        if not reason:
            raise Phase2Error("Waste reason is required.")
        event_date = parse_date(event_date or date.today().isoformat())
        with self.workspace.connect() as conn:
            item = conn.execute("SELECT current_price,units_per_purchase_unit FROM items WHERE item_id=?", (item_id,)).fetchone()
            if not item:
                raise Phase2Error("Inventory item not found.")
            unit_cost = dec(item["current_price"]) / (dec(item["units_per_purchase_unit"], "1") or Decimal("1"))
            cost = money(amount * unit_cost)
            waste_id = f"WASTE-{uuid.uuid4().hex[:12].upper()}"
            conn.execute(
                """INSERT INTO waste_events(waste_id,event_date,item_id,quantity_count_units,reason,shift,estimated_cost,
                   notes,photo_path,created_by,created_at) VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                (waste_id, event_date, item_id, f"{amount:.4f}", reason, shift, f"{cost:.2f}", notes, photo_path, created_by, now_iso()),
            )
        return waste_id

    def list_waste(self, start: str | None = None, end: str | None = None, limit: int = 1000) -> list[sqlite3.Row]:
        start = start or f"{date.today().year}-01-01"
        end = end or date.today().isoformat()
        with self.workspace.connect() as conn:
            return conn.execute(
                """SELECT w.*,i.item_name,i.vendor_name,i.count_unit FROM waste_events w JOIN items i ON i.item_id=w.item_id
                   WHERE w.event_date BETWEEN ? AND ? ORDER BY w.event_date DESC,w.created_at DESC LIMIT ?""",
                (start, end, limit),
            ).fetchall()

    # ------------------------------------------------------------------
    # Mobile inventory counts
    # ------------------------------------------------------------------
    def create_mobile_count_session(self, count_date: str | None = None, *, created_by: str = "system", hours: int = 12) -> dict[str, str]:
        count_date = parse_date(count_date or date.today().isoformat())
        session_id = f"MCOUNT-{datetime.now().strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:5].upper()}"
        token = secrets.token_urlsafe(18)
        token_hash = hashlib.sha256(token.encode()).hexdigest()
        expires = (datetime.now() + timedelta(hours=max(1, int(hours)))).replace(microsecond=0).isoformat()
        with self.workspace.connect() as conn:
            conn.execute(
                """INSERT INTO mobile_count_sessions(session_id,count_date,count_month,status,token_hash,expires_at,
                   created_by,created_at) VALUES(?,?,?,'Open',?,?,?,?)""",
                (session_id, count_date, count_date[:7], token_hash, expires, created_by, now_iso()),
            )
        return {"session_id": session_id, "token": token, "expires_at": expires}

    def _mobile_items(self) -> list[dict[str, Any]]:
        try:
            estimates = {row["item_id"]: row for row in self.planning.estimate_inventory()}
        except Exception:
            estimates = {}
        with self.workspace.connect() as conn:
            rows = conn.execute("SELECT item_id,item_name,category,vendor_name,count_unit,unit FROM items WHERE active=1 ORDER BY category,item_name").fetchall()
        return [{**dict(row), "estimate": estimates.get(row["item_id"], {}).get("estimated_on_hand", "")} for row in rows]

    def save_mobile_entries(self, session_id: str, entries: dict[str, Any], *, submitted: bool = False) -> int:
        stamp = now_iso()
        count = 0
        with self.workspace.connect() as conn:
            session = conn.execute("SELECT * FROM mobile_count_sessions WHERE session_id=?", (session_id,)).fetchone()
            if not session:
                raise Phase2Error("Mobile count session not found.")
            if session["status"] == "Finalized":
                raise Phase2Error("Mobile count is already finalized.")
            for item_id, raw in entries.items():
                if raw in (None, ""):
                    continue
                amount = qty(raw)
                if amount < 0:
                    raise Phase2Error("Counts cannot be negative.")
                conn.execute(
                    """INSERT INTO mobile_count_entries(session_id,item_id,quantity_on_hand,updated_at)
                       VALUES(?,?,?,?) ON CONFLICT(session_id,item_id) DO UPDATE SET
                       quantity_on_hand=excluded.quantity_on_hand,updated_at=excluded.updated_at""",
                    (session_id, item_id, f"{amount:.4f}", stamp),
                )
                count += 1
            if submitted:
                conn.execute("UPDATE mobile_count_sessions SET status='Submitted',submitted_at=? WHERE session_id=?", (stamp, session_id))
        return count

    def list_mobile_sessions(self, limit: int = 100) -> list[sqlite3.Row]:
        with self.workspace.connect() as conn:
            return conn.execute(
                """SELECT s.*,COUNT(e.entry_id) AS entry_count FROM mobile_count_sessions s
                   LEFT JOIN mobile_count_entries e ON e.session_id=s.session_id GROUP BY s.session_id
                   ORDER BY s.created_at DESC LIMIT ?""", (limit,)
            ).fetchall()

    def get_mobile_entries(self, session_id: str) -> list[sqlite3.Row]:
        with self.workspace.connect() as conn:
            return conn.execute(
                """SELECT e.*,i.item_name,i.vendor_name,i.category,i.count_unit,i.current_price,i.units_per_purchase_unit
                   FROM mobile_count_entries e JOIN items i ON i.item_id=e.item_id WHERE e.session_id=? ORDER BY i.category,i.item_name""",
                (session_id,),
            ).fetchall()

    def finalize_mobile_count(self, session_id: str, *, source_file: str | None = None) -> dict[str, Any]:
        stamp = now_iso()
        imported = 0
        with self.workspace.connect() as conn:
            session = conn.execute("SELECT * FROM mobile_count_sessions WHERE session_id=?", (session_id,)).fetchone()
            if not session:
                raise Phase2Error("Mobile count session not found.")
            if session["status"] == "Finalized":
                raise Phase2Error("Mobile count is already finalized.")
            entries = conn.execute(
                """SELECT e.*,i.current_price,i.units_per_purchase_unit,i.count_unit,i.unit
                   FROM mobile_count_entries e JOIN items i ON i.item_id=e.item_id WHERE e.session_id=?""", (session_id,)
            ).fetchall()
            if not entries:
                raise Phase2Error("No mobile count entries have been submitted.")
            source = source_file or f"MobileCount_{session_id}"
            for row in entries:
                amount = qty(row["quantity_on_hand"])
                units_per = dec(row["units_per_purchase_unit"], "1") or Decimal("1")
                unit_cost = money(dec(row["current_price"]) / units_per)
                value = money(amount * unit_cost)
                conn.execute(
                    """INSERT INTO inventory_counts(count_date,count_month,item_id,quantity_on_hand,count_unit,unit_cost,
                       inventory_value,source_file,notes,finalized,created_at) VALUES(?,?,?,?,?,?,?,?,?,1,?)
                       ON CONFLICT(count_date,item_id,source_file) DO UPDATE SET quantity_on_hand=excluded.quantity_on_hand,
                       count_unit=excluded.count_unit,unit_cost=excluded.unit_cost,inventory_value=excluded.inventory_value,
                       notes=excluded.notes,finalized=1""",
                    (session["count_date"], session["count_month"], row["item_id"], f"{amount:.4f}",
                     row["count_unit"] or row["unit"] or "each", f"{unit_cost:.2f}", f"{value:.2f}", source,
                     "Submitted through mobile count", stamp),
                )
                imported += 1
            conn.execute("UPDATE mobile_count_sessions SET status='Finalized',finalized_at=? WHERE session_id=?", (stamp, session_id))
        return {"session_id": session_id, "imported": imported, "count_date": session["count_date"]}

    def _local_ip(self) -> str:
        try:
            with socket.socket(socket.AF_INET, socket.SOCK_DGRAM) as s:
                s.connect(("8.8.8.8", 80))
                return s.getsockname()[0]
        except OSError:
            return "127.0.0.1"

    def start_mobile_count_server(self, session_id: str, token: str, port: int = 0) -> MobileServerHandle:
        if self.mobile_handle:
            self.mobile_handle.stop()
            self.mobile_handle = None
        expected_hash = hashlib.sha256(token.encode()).hexdigest()
        with self.workspace.connect() as conn:
            session = conn.execute("SELECT * FROM mobile_count_sessions WHERE session_id=?", (session_id,)).fetchone()
        if not session or session["token_hash"] != expected_hash:
            raise Phase2Error("Invalid mobile count session token.")
        try:
            expires_at = datetime.fromisoformat(str(session["expires_at"]))
        except (TypeError, ValueError) as exc:
            raise Phase2Error("Mobile count session has an invalid expiration time.") from exc
        if datetime.now() > expires_at:
            raise Phase2Error("Mobile count session has expired. Create a new mobile count session.")
        service = self
        items = self._mobile_items()

        class Handler(BaseHTTPRequestHandler):
            def log_message(self, _format: str, *_args: Any) -> None:
                return

            def _authorized(self) -> bool:
                if datetime.now() > expires_at:
                    return False
                query = parse_qs(urlparse(self.path).query)
                supplied = query.get("token", [""])[0]
                return secrets.compare_digest(hashlib.sha256(supplied.encode()).hexdigest(), expected_hash)

            def _send(self, status: int, body: str) -> None:
                payload = body.encode("utf-8")
                self.send_response(status)
                self.send_header("Content-Type", "text/html; charset=utf-8")
                self.send_header("Content-Length", str(len(payload)))
                self.send_header("Cache-Control", "no-store")
                self.end_headers()
                self.wfile.write(payload)

            def do_GET(self) -> None:  # noqa: N802
                if not self._authorized():
                    self._send(403, "<h1>Access denied</h1>")
                    return
                rows = []
                current_category = None
                for item in items:
                    if item["category"] != current_category:
                        current_category = item["category"]
                        rows.append(f"<h2>{html.escape(current_category or 'Unclassified')}</h2>")
                    item_id = html.escape(item["item_id"])
                    estimate = html.escape(str(item.get("estimate") or ""))
                    rows.append(
                        f"<label><span><b>{html.escape(item['item_name'])}</b><small>{html.escape(item['vendor_name'])} · {html.escape(item['count_unit'] or item['unit'] or 'each')} · est. {estimate}</small></span>"
                        f"<input inputmode='decimal' type='number' min='0' step='0.0001' name='item_{item_id}'></label>"
                    )
                page = f"""<!doctype html><html><head><meta name='viewport' content='width=device-width,initial-scale=1'>
                <title>Mobile Inventory Count</title><style>
                body{{font-family:system-ui;margin:0;background:#f5f7fa;color:#17324d}}header{{position:sticky;top:0;background:#17324d;color:white;padding:14px}}
                form{{padding:12px;max-width:760px;margin:auto}}h2{{margin-top:22px;border-bottom:1px solid #ccd5df;padding-bottom:6px}}
                label{{display:flex;align-items:center;gap:10px;background:white;padding:10px;margin:7px 0;border-radius:8px;box-shadow:0 1px 3px #0002}}
                label span{{flex:1}}small{{display:block;color:#667085;margin-top:3px}}input{{width:110px;font-size:18px;padding:9px}}
                button{{position:sticky;bottom:10px;width:100%;font-size:18px;padding:14px;background:#1f6f78;color:white;border:0;border-radius:8px}}
                </style></head><body><header><b>Inventory Count</b><br><small>{html.escape(session['count_date'])}</small></header>
                <form method='post' action='/?token={html.escape(token)}'>{''.join(rows)}<button type='submit'>Submit count for manager review</button></form></body></html>"""
                self._send(200, page)

            def do_POST(self) -> None:  # noqa: N802
                if not self._authorized():
                    self._send(403, "<h1>Access denied</h1>")
                    return
                length = int(self.headers.get("Content-Length", "0"))
                values = parse_qs(self.rfile.read(length).decode("utf-8"), keep_blank_values=True)
                entries = {key[5:]: vals[0] for key, vals in values.items() if key.startswith("item_") and vals and vals[0].strip()}
                try:
                    count = service.save_mobile_entries(session_id, entries, submitted=True)
                    self._send(200, f"<h1>Count submitted</h1><p>{count} item(s) saved for manager review.</p><p>You may close this page.</p>")
                except Exception as exc:
                    self._send(400, f"<h1>Could not save count</h1><p>{html.escape(str(exc))}</p>")

        server = ThreadingHTTPServer(("0.0.0.0", int(port)), Handler)
        actual_port = server.server_address[1]
        url = f"http://{self._local_ip()}:{actual_port}/?token={token}"
        thread = threading.Thread(target=server.serve_forever, daemon=True)
        thread.start()
        self.mobile_handle = MobileServerHandle(server, thread, url, session_id)
        return self.mobile_handle

    def stop_mobile_count_server(self) -> None:
        if self.mobile_handle:
            self.mobile_handle.stop()
            self.mobile_handle = None

    # ------------------------------------------------------------------
    # Vendor-ready purchase orders
    # ------------------------------------------------------------------
    def list_vendor_profiles(self) -> list[sqlite3.Row]:
        with self.workspace.connect() as conn:
            return conn.execute("SELECT * FROM vendor_order_profiles ORDER BY vendor_name").fetchall()

    def save_vendor_profile(self, vendor_name: str, **values: Any) -> None:
        vendor_name = str(vendor_name or "").strip()
        if not vendor_name:
            raise Phase2Error("Vendor name is required.")
        with self.workspace.connect() as conn:
            conn.execute(
                """INSERT INTO vendor_order_profiles(vendor_name,vendor_email,account_number,delivery_days,payment_terms,po_notes,updated_at)
                   VALUES(?,?,?,?,?,?,?) ON CONFLICT(vendor_name) DO UPDATE SET vendor_email=excluded.vendor_email,
                   account_number=excluded.account_number,delivery_days=excluded.delivery_days,payment_terms=excluded.payment_terms,
                   po_notes=excluded.po_notes,updated_at=excluded.updated_at""",
                (vendor_name, values.get("vendor_email", ""), values.get("account_number", ""),
                 values.get("delivery_days", ""), values.get("payment_terms", ""), values.get("po_notes", ""), now_iso()),
            )

    def generate_purchase_orders(self, batch_id: str | None = None, *, created_by: str = "system") -> dict[str, Any]:
        with self.workspace.connect() as conn:
            if not batch_id:
                batch = conn.execute("SELECT * FROM order_batches ORDER BY created_at DESC LIMIT 1").fetchone()
                if not batch:
                    raise Phase2Error("Generate an order prediction batch first.")
                batch_id = batch["batch_id"]
            predictions = conn.execute(
                """SELECT * FROM order_predictions WHERE batch_id=? AND
                   CAST(COALESCE(NULLIF(manager_order_quantity,''),suggested_order_quantity,'0') AS REAL)>0
                   ORDER BY vendor_name,item_name""", (batch_id,)
            ).fetchall()
            if not predictions:
                raise Phase2Error("The selected order batch contains no positive quantities.")
            grouped: dict[str, list[sqlite3.Row]] = defaultdict(list)
            for row in predictions:
                grouped[row["vendor_name"] or "Unassigned Vendor"].append(row)
            po_ids = []
            for vendor, rows in grouped.items():
                po_id = f"PO-{datetime.now().strftime('%Y%m%d')}-{hashlib.sha256((batch_id+vendor).encode()).hexdigest()[:7].upper()}"
                subtotal = sum((money(dec(row["manager_order_quantity"] or row["suggested_order_quantity"]) * dec(row["current_price"])) for row in rows), Decimal("0"))
                conn.execute(
                    """INSERT INTO purchase_orders(po_id,batch_id,vendor_name,po_date,status,subtotal,created_by,created_at)
                       VALUES(?,?,?,?,'Draft',?,?,?) ON CONFLICT(batch_id,vendor_name) DO UPDATE SET subtotal=excluded.subtotal,
                       status=CASE WHEN purchase_orders.status='Approved' THEN 'Approved' ELSE 'Draft' END""",
                    (po_id, batch_id, vendor, date.today().isoformat(), f"{money(subtotal):.2f}", created_by, now_iso()),
                )
                actual = conn.execute("SELECT po_id FROM purchase_orders WHERE batch_id=? AND vendor_name=?", (batch_id, vendor)).fetchone()["po_id"]
                conn.execute("DELETE FROM purchase_order_lines WHERE po_id=?", (actual,))
                for row in rows:
                    order_qty = qty(row["manager_order_quantity"] or row["suggested_order_quantity"])
                    unit_price = money(row["current_price"])
                    conn.execute(
                        """INSERT INTO purchase_order_lines(po_id,prediction_id,item_id,vendor_sku,item_name,quantity,
                           purchase_unit,unit_price,line_total,notes) VALUES(?,?,?,?,?,?,?,?,?,?)""",
                        (actual, row["prediction_id"], row["item_id"], row["vendor_sku"], row["item_name"],
                         f"{order_qty:.4f}", row["purchase_unit"], f"{unit_price:.2f}",
                         f"{money(order_qty * unit_price):.2f}", row["notes"] or ""),
                    )
                po_ids.append(actual)
            return {"batch_id": batch_id, "purchase_orders": po_ids, "vendor_count": len(po_ids)}

    def list_purchase_orders(self, limit: int = 300) -> list[sqlite3.Row]:
        with self.workspace.connect() as conn:
            return conn.execute(
                """SELECT p.*,COUNT(l.po_line_id) AS line_count FROM purchase_orders p LEFT JOIN purchase_order_lines l ON l.po_id=p.po_id
                   GROUP BY p.po_id ORDER BY p.po_date DESC,p.vendor_name LIMIT ?""", (limit,)
            ).fetchall()

    def approve_purchase_order(self, po_id: str) -> None:
        with self.workspace.connect() as conn:
            conn.execute("UPDATE purchase_orders SET status='Approved',approved_at=? WHERE po_id=?", (now_iso(), po_id))

    def export_purchase_orders(self, po_ids: Iterable[str] | None = None, destination: Path | None = None) -> Path:
        destination = destination or Path(self.workspace.folders["purchase_orders"]) / f"Vendor_POs_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        destination.mkdir(parents=True, exist_ok=True)
        with self.workspace.connect() as conn:
            if po_ids:
                ids = list(po_ids)
                placeholders = ",".join("?" for _ in ids)
                pos = conn.execute(f"SELECT * FROM purchase_orders WHERE po_id IN ({placeholders}) ORDER BY vendor_name", ids).fetchall()
            else:
                pos = conn.execute("SELECT * FROM purchase_orders WHERE status IN ('Draft','Approved') ORDER BY po_date DESC,vendor_name").fetchall()
            summary_rows = []
            for po in pos:
                lines = conn.execute("SELECT * FROM purchase_order_lines WHERE po_id=? ORDER BY item_name", (po["po_id"],)).fetchall()
                profile = conn.execute("SELECT * FROM vendor_order_profiles WHERE vendor_name=? COLLATE NOCASE", (po["vendor_name"],)).fetchone()
                base = safe_name(f"{po['po_id']}_{po['vendor_name']}")
                csv_path = destination / f"{base}.csv"
                with csv_path.open("w", encoding="utf-8", newline="") as fh:
                    writer = csv.writer(fh)
                    writer.writerow(["Purchase Order", po["po_id"]])
                    writer.writerow(["Vendor", po["vendor_name"]])
                    writer.writerow(["Vendor Email", profile["vendor_email"] if profile else ""])
                    writer.writerow(["Account Number", profile["account_number"] if profile else ""])
                    writer.writerow(["PO Date", po["po_date"]])
                    writer.writerow(["Expected Delivery", po["expected_delivery_date"] or ""])
                    writer.writerow([])
                    writer.writerow(["Vendor SKU", "Item", "Quantity", "Purchase Unit", "Unit Price", "Line Total", "Notes"])
                    for line in lines:
                        writer.writerow([line["vendor_sku"], line["item_name"], line["quantity"], line["purchase_unit"], line["unit_price"], line["line_total"], line["notes"]])
                    writer.writerow([])
                    writer.writerow(["Subtotal", po["subtotal"]])
                html_path = destination / f"{base}.html"
                line_html = "".join(
                    f"<tr><td>{html.escape(str(line['vendor_sku'] or ''))}</td><td>{html.escape(line['item_name'])}</td><td>{line['quantity']}</td><td>{html.escape(str(line['purchase_unit'] or ''))}</td><td>${line['unit_price']}</td><td>${line['line_total']}</td></tr>"
                    for line in lines
                )
                html_path.write_text(
                    f"""<!doctype html><html><head><meta charset='utf-8'><title>{po['po_id']}</title><style>
                    body{{font-family:Arial;margin:32px;color:#17324d}}table{{border-collapse:collapse;width:100%}}th,td{{border:1px solid #ccd5df;padding:8px;text-align:left}}th{{background:#17324d;color:white}}
                    </style></head><body><h1>Purchase Order {po['po_id']}</h1><p><b>Vendor:</b> {html.escape(po['vendor_name'])}<br>
                    <b>Email:</b> {html.escape(profile['vendor_email'] if profile else '')}<br><b>Account:</b> {html.escape(profile['account_number'] if profile else '')}<br>
                    <b>PO Date:</b> {po['po_date']}<br><b>Status:</b> {po['status']}</p><table><tr><th>SKU</th><th>Item</th><th>Qty</th><th>Unit</th><th>Price</th><th>Total</th></tr>{line_html}</table>
                    <h2>Subtotal: ${po['subtotal']}</h2><p>{html.escape(profile['po_notes'] if profile else '')}</p></body></html>""",
                    encoding="utf-8",
                )
                summary_rows.append([po["po_id"], po["vendor_name"], po["po_date"], po["status"], po["subtotal"], csv_path.name, html_path.name])
            with (destination / "PO_Summary.csv").open("w", encoding="utf-8", newline="") as fh:
                writer = csv.writer(fh)
                writer.writerow(["PO ID", "Vendor", "PO Date", "Status", "Subtotal", "CSV", "Printable HTML"])
                writer.writerows(summary_rows)
        return destination

    # ------------------------------------------------------------------
    # Accounting exports
    # ------------------------------------------------------------------
    def accounting_mappings(self) -> dict[str, tuple[str, str]]:
        with self.workspace.connect() as conn:
            return {row["mapping_key"]: (row["debit_account"], row["credit_account"]) for row in conn.execute("SELECT * FROM accounting_mappings")}

    def set_accounting_mapping(self, key: str, debit: str, credit: str) -> None:
        with self.workspace.connect() as conn:
            conn.execute(
                """INSERT INTO accounting_mappings(mapping_key,debit_account,credit_account,updated_at) VALUES(?,?,?,?)
                   ON CONFLICT(mapping_key) DO UPDATE SET debit_account=excluded.debit_account,credit_account=excluded.credit_account,updated_at=excluded.updated_at""",
                (key, debit, credit, now_iso()),
            )

    def _journal_rows(self, start: str, end: str) -> list[dict[str, Any]]:
        mappings = self.accounting_mappings()
        rows: list[dict[str, Any]] = []
        with self.workspace.connect() as conn:
            for invoice in conn.execute("SELECT * FROM invoices WHERE status='Approved' AND invoice_date BETWEEN ? AND ? ORDER BY invoice_date", (start, end)):
                debit, credit = mappings["inventory_purchase"]
                amount = money(invoice["total"])
                memo = f"Invoice {invoice['invoice_number']} - {invoice['vendor']}"
                rows.extend([
                    {"date": invoice["invoice_date"], "reference": invoice["invoice_number"], "vendor": invoice["vendor"], "memo": memo, "account": debit, "debit": amount, "credit": Decimal("0"), "source_type": "invoice", "source_id": invoice["invoice_id"]},
                    {"date": invoice["invoice_date"], "reference": invoice["invoice_number"], "vendor": invoice["vendor"], "memo": memo, "account": credit, "debit": Decimal("0"), "credit": amount, "source_type": "invoice", "source_id": invoice["invoice_id"]},
                ])
            for sale in preferred_sales_rows(conn, start, end):
                gross = money(sale["gross_sales"]); discounts = money(sale["discounts"]); refunds = money(sale["refunds"]); tax = money(sale["sales_tax"]); net = money(sale["net_sales"])
                if gross == 0 and net:
                    gross = money(net + discounts + refunds)
                cash_account, sales_account = mappings["sales"]
                memo = f"Sales {sale['period_start']} to {sale['period_end']}"
                cash = money(net + tax)
                rows.append(
                    {"date": sale["period_end"], "reference": f"SALES-{sale['sales_id']}", "vendor": "", "memo": memo, "account": cash_account, "debit": cash, "credit": Decimal("0"), "source_type": "sales", "source_id": sale["sales_id"]}
                )
                rows.append(
                    {"date": sale["period_end"], "reference": f"SALES-{sale['sales_id']}", "vendor": "", "memo": memo, "account": sales_account, "debit": Decimal("0"), "credit": gross, "source_type": "sales", "source_id": sale["sales_id"]}
                )
                if discounts:
                    rows.append(
                        {"date": sale["period_end"], "reference": f"SALES-{sale['sales_id']}", "vendor": "", "memo": memo, "account": mappings["discounts"][0], "debit": discounts, "credit": Decimal("0"), "source_type": "sales", "source_id": sale["sales_id"]}
                    )
                if refunds:
                    rows.append(
                        {"date": sale["period_end"], "reference": f"SALES-{sale['sales_id']}", "vendor": "", "memo": memo, "account": mappings["refunds"][0], "debit": refunds, "credit": Decimal("0"), "source_type": "sales", "source_id": sale["sales_id"]}
                    )
                if tax:
                    rows.append(
                        {"date": sale["period_end"], "reference": f"SALES-{sale['sales_id']}", "vendor": "", "memo": memo, "account": mappings["sales_tax"][1], "debit": Decimal("0"), "credit": tax, "source_type": "sales", "source_id": sale["sales_id"]}
                    )
            for cost in conn.execute("SELECT * FROM operating_costs WHERE cost_date BETWEEN ? AND ? ORDER BY cost_date", (start, end)):
                debit, credit = mappings.get(f"cost:{normalize(cost['category'])}", mappings["operating_cost"])
                amount = money(cost["amount"])
                memo = f"{cost['category']} - {cost['description']}"
                rows.extend([
                    {"date": cost["cost_date"], "reference": f"COST-{cost['cost_id']}", "vendor": "", "memo": memo, "account": debit, "debit": amount, "credit": Decimal("0"), "source_type": "operating_cost", "source_id": cost["cost_id"]},
                    {"date": cost["cost_date"], "reference": f"COST-{cost['cost_id']}", "vendor": "", "memo": memo, "account": credit, "debit": Decimal("0"), "credit": amount, "source_type": "operating_cost", "source_id": cost["cost_id"]},
                ])
        return rows

    def export_accounting(self, start: str, end: str, export_type: str = "General Journal Excel") -> Path:
        export_type = str(export_type or "").strip()
        if export_type == "General Journal CSV":
            return self._export_general_journal_csv(start, end)
        if export_type == "General Journal Excel":
            return self._export_general_journal_xlsx(start, end)
        if export_type == "QuickBooks IIF":
            return self._export_quickbooks_iif(start, end)
        raise ValueError(f"Unsupported export type: {export_type}")

    def _export_general_journal_csv(self, start: str, end: str) -> Path:
        start = parse_date(start)
        end = parse_date(end)
        if end < start:
            raise Phase2Error("Accounting export end date cannot be before start date.")
        rows = self._journal_rows(start, end)
        total_debits = money(sum((row["debit"] for row in rows), Decimal("0")))
        total_credits = money(sum((row["credit"] for row in rows), Decimal("0")))
        if total_debits != total_credits:
            raise Phase2Error(f"Journal is not balanced: debits {total_debits}, credits {total_credits}")
        export_id = f"ACCT-{datetime.now().strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:5].upper()}"
        path = Path(self.workspace.folders["accounting"]) / safe_name(f"{export_id}_General Journal.csv")
        headers = ["Date", "Reference", "Vendor", "Memo", "Account", "Debit", "Credit", "Source Type", "Source ID"]
        with path.open("w", encoding="utf-8", newline="") as fh:
            writer = csv.writer(fh)
            writer.writerow(headers)
            for row in rows:
                writer.writerow([row["date"], row["reference"], row["vendor"], row["memo"], row["account"], f"{money(row['debit']):.2f}", f"{money(row['credit']):.2f}", row["source_type"], row["source_id"]])
        with self.workspace.connect() as conn:
            conn.execute(
                """INSERT INTO accounting_export_history(export_id,export_type,period_start,period_end,file_path,row_count,
                   total_debits,total_credits,created_by,created_at) VALUES(?,?,?,?,?,?,?,?,?,?)""",
                (export_id, "General Journal CSV", start, end, str(path), len(rows), f"{total_debits:.2f}", f"{total_credits:.2f}",
                 self.controls.current_user.username if self.controls.current_user else "system", now_iso()),
            )
        return path

    def _export_general_journal_xlsx(self, start: str, end: str) -> Path:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter

        start = parse_date(start)
        end = parse_date(end)
        if end < start:
            raise Phase2Error("Accounting export end date cannot be before start date.")
        rows = self._journal_rows(start, end)
        total_debits = money(sum((row["debit"] for row in rows), Decimal("0")))
        total_credits = money(sum((row["credit"] for row in rows), Decimal("0")))
        if total_debits != total_credits:
            raise Phase2Error(f"Journal is not balanced: debits {total_debits}, credits {total_credits}")
        export_id = f"ACCT-{datetime.now().strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:5].upper()}"
        path = Path(self.workspace.folders["accounting"]) / safe_name(f"{export_id}_General Journal.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "General Journal"
        headers = ["Date", "Reference", "Vendor", "Memo", "Account", "Debit", "Credit", "Source Type", "Source ID"]
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = PatternFill("solid", fgColor="17324D")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append([row["date"], row["reference"], row["vendor"], row["memo"], row["account"], f"{money(row['debit']):.2f}", f"{money(row['credit']):.2f}", row["source_type"], row["source_id"]])
        for column in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in column)
            ws.column_dimensions[get_column_letter(column[0].column)].width = min(max(max_len + 2, 12), 40)
        ws.freeze_panes = "A2"
        wb.save(str(path))
        with self.workspace.connect() as conn:
            conn.execute(
                """INSERT INTO accounting_export_history(export_id,export_type,period_start,period_end,file_path,row_count,
                   total_debits,total_credits,created_by,created_at) VALUES(?,?,?,?,?,?,?,?,?,?)""",
                (export_id, "General Journal Excel", start, end, str(path), len(rows), f"{total_debits:.2f}", f"{total_credits:.2f}",
                 self.controls.current_user.username if self.controls.current_user else "system", now_iso()),
            )
        return path

    def _export_quickbooks_iif(self, start: str, end: str) -> Path:
        start = parse_date(start)
        end = parse_date(end)
        if end < start:
            raise Phase2Error("Accounting export end date cannot be before start date.")
        rows = self._journal_rows(start, end)
        total_debits = money(sum((row["debit"] for row in rows), Decimal("0")))
        total_credits = money(sum((row["credit"] for row in rows), Decimal("0")))
        if total_debits != total_credits:
            raise Phase2Error(f"Journal is not balanced: debits {total_debits}, credits {total_credits}")
        export_id = f"ACCT-{datetime.now().strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:5].upper()}"
        path = Path(self.workspace.folders["accounting"]) / safe_name(f"{export_id}_QuickBooks IIF.iif")
        with path.open("w", encoding="utf-8", newline="") as fh:
            fh.write("!TRNS\tTRNSTYPE\tDATE\tACCNT\tNAME\tAMOUNT\tDOCNUM\tMEMO\n")
            fh.write("!SPL\tTRNSTYPE\tDATE\tACCNT\tNAME\tAMOUNT\tDOCNUM\tMEMO\n")
            fh.write("!ENDTRNS\n")
            grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
            for row in rows:
                grouped[f"{row['date']}|{row['reference']}"].append(row)
            for group in grouped.values():
                primary = group[0]
                amount = primary["debit"] - primary["credit"]
                fh.write(
                    f"TRNS\tGENERAL JOURNAL\t{primary['date']}\t{primary['account']}\t"
                    f"{primary['vendor']}\t{amount}\t{primary['reference']}\t{primary['memo']}\n"
                )
                for split in group[1:]:
                    split_amount = split["debit"] - split["credit"]
                    fh.write(
                        f"SPL\tGENERAL JOURNAL\t{split['date']}\t{split['account']}\t"
                        f"{split['vendor']}\t{split_amount}\t{split['reference']}\t{split['memo']}\n"
                    )
                fh.write("ENDTRNS\n")
        with self.workspace.connect() as conn:
            conn.execute(
                """INSERT INTO accounting_export_history(export_id,export_type,period_start,period_end,file_path,row_count,
                   total_debits,total_credits,created_by,created_at) VALUES(?,?,?,?,?,?,?,?,?,?)""",
                (
                    export_id,
                    "QuickBooks IIF",
                    start,
                    end,
                    str(path),
                    len(rows),
                    f"{total_debits:.2f}",
                    f"{total_credits:.2f}",
                    self.controls.current_user.username if self.controls.current_user else "system",
                    now_iso(),
                ),
            )
        return path

    def list_accounting_exports(self, limit: int = 200) -> list[sqlite3.Row]:
        with self.workspace.connect() as conn:
            return conn.execute("SELECT * FROM accounting_export_history ORDER BY created_at DESC LIMIT ?", (limit,)).fetchall()

    # ------------------------------------------------------------------
    # Dashboard and exports
    # ------------------------------------------------------------------
    def dashboard_summary(self) -> dict[str, Any]:
        month_start = date.today().replace(day=1).isoformat()
        with self.workspace.connect() as conn:
            pos_runs = conn.execute("SELECT COUNT(*) FROM pos_import_runs WHERE status='Imported'").fetchone()[0]
            menu_items = conn.execute("SELECT COUNT(*) FROM menu_items WHERE active=1").fetchone()[0]
            recipes = conn.execute("SELECT COUNT(DISTINCT menu_item_id) FROM recipe_ingredients").fetchone()[0]
            waste_cost = money(conn.execute("SELECT COALESCE(SUM(CAST(estimated_cost AS REAL)),0) FROM waste_events WHERE event_date>=?", (month_start,)).fetchone()[0])
            open_mobile = conn.execute("SELECT COUNT(*) FROM mobile_count_sessions WHERE status IN ('Open','Submitted')").fetchone()[0]
            draft_pos = conn.execute("SELECT COUNT(*) FROM purchase_orders WHERE status='Draft'").fetchone()[0]
            last_pos = conn.execute("SELECT MAX(imported_at) FROM pos_import_runs WHERE status='Imported'").fetchone()[0]
        return {
            "pos_import_runs": pos_runs, "menu_items": menu_items, "recipes_configured": recipes,
            "month_waste_cost": waste_cost, "open_mobile_counts": open_mobile,
            "draft_purchase_orders": draft_pos, "last_pos_import": last_pos or "",
        }

    def export_csvs(self) -> list[Path]:
        from excel_io import write_table_as
        queries = {
            "pos_import_runs.csv": "SELECT * FROM pos_import_runs ORDER BY imported_at DESC",
            "pos_sales_lines.csv": "SELECT * FROM pos_sales_lines ORDER BY business_date,sale_line_id",
            "menu_items.csv": "SELECT * FROM menu_items ORDER BY category,menu_item_name",
            "recipe_ingredients.csv": "SELECT r.*,m.menu_item_name,i.item_name FROM recipe_ingredients r JOIN menu_items m ON m.menu_item_id=r.menu_item_id JOIN items i ON i.item_id=r.item_id ORDER BY m.menu_item_name,i.item_name",
            "waste_events.csv": "SELECT w.*,i.item_name,i.vendor_name FROM waste_events w JOIN items i ON i.item_id=w.item_id ORDER BY event_date DESC",
            "mobile_count_sessions.csv": "SELECT * FROM mobile_count_sessions ORDER BY created_at DESC",
            "mobile_count_entries.csv": "SELECT e.*,i.item_name FROM mobile_count_entries e JOIN items i ON i.item_id=e.item_id ORDER BY session_id,i.item_name",
            "purchase_orders.csv": "SELECT * FROM purchase_orders ORDER BY po_date DESC,vendor_name",
            "purchase_order_lines.csv": "SELECT * FROM purchase_order_lines ORDER BY po_id,item_name",
            "accounting_export_history.csv": "SELECT * FROM accounting_export_history ORDER BY created_at DESC",
        }
        paths = []
        with self.workspace.connect() as conn:
            for filename, query in queries.items():
                path = Path(self.workspace.folders["exports"]) / filename
                records = [dict(row) for row in conn.execute(query).fetchall()]
                write_table_as(path, records, "csv")
                paths.append(path)
        return paths
