#!/usr/bin/env python3
"""Recipe costing engine for Restaurant Cost Controller v3.5.

Reads a spreadsheet describing each menu item's ingredient breakdown,
calculates per-menu-item food costs, and exports a costed recipe report.

Spreadsheet columns (CSV or XLSX), case-insensitive:
    Menu Item Name  (required)
    POS Item Key    (optional, for menu_items table lookup)
    Menu Category   (optional, defaults to "Unclassified")
    Menu Price      (required, the selling price of the menu item)
    Ingredient Name (required)
    Quantity        (required, numeric)
    Unit            (required, e.g. oz, lb, each, g, kg)
    Unit Cost       (optional — if missing, pulled from the item master)
    Yield Percent   (optional, defaults to 100)
    Notes           (optional)

If Unit Cost is blank the system looks up the ingredient in the
``items`` table.  Each inventory item carries:
    current_price              – price per purchase unit (e.g. $25 per case)
    units_per_purchase_unit    – count units per purchase unit (e.g. 160 oz in case)
    count_unit                 – the unit of measure (e.g. "oz")
    unit                       – purchase unit (e.g. "case")

The per-ingredient cost = Unit Cost / (count-unit conversion factor).
Unit conversions supported: oz↔lb, lb↔each (16 oz = 1 lb), g↔kg,
fl oz↔cup (8 fl oz = 1 cup), cup↔pt, pt↔qt, qt↔gal.

If the ingredient cannot be matched to an inventory item, the row is
flagged for manager review.
"""
from __future__ import annotations

import csv
import io
import re
import sqlite3
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable

from excel_io import is_excel_path, read_xlsx
from inventory_planning import QTY as QTY_SCALE, d as dec, normalize

MONEY = Decimal("0.01")


# ---------------------------------------------------------------------------
# Unit conversion
# ---------------------------------------------------------------------------

# Base conversion table: how many of each unit equal 1 "base unit".
# The base unit for weight is the **ounce**, for volume is the **fluid ounce**,
# and for count is the **each**.
_WEIGHT_OZ: dict[str, Decimal] = {
    "oz": Decimal("1"), "ounce": Decimal("1"), "ounces": Decimal("1"),
    "lb": Decimal("16"), "lbs": Decimal("16"), "pound": Decimal("16"), "pounds": Decimal("16"),
    "g": Decimal("0.035274"), "gram": Decimal("0.035274"), "grams": Decimal("0.035274"),
    "kg": Decimal("35.274"), "kilogram": Decimal("35.274"), "kilograms": Decimal("35.274"),
    "mg": Decimal("0.000035274"),
    "ton": Decimal("32000"), "tonne": Decimal("2204.62"),
}

_VOLUME_FLOZ: dict[str, Decimal] = {
    "fl oz": Decimal("1"), "fluid ounce": Decimal("1"), "fluid ounces": Decimal("1"),
    "cup": Decimal("8"), "cups": Decimal("8"),
    "pt": Decimal("16"), "pint": Decimal("16"), "pints": Decimal("16"),
    "qt": Decimal("32"), "quart": Decimal("32"), "quarts": Decimal("32"),
    "gal": Decimal("128"), "gallon": Decimal("128"), "gallons": Decimal("128"),
    "ml": Decimal("0.033814"), "milliliter": Decimal("0.033814"), "milliliters": Decimal("0.033814"),
    "l": Decimal("33.814"), "liter": Decimal("33.814"), "liters": Decimal("33.814"),
}

_COUNT: dict[str, Decimal] = {
    "each": Decimal("1"), "ea": Decimal("1"), "ct": Decimal("1"), "count": Decimal("1"),
    "case": Decimal("1"), "bunch": Decimal("1"), "bottle": Decimal("1"),
}


def _money(value: Any) -> Decimal:
    """Parse a money value into a 2-dp Decimal."""
    try:
        return dec(value).quantize(MONEY, rounding=ROUND_HALF_UP)
    except (ValueError, TypeError):
        return Decimal("0.00")


def _qty(value: Any) -> Decimal:
    """Parse a quantity value into a 4-dp Decimal."""
    try:
        return dec(value).quantize(QTY_SCALE, rounding=ROUND_HALF_UP)
    except (ValueError, TypeError):
        return Decimal("0")


def normalize_unit(value: Any) -> str:
    """Normalize a unit string to a canonical key."""
    return re.sub(r"\s+", " ", str(value or "").lower().strip())


def convert_quantity(quantity: Decimal, from_unit: str, to_unit: str) -> Decimal:
    """Convert *quantity* from *from_unit* to *to_unit*.

    Raises ``ValueError`` if the units belong to different categories
    (e.g. weight → volume).
    """
    fu = normalize_unit(from_unit)
    tu = normalize_unit(to_unit)

    # Direct match
    if fu == tu:
        return quantity

    # Look up conversion factors
    if fu in _WEIGHT_OZ and tu in _WEIGHT_OZ:
        return (quantity * _WEIGHT_OZ[fu] / _WEIGHT_OZ[tu]).quantize(QTY_SCALE, rounding=ROUND_HALF_UP)

    if fu in _VOLUME_FLOZ and tu in _VOLUME_FLOZ:
        return (quantity * _VOLUME_FLOZ[fu] / _VOLUME_FLOZ[tu]).quantize(QTY_SCALE, rounding=ROUND_HALF_UP)

    if fu in _COUNT and tu in _COUNT:
        return quantity  # count-to-count is 1:1 (each↔case handled by ingredient-specific units)

    raise ValueError(f"Cannot convert '{from_unit}' to '{to_unit}': incompatible unit categories")


def _canonical_unit(unit: str) -> tuple[str, Decimal]:
    """Return (canonical_unit_name, factor_to_ounce_or_fluid_oz).

    Returns the base category so we can detect mix-ups.
    """
    u = normalize_unit(unit)
    if u in _WEIGHT_OZ:
        return ("weight_oz", _WEIGHT_OZ[u])
    if u in _VOLUME_FLOZ:
        return ("volume_floz", _VOLUME_FLOZ[u])
    if u in _COUNT:
        return ("count", Decimal("1"))
    # Unknown unit — treat as "each" (count)
    return ("count", Decimal("1"))


# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class RecipeLineResult:
    ingredient_name: str
    quantity: Decimal
    unit: str
    unit_cost: Decimal
    ingredient_cost: Decimal
    source: str  # "spreadsheet" or "item_master"
    notes: str = ""


@dataclass
class MenuCostResult:
    menu_item_name: str
    menu_item_id: str
    menu_price: Decimal
    recipe_cost: Decimal
    food_cost_pct: Decimal
    contribution_margin: Decimal
    recommended_price: Decimal = Decimal("0.00")
    recommended_markup_pct: Decimal = Decimal("0.00")
    lines: list[RecipeLineResult] = field(default_factory=list)
    missing_cost_ingredients: list[str] = field(default_factory=list)


@dataclass
class RecipeImportResult:
    menu_items: list[MenuCostResult]
    missing_ingredients: list[str]
    errors: list[str]
    rows_processed: int


# ---------------------------------------------------------------------------
# Core costing engine
# ---------------------------------------------------------------------------

class RecipeCostingService:
    """Cost recipes from a spreadsheet using local item-master data."""

    def __init__(self, workspace: Any):
        self.workspace = workspace

    # ------------------------------------------------------------------
    # Ingredient / item master lookup
    # ------------------------------------------------------------------

    def _fetch_item(self, conn: sqlite3.Connection, ingredient_name: str) -> dict[str, Any] | None:
        """Look up an inventory item by name (fuzzy, case-insensitive).

        Tries an exact normalized match first, then a LIKE on the raw name.
        """
        norm = normalize(ingredient_name)
        row = conn.execute(
            "SELECT item_id, item_name, current_price, units_per_purchase_unit, "
            "count_unit, unit, vendor_name, vendor_sku "
            "FROM items WHERE normalized_description=? ORDER BY last_purchase_date DESC LIMIT 1",
            (norm,),
        ).fetchone()
        if row:
            return dict(row)
        # Fallback: partial name match
        row = conn.execute(
            "SELECT item_id, item_name, current_price, units_per_purchase_unit, "
            "count_unit, unit, vendor_name, vendor_sku "
            "FROM items WHERE LOWER(item_name) LIKE ? ORDER BY last_purchase_date DESC LIMIT 1",
            (f"%{ingredient_name.lower().strip()}%",),
        ).fetchone()
        if row:
            return dict(row)
        return None

    def _per_unit_cost_from_inventory(
        self, item: dict[str, Any], target_unit: str, *, ingredient_name: str = "", row_num: int = 0
    ) -> tuple[Decimal, str]:
        """Calculate the cost per *target_unit* from inventory price data.

        Returns ``(cost_per_target_unit, error_message)``. When the target unit
        is incompatible with the item's count_unit (e.g. "oz" for an item
        measured in "each"), the error is returned and the caller flags it.
        """
        price = _money(item.get("current_price") or 0)
        if price == 0:
            return Decimal("0.00"), ""

        units_per_purchase = _qty(item.get("units_per_purchase_unit") or 1) or Decimal("1")
        item_count_unit = (item.get("count_unit") or item.get("unit") or "each").strip()

        cost_per_count_unit = (price / units_per_purchase).quantize(
            Decimal("0.00000001"), rounding=ROUND_HALF_UP
        )

        try:
            factor = convert_quantity(Decimal("1"), item_count_unit, target_unit)
            return (cost_per_count_unit / factor).quantize(
                Decimal("0.00000001"), rounding=ROUND_HALF_UP
            ), ""
        except ValueError:
            # Incompatible unit categories — flag for review
            return cost_per_count_unit, (
                f"Row {row_num} ({ingredient_name}): unit '{target_unit}' is incompatible "
                f"with inventory count unit '{item_count_unit}' for "
                f"{item.get('item_name', ingredient_name)} — please use matching units"
            )

    # ------------------------------------------------------------------
    # Spreadsheet reading
    # ------------------------------------------------------------------

    @staticmethod
    def _read_rows(path: Path) -> list[dict[str, Any]]:
        """Read a CSV or XLSX file into a list of dicts with stripped keys."""
        path = Path(path)
        if is_excel_path(path):
            return read_xlsx(path)
        rows: list[dict[str, Any]] = []
        with path.open("r", encoding="utf-8-sig", newline="") as fh:
            reader = csv.DictReader(fh)
            for row in reader:
                cleaned: dict[str, Any] = {}
                for key, value in row.items():
                    if key is None:
                        continue
                    cleaned[key.strip()] = value.strip() if isinstance(value, str) else value
                rows.append(cleaned)
        return rows

    @staticmethod
    def _get_row(row: dict[str, Any], *aliases: str, default: Any = "") -> str:
        """Try multiple column names and return the first non-empty value."""
        for alias in aliases:
            value = row.get(alias)
            if value is not None and str(value).strip():
                return str(value).strip()
        return str(default) if default is not None else ""

    # ------------------------------------------------------------------
    # Main entry point
    # ------------------------------------------------------------------

    def import_recipes(self, path: Path) -> RecipeImportResult:
        """Import recipes from a spreadsheet and cost them against inventory.

        Returns a :class:`RecipeImportResult` with per-menu-item costs,
        any missing ingredients, and row-level errors.
        """
        path = Path(path)
        rows = self._read_rows(path)
        errors: list[str] = []
        missing_ingredients: list[str] = []
        menu_items: dict[str, MenuCostResult] = {}

        with self.workspace.connect() as conn:
            for idx, row in enumerate(rows, start=2):
                menu_name = self._get_row(row, "Menu Item Name", "menu_item_name", "menu item name")
                if not menu_name:
                    errors.append(f"Row {idx}: Menu Item Name is required")
                    continue

                ingredient_name = self._get_row(
                    row, "Ingredient Name", "ingredient", "ingredient_name", "Ingredient"
                )
                if not ingredient_name:
                    errors.append(f"Row {idx} ({menu_name}): Ingredient Name is required")
                    continue

                quantity_raw = self._get_row(
                    row, "Quantity", "quantity", "qty", "amount"
                )
                if not quantity_raw:
                    errors.append(f"Row {idx} ({menu_name} → {ingredient_name}): Quantity is required")
                    continue

                unit = self._get_row(row, "Unit", "unit", "uom", "measurement")
                if not unit:
                    unit = "each"

                # Menu price
                menu_price_raw = self._get_row(row, "Menu Price", "menu_price", "selling price")
                menu_price = _money(menu_price_raw or 0)

                # Menu item ID (for lookup or generation)
                menu_key = self._get_row(row, "POS Item Key", "pos_item_key", "item key")
                menu_category = self._get_row(row, "Menu Category", "menu_category", "category", default="Unclassified")

                # Yield / notes
                yield_pct = self._get_row(row, "Yield Percent", "yield_percent", default="100")
                notes = self._get_row(row, "Notes", "notes")

                # Unit cost from spreadsheet (if provided)
                unit_cost_raw = self._get_row(row, "Unit Cost", "unit_cost", "Cost Per Unit", "cost")
                unit_cost_from_sheet: Decimal | None = None
                if unit_cost_raw:
                    unit_cost_from_sheet = _money(unit_cost_raw)

                # --- Look up or create menu item ---
                menu_key_norm = menu_key or normalize(menu_name)
                existing_menu = conn.execute(
                    "SELECT menu_item_id FROM menu_items WHERE pos_item_key=? COLLATE NOCASE OR menu_item_name=? LIMIT 1",
                    (menu_key_norm, menu_name),
                ).fetchone()
                if existing_menu:
                    menu_item_id = existing_menu["menu_item_id"]
                else:
                    menu_item_id = f"MENU-{hashlib_sha256(menu_key_norm.encode()).hexdigest()[:14].upper()}"
                    conn.execute(
                        "INSERT INTO menu_items(menu_item_id,pos_item_key,menu_item_name,category,menu_price,created_at,updated_at) "
                        "VALUES(?,?,?,?,?,?,?) "
                        "ON CONFLICT(menu_item_id) DO UPDATE SET "
                        "menu_item_name=excluded.menu_item_name, category=excluded.category, "
                        "menu_price=excluded.menu_price, updated_at=excluded.updated_at",
                        (menu_item_id, menu_key_norm, menu_name, menu_category, f"{menu_price:.2f}", _now(), _now()),
                    )

                # --- Resolve ingredient cost ---
                quantity = _qty(quantity_raw)
                source = ""
                resolved_unit_cost: Decimal

                if unit_cost_from_sheet is not None and unit_cost_from_sheet > 0:
                    resolved_unit_cost = unit_cost_from_sheet
                    source = "spreadsheet"
                else:
                    item = self._fetch_item(conn, ingredient_name)
                    if item:
                        resolved_unit_cost, unit_err = self._per_unit_cost_from_inventory(
                            item, unit, ingredient_name=ingredient_name, row_num=idx
                        )
                        source = f"item_master:{item.get('item_id', '?')}"
                        if unit_err:
                            missing_ingredients.append(ingredient_name)
                            errors.append(unit_err)
                            resolved_unit_cost = Decimal("0.00")
                        elif resolved_unit_cost == 0:
                            source = "item_master:zero-price"
                            missing_ingredients.append(ingredient_name)
                            errors.append(
                                f"Row {idx} ({menu_name} → {ingredient_name}): "
                                f"Item found in inventory but current_price is 0"
                            )
                    else:
                        source = "not_found"
                        missing_ingredients.append(ingredient_name)
                        errors.append(
                            f"Row {idx} ({menu_name} → {ingredient_name}): "
                            f"No matching inventory item and no Unit Cost provided"
                        )
                        resolved_unit_cost = Decimal("0.00")

                # Adjust for yield
                yield_factor = dec(yield_pct, "100") / Decimal("100")
                if yield_factor <= 0 or yield_factor > 1:
                    yield_factor = Decimal("1")
                effective_qty = quantity / yield_factor
                ingredient_cost = _money(resolved_unit_cost * effective_qty)

                # Store in recipe_ingredients table
                item_row = self._fetch_item(conn, ingredient_name) if source.startswith("item_master") else None
                item_id = item_row["item_id"] if item_row else None
                if item_id:
                    conn.execute(
                        "INSERT INTO recipe_ingredients(menu_item_id,item_id,quantity_count_units,yield_percent,notes,created_at,updated_at) "
                        "VALUES(?,?,?,?,?,?,?) "
                        "ON CONFLICT(menu_item_id,item_id) DO UPDATE SET "
                        "quantity_count_units=excluded.quantity_count_units, "
                        "yield_percent=excluded.yield_percent, notes=excluded.notes, updated_at=excluded.updated_at",
                        (menu_item_id, item_id, f"{quantity:.4f}", f"{dec(yield_pct, '100'):.2f}", notes, _now(), _now()),
                    )

                # Accumulate into menu_items dict
                if menu_item_id not in menu_items:
                    menu_items[menu_item_id] = MenuCostResult(
                        menu_item_name=menu_name,
                        menu_item_id=menu_item_id,
                        menu_price=menu_price,
                        recipe_cost=Decimal("0"),
                        food_cost_pct=Decimal("0"),
                        contribution_margin=Decimal("0"),
                    )
                mr = menu_items[menu_item_id]
                mr.lines.append(RecipeLineResult(
                    ingredient_name=ingredient_name,
                    quantity=quantity,
                    unit=unit,
                    unit_cost=resolved_unit_cost,
                    ingredient_cost=ingredient_cost,
                    source=source,
                    notes=notes,
                ))
                mr.recipe_cost += ingredient_cost

            # Calculate food cost % and contribution margin for each menu item
            for mr in menu_items.values():
                mr.recipe_cost = _money(mr.recipe_cost)
                if mr.menu_price and mr.menu_price > 0:
                    mr.food_cost_pct = _money(
                        (mr.recipe_cost / mr.menu_price * Decimal("100"))
                        .quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                    )
                else:
                    mr.food_cost_pct = Decimal("0.00")
                mr.contribution_margin = _money(mr.menu_price - mr.recipe_cost)
                # Price recommendation: at least 3x recipe cost
                three_times = mr.recipe_cost * Decimal("3")
                if three_times > mr.menu_price:
                    mr.recommended_price = _money(three_times)
                    if mr.recipe_cost > 0:
                        mr.recommended_markup_pct = _money(
                            ((mr.recommended_price / mr.recipe_cost * Decimal("100")) - Decimal("100"))
                            .quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
                        )
                    else:
                        mr.recommended_markup_pct = Decimal("0.00")
                else:
                    mr.recommended_price = mr.menu_price
                    mr.recommended_markup_pct = Decimal("0.00")

        return RecipeImportResult(
            menu_items=list(menu_items.values()),
            missing_ingredients=list(dict.fromkeys(missing_ingredients)),
            errors=errors,
            rows_processed=len(rows),
        )

    # ------------------------------------------------------------------
    # Daily / weekly / monthly sales costing
    # ------------------------------------------------------------------

    def get_recipe_costs(self) -> dict[str, Decimal]:
        """Return a mapping of menu item name → recipe cost from the database.

        Looks up costed recipes from the ``recipe_ingredients`` table joined
        with ``menu_items`` and ``items``.
        """
        costs: dict[str, Decimal] = {}
        with self.workspace.connect() as conn:
            rows = conn.execute(
                """SELECT m.menu_item_name,
                          COALESCE(SUM(ri.quantity_count_units *
                                      (i.current_price / NULLIF(i.units_per_purchase_unit, '0')
                                       / (NULLIF(ri.yield_percent, '0') / 100.0))), 0) AS recipe_cost
                   FROM recipe_ingredients ri
                   JOIN menu_items m ON m.menu_item_id = ri.menu_item_id
                   JOIN items i ON i.item_id = ri.item_id
                   WHERE i.current_price IS NOT NULL
                   GROUP BY m.menu_item_name"""
            ).fetchall()
            for row in rows:
                costs[row["menu_item_name"]] = _money(row["recipe_cost"])
        return costs

    def calculate_daily_sales_cost(
        self, sales_document: Path | str, *, recipe_results: RecipeImportResult | None = None
    ) -> DailySalesResult:
        """Cost a daily/weekly/monthly sales document against recipe costs.

        Reads *sales_document* (CSV, XLSX, TXT, or PDF), matches each sales
        line to a menu item, calculates the total recipe cost for items sold,
        and produces price recommendations where the current menu price is
        below 3× the recipe cost.

        If *recipe_results* is provided (from a recent :meth:`import_recipes`
        call), those costs are used directly.  Otherwise the method looks up
        recipe costs from the database via :meth:`get_recipe_costs`.
        """
        path = Path(sales_document)
        rows = read_document(path)
        sales_rows, errors, period_start, period_end = _parse_sales_rows(rows)

        # Build menu item name → recipe cost mapping
        if recipe_results is not None:
            recipe_costs = {mr.menu_item_name: mr.recipe_cost for mr in recipe_results.menu_items}
            menu_items = list(recipe_results.menu_items)
        else:
            recipe_costs = self.get_recipe_costs()
            menu_items = []

        total_net_sales = Decimal("0")
        total_recipe_cost = Decimal("0")
        matched_sales: list[DailySalesRow] = []

        for sr in sales_rows:
            net_sales = sr.net_sales
            total_net_sales += net_sales
            recipe_cost_per_item = recipe_costs.get(sr.menu_item_name)

            if recipe_cost_per_item is not None:
                item_cost = _money(recipe_cost_per_item * sr.quantity)
                total_recipe_cost += item_cost
                matched_sales.append(sr)

                # Price recommendation: at least 3x recipe cost
                current_price = net_sales / sr.quantity if sr.quantity > 0 else Decimal("0")
                recommended = _money(recipe_cost_per_item * Decimal("3"))
                if recommended > current_price:
                    price_rec = {
                        "menu_item_name": sr.menu_item_name,
                        "quantity_sold": f"{sr.quantity:.4f}",
                        "current_price": f"{current_price:.2f}",
                        "recipe_cost": f"{recipe_cost_per_item:.2f}",
                        "recommended_price": f"{recommended:.2f}",
                        "reason": "below 3x markup",
                    }
                else:
                    price_rec = None
            else:
                errors.append(
                    f"Row {sr.source_row}: No recipe cost found for '{sr.menu_item_name}'"
                )
                price_rec = None

        total_gross_profit = _money(total_net_sales - total_recipe_cost)
        overall_food_cost_pct = (
            _money(total_recipe_cost / total_net_sales * Decimal("100"))
            if total_net_sales > 0 else Decimal("0.00")
        )

        # Collect price recommendations
        recommendations: list[dict[str, Any]] = []
        seen_names: set[str] = set()
        for sr in sales_rows:
            if sr.menu_item_name in seen_names:
                continue
            seen_names.add(sr.menu_item_name)
            rc = recipe_costs.get(sr.menu_item_name)
            if rc is not None:
                recommended = _money(rc * Decimal("3"))
                # Calculate average price from sales
                item_sales = [s for s in sales_rows if s.menu_item_name == sr.menu_item_name]
                total_qty = sum((s.quantity for s in item_sales), Decimal("0"))
                total_rev = sum((s.net_sales for s in item_sales), Decimal("0"))
                current_price = total_rev / total_qty if total_qty > 0 else Decimal("0")
                if recommended > current_price:
                    recommendations.append({
                        "menu_item_name": sr.menu_item_name,
                        "total_quantity_sold": f"{total_qty:.4f}",
                        "total_revenue": f"{total_rev:.2f}",
                        "avg_current_price": f"{current_price:.2f}",
                        "recipe_cost": f"{rc:.2f}",
                        "recommended_price": f"{recommended:.2f}",
                        "potential_extra_profit": f"{_money((recommended - current_price) * total_qty):.2f}" if total_qty > 0 else "$0.00",
                    })

        return DailySalesResult(
            period_start=period_start,
            period_end=period_end,
            menu_items=menu_items,
            sales_rows=matched_sales,
            total_net_sales=_money(total_net_sales),
            total_recipe_cost=_money(total_recipe_cost),
            total_gross_profit=total_gross_profit,
            overall_food_cost_pct=_money(overall_food_cost_pct),
            price_recommendations=recommendations,
            errors=errors,
        )

    # ------------------------------------------------------------------
    # Exporters
    # ------------------------------------------------------------------

    def export_recipe_cost_report(
        self, results: RecipeImportResult, destination: Path | None = None
    ) -> Path:
        """Export a recipe cost report as CSV."""
        destination = destination or Path(self.workspace.folders.get(
            "recipes", self.workspace.root / "Recipes"
        )) / "Recipe_Cost_Report.csv"
        destination = Path(destination)
        destination.parent.mkdir(parents=True, exist_ok=True)

        fieldnames = [
            "Menu Item ID", "Menu Item Name", "Menu Price", "Recipe Cost",
            "Food Cost %", "Contribution Margin", "Recommended Price", "Recommended Markup %",
            "Ingredient Name", "Quantity", "Unit", "Unit Cost",
            "Ingredient Cost", "Cost Source", "Notes",
        ]
        with destination.open("w", encoding="utf-8", newline="") as fh:
            writer = csv.writer(fh)
            writer.writerow(fieldnames)
            for mr in results.menu_items:
                for line in mr.lines:
                    writer.writerow([
                        mr.menu_item_id, mr.menu_item_name,
                        f"{mr.menu_price:.2f}", f"{mr.recipe_cost:.2f}",
                        f"{mr.food_cost_pct:.2f}", f"{mr.contribution_margin:.2f}",
                        f"{mr.recommended_price:.2f}", f"{mr.recommended_markup_pct:.2f}",
                        line.ingredient_name, f"{line.quantity:.4f}", line.unit,
                        f"{line.unit_cost:.6f}", f"{line.ingredient_cost:.4f}",
                        line.source, line.notes,
                    ])
        return destination

    def export_recipe_cost_report_xlsx(self, results: RecipeImportResult, destination: Path | None = None) -> Path:
        """Export a recipe cost report as XLSX (summary + detail sheets)."""
        from excel_io import write_xlsx

        destination = destination or Path(self.workspace.folders.get(
            "recipes", self.workspace.root / "Recipes"
        )) / "Recipe_Cost_Report.xlsx"
        destination = Path(destination)
        destination.parent.mkdir(parents=True, exist_ok=True)

        # Summary sheet
        summary_records = []
        for mr in results.menu_items:
            summary_records.append({
                "Menu Item ID": mr.menu_item_id,
                "Menu Item Name": mr.menu_item_name,
                "Menu Price": f"{mr.menu_price:.2f}",
                "Recipe Cost": f"{mr.recipe_cost:.2f}",
                "Food Cost %": f"{mr.food_cost_pct:.2f}",
                "Contribution Margin": f"{mr.contribution_margin:.2f}",
                "Recommended Price": f"{mr.recommended_price:.2f}",
                "Recommended Markup %": f"{mr.recommended_markup_pct:.2f}",
                "Ingredients": len(mr.lines),
            })

        # Detail sheet
        detail_records = []
        for mr in results.menu_items:
            for line in mr.lines:
                detail_records.append({
                    "Menu Item ID": mr.menu_item_id,
                    "Menu Item Name": mr.menu_item_name,
                    "Menu Price": f"{mr.menu_price:.2f}",
                    "Recipe Cost": f"{mr.recipe_cost:.2f}",
                    "Food Cost %": f"{mr.food_cost_pct:.2f}",
                    "Ingredient Name": line.ingredient_name,
                    "Quantity": f"{line.quantity:.4f}",
                    "Unit": line.unit,
                    "Unit Cost": f"{line.unit_cost:.6f}",
                    "Ingredient Cost": f"{line.ingredient_cost:.4f}",
                    "Cost Source": line.source,
                    "Notes": line.notes,
                })

        # Write both sheets into a single XLSX
        import openpyxl  # type: ignore
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        if summary_records:
            headers = list(summary_records[0].keys())
            ws.append(headers)
            for record in summary_records:
                ws.append(list(record.values()))
        ws2 = wb.create_sheet("Detail")
        if detail_records:
            headers2 = list(detail_records[0].keys())
            ws2.append(headers2)
            for record in detail_records:
                ws2.append(list(record.values()))
        # Errors sheet
        if results.errors:
            ws3 = wb.create_sheet("Errors")
            ws3.append(["Row", "Error"])
            for err in results.errors:
                ws3.append([None, err])
        wb.save(str(destination))
        return destination


# ---------------------------------------------------------------------------
# Daily sales costing
# ---------------------------------------------------------------------------

@dataclass
class DailySalesRow:
    """A single line from a sales document."""
    menu_item_name: str
    quantity: Decimal
    net_sales: Decimal
    gross_sales: Decimal
    source_row: int
    raw: dict[str, Any]


@dataclass
class DailySalesResult:
    """Result of costing a daily/weekly/monthly sales document."""
    period_start: str
    period_end: str
    menu_items: list[MenuCostResult]
    sales_rows: list[DailySalesRow]
    total_net_sales: Decimal
    total_recipe_cost: Decimal
    total_gross_profit: Decimal
    overall_food_cost_pct: Decimal
    price_recommendations: list[dict[str, Any]]
    errors: list[str] = field(default_factory=list)


# ---------------------------------------------------------------------------
# Document format detection and reading
# ---------------------------------------------------------------------------

# Flexible column name matching for sales columns
_SALES_NAME_ALIASES = ("menu item name", "item name", "product name", "menu item", "item", "description")
_SALES_QTY_ALIASES = ("quantity", "qty", "units sold", "count", "servings", "number sold")
_SALES_NET_ALIASES = ("net sales", "net", "net_sales", "sales")
_SALES_GROSS_ALIASES = ("gross sales", "gross", "gross_sales", "total", "sales amount")
_SALES_DATE_ALIASES = ("date", "period", "business date", "sale date", "day", "week", "month")


def _find_col(row: dict[str, Any], aliases: Iterable[str]) -> str | None:
    """Return the first key in *row* that matches any of *aliases* (case-insensitive, flexible)."""
    for key in row:
        key_clean = re.sub(r"\s+", " ", str(key).strip().lower())
        for alias in aliases:
            if key_clean == alias or alias in key_clean:
                return key
    return None


def _read_csv_rows(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as fh:
        sample = fh.read(8192)
        fh.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|\t")
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(fh, dialect=str(dialect) if isinstance(dialect, str) else dialect)
        for row in reader:
            rows.append({str(k).strip() if k else "": v for k, v in row.items()})
    return rows


def _read_xlsx_rows(path: Path) -> list[dict[str, Any]]:
    from excel_io import read_xlsx
    return read_xlsx(path)


def _read_txt_rows(path: Path) -> list[dict[str, Any]]:
    """Read a tab- or pipe-delimited text file."""
    rows: list[dict[str, Any]] = []
    with path.open("r", encoding="utf-8-sig") as fh:
        sample = fh.read(8192)
        fh.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters="\t,;|\t")
        except csv.Error:
            dialect = csv.excel
        reader = csv.DictReader(fh, dialect=str(dialect) if isinstance(dialect, str) else dialect)
        for row in reader:
            rows.append({str(k).strip() if k else "": v for k, v in row.items()})
    return rows


def _read_pdf_rows(path: Path) -> list[dict[str, Any]]:
    """Extract tabular data from a PDF using PyMuPDF."""
    rows: list[dict[str, Any]] = []
    try:
        import fitz  # type: ignore
    except ImportError:
        return rows
    doc = fitz.open(str(path))
    if doc.page_count == 0:
        return rows
    # Use table extraction (PyMuPDF 1.23+)
    tables = []
    for page in doc:
        try:
            page_tables = page.find_tables()
            if page_tables:
                tables.extend(page_tables)
        except Exception:
            pass
    if tables:
        for table in tables:
            rows.extend(table.extract_dict()["rows"])
    else:
        # Fallback: extract text lines and parse whitespace-separated columns
        for page in doc:
            text = page.get_text()
            lines = [line for line in text.split("\n") if line.strip()]
            if not lines:
                continue
            header = re.split(r"\s{2,}", lines[0])
            header = [h.strip() for h in header]
            for line in lines[1:]:
                values = re.split(r"\s{2,}", line)
                if len(values) == len(header):
                    rows.append(dict(zip(header, [v.strip() for v in values])))
    return rows


def read_document(path: Path) -> list[dict[str, Any]]:
    """Read a sales or recipe spreadsheet from any supported format.

    Supports: .csv, .tsv, .txt, .xlsx, .xlsm, .pdf
    Returns a list of row dicts with stripped string keys.
    """
    path = Path(path)
    suffix = path.suffix.lower()

    if suffix in (".csv", ".tsv"):
        return _read_csv_rows(path)
    if suffix in (".xlsx", ".xlsm"):
        return _read_xlsx_rows(path)
    if suffix == ".txt":
        return _read_txt_rows(path)
    if suffix == ".pdf":
        return _read_pdf_rows(path)
    # Unknown extension: try CSV first, then TXT
    try:
        return _read_csv_rows(path)
    except Exception:
        return _read_txt_rows(path)


def parse_date(value: Any) -> str:
    """Parse a date string in various formats to ISO 'YYYY-MM-DD'."""
    from datetime import datetime as dt
    text = str(value or "").strip()
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m-%d-%Y", "%m/%d/%y", "%Y/%m/%d", "%m/%d/%Y %H:%M"):
        try:
            return dt.strptime(text.split()[0], fmt).date().isoformat()
        except ValueError:
            continue
    return text


def _parse_sales_rows(rows: list[dict[str, Any]]) -> tuple[list[DailySalesRow], list[str], str, str]:
    """Parse raw rows into DailySalesRow entries with flexible column matching.

    Returns (sales_rows, errors, period_start, period_end)
    """
    sales_rows: list[DailySalesRow] = []
    errors: list[str] = []
    period_start = ""
    period_end = ""

    for idx, row in enumerate(rows, start=2):  # 1 = header
        name_key = _find_col(row, _SALES_NAME_ALIASES)
        qty_key = _find_col(row, _SALES_QTY_ALIASES)
        net_key = _find_col(row, _SALES_NET_ALIASES)
        gross_key = _find_col(row, _SALES_GROSS_ALIASES)
        date_key = _find_col(row, _SALES_DATE_ALIASES)

        menu_name = row.get(name_key) if name_key else ""
        menu_name = str(menu_name or "").strip()
        if not menu_name:
            continue

        qty_raw = row.get(qty_key) if qty_key else ""
        quantity = _qty(qty_raw)
        if quantity == 0:
            quantity = Decimal("1")  # Assume 1 if only revenue is tracked

        net_raw = row.get(net_key) if net_key else ""
        gross_raw = row.get(gross_key) if gross_key else ""

        net_sales = _money(net_raw) or _money(gross_raw)
        gross_sales = _money(gross_raw) or net_sales

        # Track date range
        date_raw = str(row.get(date_key) or "").strip()
        if date_raw:
            try:
                parsed_date = parse_date(date_raw)
                if not period_start or parsed_date < period_start:
                    period_start = parsed_date
                if not period_end or parsed_date > period_end:
                    period_end = parsed_date
            except Exception:
                pass

        sales_rows.append(DailySalesRow(
            menu_item_name=menu_name,
            quantity=quantity,
            net_sales=net_sales,
            gross_sales=gross_sales,
            source_row=idx,
            raw=dict(row),
        ))

    return sales_rows, errors, period_start, period_end


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _now() -> str:
    return datetime.now().replace(microsecond=0).isoformat()


def hashlib_sha256(data: bytes) -> Any:
    """Wrapper to keep the hashing import localized."""
    import hashlib
    return hashlib.sha256(data)


def export_daily_sales_report(
    workspace: Any, sales_result: DailySalesResult, destination: Path | None = None
) -> Path:
    """Export a daily/periodic sales cost report as CSV.

    Columns: Period Start, Period End, Menu Item, Qty Sold, Net Sales,
    Recipe Cost, Gross Profit, Food Cost %.
    Also writes a Price Recommendations sheet as a separate CSV.
    """
    destination = destination or Path(workspace.folders.get(
        "recipes", Path(workspace.root) / "Recipes"
    )) / "Daily_Sales_Cost_Report.csv"
    destination = Path(destination)
    destination.parent.mkdir(parents=True, exist_ok=True)

    fieldnames = [
        "Period Start", "Period End",
        "Menu Item Name", "Quantity Sold", "Net Sales",
        "Recipe Cost", "Gross Profit", "Food Cost %",
    ]
    with destination.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow(fieldnames)

        # Build per-item totals
        item_totals: dict[str, Decimal] = {}
        item_qty: dict[str, Decimal] = {}
        for sr in sales_result.sales_rows:
            key = sr.menu_item_name
            item_totals[key] = item_totals.get(key, Decimal("0")) + sr.net_sales
            item_qty[key] = item_qty.get(key, Decimal("0")) + sr.quantity

        recipe_costs = {mr.menu_item_name: mr.recipe_cost for mr in sales_result.menu_items}

        for name in sorted(item_totals.keys()):
            qty = item_qty[name]
            net = _money(item_totals[name])
            rc = recipe_costs.get(name, Decimal("0"))
            cost = _money(rc * qty)
            gross = _money(net - cost)
            fcpct = _money(cost / net * Decimal("100")) if net > 0 else Decimal("0")
            writer.writerow([
                sales_result.period_start, sales_result.period_end,
                name, f"{qty:.4f}", f"{net:.2f}",
                f"{cost:.2f}", f"{gross:.2f}", f"{fcpct:.2f}",
            ])

        # Summary row
        writer.writerow([])
        writer.writerow(["SUMMARY"])
        writer.writerow(["Total Net Sales", f"{sales_result.total_net_sales:.2f}"])
        writer.writerow(["Total Recipe Cost", f"{sales_result.total_recipe_cost:.2f}"])
        writer.writerow(["Total Gross Profit", f"{sales_result.total_gross_profit:.2f}"])
        writer.writerow(["Overall Food Cost %", f"{sales_result.overall_food_cost_pct:.2f}"])

    # Price recommendations CSV
    rec_path = destination.parent / "Price_Recommendations.csv"
    with rec_path.open("w", encoding="utf-8", newline="") as fh:
        writer = csv.writer(fh)
        writer.writerow([
            "Menu Item", "Qty Sold", "Current Revenue", "Avg Price",
            "Recipe Cost", "3x Cost", "Recommended Price", "Potential Extra Profit",
        ])
        for rec in sales_result.price_recommendations:
            writer.writerow([
                rec["menu_item_name"], rec["total_quantity_sold"], rec["total_revenue"],
                rec["avg_current_price"], rec["recipe_cost"],
                f"{_money(Decimal(rec['recipe_cost']) * Decimal('3')):.2f}",
                rec["recommended_price"], rec["potential_extra_profit"],
            ])

    return destination


# ---------------------------------------------------------------------------
# Convenience: cost a single spreadsheet and print summary
# ---------------------------------------------------------------------------

def cost_recipes(workspace: Any, spreadsheet_path: Path) -> RecipeImportResult:
    """Quick one-shot call: cost recipes from *spreadsheet_path* in *workspace*."""
    service = RecipeCostingService(workspace)
    return service.import_recipes(Path(spreadsheet_path))


def cost_daily_sales(
    workspace: Any, sales_document: Path | str,
    *, recipe_results: RecipeImportResult | None = None,
) -> DailySalesResult:
    """Quick one-shot call: cost a daily/weekly/monthly sales document."""
    service = RecipeCostingService(workspace)
    return service.calculate_daily_sales_cost(Path(sales_document), recipe_results=recipe_results)
