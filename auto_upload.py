#!/usr/bin/env python3
"""Single-folder automatic intake and routing for MarginMise.

Each restaurant receives a visible Desktop upload folder. Files dropped directly
into that folder are claimed atomically, classified from their contents, routed
through the application's existing service layer, and organized into Processed,
Needs Review, or Failed archives.

The service intentionally polls the filesystem instead of requiring an external
watchdog package. That keeps the desktop installation small and makes the same
runtime work on Windows, macOS, and Linux.
"""
from __future__ import annotations

import csv
import ctypes
import hashlib
import json
import os
import re
import shutil
import subprocess
import sys
import threading
import time
import uuid
import zipfile
from dataclasses import asdict, dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, NamedTuple, Callable, Iterable

from invoice_pipeline import (
    InvoicePipeline,
    RestaurantWorkspace,
    parse_date,
    safe_filename,
    canonical_inventory_category,
)
from operational_controls import AuthenticatedUser
AUTO_UPLOAD_SCHEMA_SQL = """
CREATE TABLE IF NOT EXISTS auto_upload_events (
    event_id INTEGER PRIMARY KEY AUTOINCREMENT,
    source_sha256 TEXT NOT NULL,
    original_name TEXT NOT NULL,
    original_path TEXT NOT NULL,
    detected_type TEXT NOT NULL,
    classification_confidence REAL NOT NULL DEFAULT 0,
    status TEXT NOT NULL,
    summary TEXT NOT NULL,
    details_json TEXT NOT NULL DEFAULT '{}',
    archived_path TEXT,
    created_at TEXT NOT NULL,
    completed_at TEXT NOT NULL
);
CREATE INDEX IF NOT EXISTS idx_auto_upload_hash ON auto_upload_events(source_sha256, status);
CREATE INDEX IF NOT EXISTS idx_auto_upload_completed ON auto_upload_events(completed_at DESC);
"""

INVOICE_SUFFIXES = {".pdf", ".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp"}
TABLE_SUFFIXES = {".csv", ".tsv", ".txt", ".xlsx", ".xlsm"}
SUPPORTED_SUFFIXES = INVOICE_SUFFIXES | TABLE_SUFFIXES | {".json", ".ics", ".zip"}
TEMP_SUFFIXES = {".tmp", ".part", ".crdownload", ".download", ".lock"}
SYSTEM_FILES = {
    "README_DROP_FILES_HERE.txt",
    ".restaurant_workspace.json",
    "AUTO_UPLOAD_ACTIVITY.csv",
    "restaurant_config.json",
}
DISCOVERY_DOCUMENT_TYPES = {
    "Invoice",
    "POS Sales",
    "Sales Summary",
    "Inventory Count",
    "Item Planning",
    "Menu Items",
    "Recipes",
    "Operating Costs",
    "Waste Log",
    "Receiving Log",
    "Distributor Catalog",
    "Distributor Confirmation",
    "Accounting Mappings",
    "Event Calendar",
    "Archive",
}
DISCOVERY_EXCLUDED_DIRECTORY_NAMES = {
    ".git",
    ".hg",
    ".svn",
    ".venv",
    "__pycache__",
    "node_modules",
    "$recycle.bin",
    "system volume information",
}


class AutoUploadError(RuntimeError):
    pass


class AutoUploadRetryLater(AutoUploadError):
    """Raised for transient failures such as a database lock or incomplete copy."""


@dataclass
class Classification:
    detected_type: str
    confidence: float
    reason: str
    headers: list[str] = field(default_factory=list)
    row_count: int = 0


@dataclass
class RoutingOutcome:
    status: str
    detected_type: str
    summary: str
    imported: int = 0
    rejected: int = 0
    details: dict[str, Any] = field(default_factory=dict)
    archived_path: str | None = None

    def as_dict(self) -> dict[str, Any]:
        return asdict(self)


@dataclass
class DocumentDiscoveryReport:
    run_id: str
    source_root: str
    started_at: str
    completed_at: str = ""
    status: str = "Running"
    scanned_files: int = 0
    supported_files: int = 0
    queued_files: int = 0
    duplicate_files: int = 0
    oversized_files: int = 0
    unclassified_files: int = 0
    error_files: int = 0
    stopped_at_limit: bool = False
    queued: list[dict[str, Any]] = field(default_factory=list)
    skipped: list[dict[str, Any]] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)
    report_path: str = ""

    def as_dict(self) -> dict[str, Any]:
        return asdict(self)

    @property
    def summary(self) -> str:
        limit_note = " Scan limit reached." if self.stopped_at_limit else ""
        return (
            f"Searched {self.scanned_files} file(s), found {self.supported_files} supported "
            f"restaurant document(s), and queued {self.queued_files} for Auto Upload. "
            f"Skipped {self.duplicate_files} duplicate(s), {self.unclassified_files} "
            f"unclassified file(s), {self.oversized_files} oversized file(s), and "
            f"recorded {self.error_files} error(s).{limit_note}"
        )


def now_iso() -> str:
    return datetime.now().replace(microsecond=0).isoformat()


# Restaurant identity discovery ------------------------------------------------
#
# When a user points MarginMise at a folder, the program should create the
# restaurant automatically and fill in whatever identity details it can find.
# Detection is intentionally conservative: only fields that were actually
# observed are returned, so callers can fill empty settings without clobbering
# values a manager already entered.

RESTAURANT_IDENTITY_FILENAMES = (
    "restaurant_info.json",
    "restaurant.json",
    "info.json",
    "company.json",
    "business.json",
)

_RESTAURANT_NAME_KEYWORDS = (
    "bar", "grill", "restaurant", "cafe", "café", "diner", "kitchen",
    "pub", "bistro", "eatery", "pizzeria", "tavern", "bakery", "kitchen",
)


def _coerce_str(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _looks_like_restaurant_name(text: str) -> bool:
    text = (text or "").strip()
    if not text or len(text) > 80:
        return False
    lowered = text.lower()
    return any(kw in lowered for kw in _RESTAURANT_NAME_KEYWORDS)


def _apply_identity_json(result: dict[str, Any], data: dict[str, Any]) -> None:
    name_keys = (
        "restaurant_name", "name", "business_name", "company_name",
        "location_name", "establishment",
    )
    if not result.get("restaurant_name"):
        for key in name_keys:
            value = _coerce_str(data.get(key))
            if value:
                result["restaurant_name"] = value
                break
    mapping = (
        ("street", "street"),
        ("address", "street"),
        ("location", "street"),
        ("city", "city"),
        ("state", "state"),
        ("region", "state"),
        ("province", "state"),
        ("zip", "zip"),
        ("zipcode", "zip"),
        ("postal_code", "zip"),
        ("postal", "zip"),
        ("latitude", "latitude"),
        ("lat", "latitude"),
        ("longitude", "longitude"),
        ("lon", "longitude"),
        ("lng", "longitude"),
        ("timezone", "timezone"),
        ("tz", "timezone"),
        ("currency", "currency"),
        ("phone", "phone"),
        ("email", "email"),
    )
    for source_key, target_key in mapping:
        if result.get(target_key):
            continue
        value = _coerce_str(data.get(source_key))
        if value:
            result[target_key] = value
    # Build a complete address from the component parts. A bare `address`/`location`
    # value is treated as the street line so the assembled address includes it.
    assembled = ", ".join(
        part for part in (result.get(k, "") for k in ("street", "city", "state", "zip")) if part
    )
    if assembled:
        result["address"] = assembled


def _excel_title_name(path: Path) -> str | None:
    try:
        import openpyxl

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            first_row = next(sheet.iter_rows(values_only=True), None)
            if not first_row:
                continue
            for cell in first_row:
                if isinstance(cell, str) and _looks_like_restaurant_name(cell):
                    return cell.strip()
    except Exception:
        return None
    return None


def _csv_title_name(path: Path) -> str | None:
    try:
        with path.open("r", encoding="utf-8-sig", errors="replace") as handle:
            first_line = handle.readline()
        cleaned = first_line.strip().strip("#").strip()
        if "," not in cleaned and _looks_like_restaurant_name(cleaned):
            return cleaned
    except Exception:
        return None
    return None


def _scan_folder_for_restaurant_name(source_root: Path) -> str | None:
    try:
        for path in sorted(source_root.iterdir()):
            if path.is_file() and path.suffix.lower() in {".xlsx", ".xlsm"}:
                name = _excel_title_name(path)
                if name:
                    return name
        for path in sorted(source_root.iterdir()):
            if path.is_file() and path.suffix.lower() == ".csv":
                name = _csv_title_name(path)
                if name:
                    return name
    except Exception:
        return None
    return None


def _scan_readme_for_restaurant_name(source_root: Path) -> str | None:
    for name in ("README.md", "README.txt", "ABOUT.md", "readme.md", "About.txt"):
        path = source_root / name
        if not path.is_file():
            continue
        try:
            lines = path.read_text(encoding="utf-8-sig", errors="replace").splitlines()[:40]
        except Exception:
            continue
        for line in lines:
            cleaned = line.lstrip("#*- ").strip()
            cleaned = re.split(r"\s[—-]\s", cleaned)[0].strip()
            if _looks_like_restaurant_name(cleaned):
                return cleaned
    return None


def discover_restaurant_identity(source_root: Path) -> dict[str, Any]:
    """Extract restaurant name and location from a records folder.

    Only fields actually observed in the folder are returned. Callers should
    apply the result by filling empty settings rather than overwriting existing
    values, so a manager's manual edits are never lost.
    """
    source_root = Path(source_root)
    result: dict[str, Any] = {}

    for filename in RESTAURANT_IDENTITY_FILENAMES:
        candidate = source_root / filename
        if candidate.is_file():
            try:
                data = json.loads(candidate.read_text(encoding="utf-8-sig"))
            except Exception:
                continue
            if isinstance(data, dict):
                _apply_identity_json(result, data)
                if result.get("restaurant_name"):
                    break

    if not result.get("restaurant_name") and (source_root / "restaurant_config.json").is_file():
        try:
            data = json.loads((source_root / "restaurant_config.json").read_text(encoding="utf-8-sig"))
        except Exception:
            data = None
        if isinstance(data, dict):
            value = _coerce_str(data.get("restaurant_name"))
            if value:
                result["restaurant_name"] = value
            for key in ("address", "latitude", "longitude", "timezone", "currency"):
                if not result.get(key):
                    v = _coerce_str(data.get(key))
                    if v:
                        result[key] = v

    if not result.get("restaurant_name"):
        scanned = _scan_folder_for_restaurant_name(source_root)
        if scanned:
            result["restaurant_name"] = scanned

    if not result.get("restaurant_name"):
        readme = _scan_readme_for_restaurant_name(source_root)
        if readme:
            result["restaurant_name"] = readme

    if not result.get("restaurant_name"):
        folder_name = re.sub(r"[-_]+", " ", source_root.name).strip()
        folder_name = re.sub(r"\s+", " ", folder_name)
        if folder_name:
            result["restaurant_name"] = folder_name.title()

    return result


def normalize(value: Any) -> str:
    return re.sub(r"\s+", " ", re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip())


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _windows_desktop() -> Path | None:
    """Resolve the actual Windows Desktop folder through the shell when possible."""
    try:
        buffer = ctypes.create_unicode_buffer(260)
        # CSIDL_DESKTOPDIRECTORY = 0x0010
        result = ctypes.windll.shell32.SHGetFolderPathW(None, 0x0010, None, 0, buffer)  # type: ignore[attr-defined]
        if result == 0 and buffer.value:
            return Path(buffer.value)
    except Exception:
        pass
    candidates = [
        Path(os.environ.get("OneDrive", "")) / "Desktop" if os.environ.get("OneDrive") else None,
        Path(os.environ.get("USERPROFILE", "")) / "Desktop" if os.environ.get("USERPROFILE") else None,
        Path.home() / "Desktop",
    ]
    return next((candidate for candidate in candidates if candidate and candidate.exists()), None)


def desktop_directory() -> Path:
    override = str(os.environ.get("RCC_DESKTOP_DIR") or "").strip()
    if override:
        path = Path(override).expanduser().resolve()
        path.mkdir(parents=True, exist_ok=True)
        return path
    if sys.platform.startswith("win"):
        found = _windows_desktop()
        if found:
            found.mkdir(parents=True, exist_ok=True)
            return found.resolve()
    elif sys.platform == "darwin":
        path = Path.home() / "Desktop"
        path.mkdir(parents=True, exist_ok=True)
        return path.resolve()
    else:
        try:
            result = subprocess.run(
                ["xdg-user-dir", "DESKTOP"], capture_output=True, text=True, timeout=3, check=False
            )
            candidate = Path(result.stdout.strip()).expanduser() if result.stdout.strip() else None
            if candidate:
                candidate.mkdir(parents=True, exist_ok=True)
                return candidate.resolve()
        except Exception:
            pass
    path = Path.home() / "Desktop"
    path.mkdir(parents=True, exist_ok=True)
    return path.resolve()


def _folder_name(restaurant_name: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*]+', " ", str(restaurant_name or "Restaurant"))
    cleaned = re.sub(r"\s+", " ", cleaned).strip().rstrip(".")
    return f"{cleaned or 'Restaurant'} - Auto Upload"


def _marker_matches(folder: Path, workspace: Path) -> bool:
    marker = folder / ".restaurant_workspace.json"
    if not marker.exists():
        return False
    try:
        payload = json.loads(marker.read_text(encoding="utf-8"))
        return Path(payload.get("workspace", "")).expanduser().resolve() == workspace.resolve()
    except Exception:
        return False


def ensure_auto_upload_folder(workspace: RestaurantWorkspace, restaurant_name: str | None = None) -> Path:
    """Create or recover the persistent Desktop auto-upload folder for a workspace."""
    settings = workspace.load_settings()
    configured = str(settings.get("auto_upload_folder") or "").strip()
    if configured:
        folder = Path(configured).expanduser()
        try:
            folder = folder.resolve()
            folder.mkdir(parents=True, exist_ok=True)
        except OSError:
            folder = None
        if folder is not None and folder.exists():
            _initialize_inbox(folder, workspace, restaurant_name or settings.get("restaurant_name", "Restaurant"))
            return folder

    desktop = desktop_directory()
    name = str(restaurant_name or settings.get("restaurant_name") or "Restaurant").strip()
    folder = desktop / _folder_name(name)
    if folder.exists() and not _marker_matches(folder, workspace.root):
        short = hashlib.sha256(str(workspace.root).encode("utf-8")).hexdigest()[:6].upper()
        folder = desktop / f"{_folder_name(name)} {short}"
    folder.mkdir(parents=True, exist_ok=True)
    _initialize_inbox(folder, workspace, name)
    settings["auto_upload_folder"] = str(folder.resolve())
    workspace.save_settings(settings)
    return folder.resolve()


def _initialize_inbox(folder: Path, workspace: RestaurantWorkspace, restaurant_name: str) -> None:
    for name in ("_Processing", "_Processed", "_Needs Review", "_Failed"):
        (folder / name).mkdir(parents=True, exist_ok=True)
    marker = {
        "restaurant": restaurant_name,
        "workspace": str(workspace.root),
        "created_by": "MarginMise",
    }
    marker_path = folder / ".restaurant_workspace.json"
    marker_text = json.dumps(marker, indent=2)
    try:
        if marker_path.read_text(encoding="utf-8") != marker_text:
            marker_path.write_text(marker_text, encoding="utf-8")
    except (FileNotFoundError, OSError):
        marker_path.write_text(marker_text, encoding="utf-8")
    readme = f"""{restaurant_name} AUTOMATIC UPLOAD FOLDER

DROP FILES DIRECTLY INTO THIS FOLDER.
Do not place them inside the underscore folders. The application watches the top
level while MarginMise is running. Files waiting while the app
is closed are processed the next time it starts.

The application identifies and routes:
- Invoice PDFs, images, and canonical invoice JSON
- POS item sales CSV/XLSX/XLSM
- Daily or period sales summaries
- Inventory counts
- Menu recipes
- Operating costs
- Waste logs
- Receiving logs matched to approved invoices
- Product planning/conversion updates
- Event calendars (.ics) and event CSVs
- Distributor catalogs and order confirmations
- Accounting mappings
- ZIP archives containing supported files

Files that cannot be identified safely are moved to _Needs Review without being
silently guessed. Processed originals are organized by type and month. Failed
files are preserved with a result report.

Workspace:
{workspace.root}
"""
    # Do not rewrite a Desktop file on every application launch. Security tools
    # commonly flag repeated writes/deletes in protected Desktop locations.
    # The README is optional documentation and is created only when explicitly
    # requested by the user/environment.
    readme_path = folder / "README_DROP_FILES_HERE.txt"
    if os.environ.get("MARGINMISE_CREATE_UPLOAD_README", "0") == "1" and not readme_path.exists():
        try:
            readme_path.write_text(readme, encoding="utf-8")
        except OSError:
            pass


def auto_upload_status(workspace: RestaurantWorkspace) -> dict[str, Any]:
    settings = workspace.load_settings()
    folder = ensure_auto_upload_folder(workspace, settings.get("restaurant_name", "Restaurant"))
    pending = sum(
        1 for path in folder.iterdir()
        if path.is_file() and path.name not in SYSTEM_FILES and not path.name.startswith(".")
    )
    with workspace.connect() as conn:
        conn.executescript(AUTO_UPLOAD_SCHEMA_SQL)
        recent = conn.execute(
            "SELECT * FROM auto_upload_events ORDER BY event_id DESC LIMIT 1"
        ).fetchone()
        review_count = conn.execute(
            "SELECT COUNT(*) FROM auto_upload_events WHERE status='Needs Review'"
        ).fetchone()[0]
        failed_count = conn.execute(
            "SELECT COUNT(*) FROM auto_upload_events WHERE status='Failed'"
        ).fetchone()[0]
    return {
        "enabled": bool(settings.get("auto_upload_enabled", True)),
        "folder": str(folder),
        "pending": pending,
        "needs_review": int(review_count),
        "failed": int(failed_count),
        "last_status": recent["status"] if recent else "",
        "last_summary": recent["summary"] if recent else "",
        "last_completed": recent["completed_at"] if recent else "",
    }


class AutoUploadRouter:
    def __init__(self, workspace: RestaurantWorkspace, restaurant_name: str | None = None):
        self.workspace = workspace
        self.settings = workspace.load_settings()
        self.restaurant_name = str(restaurant_name or self.settings.get("restaurant_name") or workspace.root.name)
        self.inbox = ensure_auto_upload_folder(workspace, self.restaurant_name)
        self.pipeline = InvoicePipeline(workspace)
        self.pipeline.controls.current_user = AuthenticatedUser(
            "SYSTEM-AUTO-UPLOAD", "auto_upload", "Automatic Upload Service", "Owner"
        )
        with self.workspace.connect() as conn:
            conn.executescript(AUTO_UPLOAD_SCHEMA_SQL)
        self._reconcile_historical_item_planning()
        self._recover_stale_claims()

    def _reconcile_historical_item_planning(self) -> None:
        """Upgrade products configured by Item Planning imports from older builds.

        The archived workbook is the evidence that the count conversion was
        explicitly supplied.  Invoice-inferred defaults are never approved by
        this migration.
        """
        with self.workspace.connect() as conn:
            pending = int(conn.execute(
                """SELECT COUNT(*) FROM items
                   WHERE planning_confirmed=0 AND review_status<>'Approved'"""
            ).fetchone()[0])
            if not pending:
                return
            archives = conn.execute(
                """SELECT archived_path FROM auto_upload_events
                   WHERE detected_type='Item Planning' AND status='Processed'
                     AND archived_path IS NOT NULL AND archived_path<>''
                   ORDER BY event_id"""
            ).fetchall()
        reconciled: set[str] = set()
        for event in archives:
            path = Path(str(event["archived_path"] or ""))
            if not path.exists() or path.suffix.casefold() not in TABLE_SUFFIXES:
                continue
            try:
                _headers, rows = self._read_table(path)
            except Exception:
                continue
            for row in rows:
                try:
                    item_id = self._match_item(row)
                    count_unit = str(self._row_value(row, "Count Unit", default="") or "").strip()
                    units_per = Decimal(str(
                        self._row_value(row, "Units Per Purchase Unit", default="0") or "0"
                    ))
                    if not count_unit or units_per <= 0:
                        continue
                    self.pipeline.update_item_planning(
                        item_id,
                        count_unit=count_unit,
                        units_per_purchase_unit=units_per,
                        planning_confirmed=True,
                    )
                    self.pipeline.approve_item_configuration(
                        item_id,
                        "Product planning approved from prior structured Item Planning import",
                    )
                    reconciled.add(item_id)
                except Exception:
                    continue
        if reconciled:
            try:
                self.pipeline.controls.audit(
                    "items.planning_reconcile",
                    "item",
                    "",
                    f"Confirmed {len(reconciled)} product planning record(s) from archived structured imports",
                    details={"item_ids": sorted(reconciled)},
                )
            except Exception:
                pass

    def _recover_stale_claims(self) -> None:
        processing = self.inbox / "_Processing"
        threshold = time.time() - 600
        for path in processing.iterdir():
            if not path.is_file():
                continue
            try:
                if path.stat().st_mtime < threshold:
                    destination = self._unique_path(self.inbox / path.name)
                    shutil.move(str(path), str(destination))
            except OSError:
                continue

    @staticmethod
    def _unique_path(path: Path) -> Path:
        if not path.exists():
            return path
        for index in range(1, 10000):
            candidate = path.with_name(f"{path.stem}_{index}{path.suffix}")
            if not candidate.exists():
                return candidate
        return path.with_name(f"{path.stem}_{uuid.uuid4().hex[:8]}{path.suffix}")

    def _claim(self, source: Path) -> Path:
        processing = self.inbox / "_Processing"
        processing.mkdir(parents=True, exist_ok=True)
        claimed = self._unique_path(processing / source.name)
        try:
            os.replace(source, claimed)
        except FileNotFoundError as exc:
            raise AutoUploadRetryLater("The file was claimed by another watcher.") from exc
        except PermissionError as exc:
            raise AutoUploadRetryLater("The file is still being copied or is open in another program.") from exc
        return claimed

    def _header_signature_score(self, headers: Iterable[str]) -> int:
        """Score a header row by how safely it maps to an existing importer."""
        normalized = {normalize(header) for header in headers if str(header or "").strip()}

        def has(*names: str) -> bool:
            return any(normalize(name) in normalized for name in names)

        if (
            has("vendor", "vendor name")
            and has("invoice number", "invoice no", "invoice #")
            and has("invoice date")
            and has("item description", "product description", "description")
            and has("quantity", "qty")
            and has("unit price", "price")
            and has("line total", "extended amount", "extension")
        ):
            return 130
        if has("mapping key") and has("debit account") and has("credit account"):
            return 120
        if has("menu item name") and has("quantity count units") and (
            has("inventory item id") or has("vendor sku") or has("inventory item name")
        ):
            return 120
        if (has("counted quantity") or has("quantity on hand") or has("ending quantity")) and (
            has("item id") or has("inventory item id") or has("vendor sku") or has("item name")
        ):
            return 120
        if has("inventory item id") and has("count unit") and has("units per purchase unit") and (
            has("lead time days") or has("order cycle days") or has("safety stock days")
        ):
            return 120
        if (has("event date") or has("waste date") or has("date")) and has("item name") and (
            has("quantity count units") or has("quantity")
        ) and has("count unit") and has("reason"):
            return 120
        if (has("date") or has("cost date")) and has("category") and has("description") and has("amount"):
            return 120
        if has("receiving date") and has("vendor") and has("invoice number") and has("received qty", "received quantity"):
            return 119
        if has("event name") and (has("start date") or has("event date")):
            return 115
        if has("period start") and has("net sales"):
            return 120
        if has("net sales") and has("date") and not (has("quantity") or has("units sold")):
            return 115
        if has("menu item name") and any(
            has(alias) for alias in ["menu price", "category", "active", "pos item key"]
        ):
            return 110
        if has("sku", "vendor sku", "distributor sku") and has(
            "description", "item", "product"
        ) and has("price", "unit price"):
            return 105
        mapping = self.pipeline.phase2.suggest_mapping(headers)
        if all(mapping.get(field) for field in ("business_date", "menu_item_name", "quantity")):
            return 118
        return 0

    def _read_excel_table(self, path: Path) -> tuple[list[str], list[dict[str, Any]]]:
        """Find import-ready worksheets and combine matching sheets.

        Restaurant workbooks commonly contain a cover/summary sheet before the
        actual import table. Inventory workbooks also commonly split beginning
        and ending counts across two sheets. Selecting only the active sheet
        silently lost or misclassified those files.
        """
        try:
            from openpyxl import load_workbook
        except ImportError as exc:
            raise AutoUploadError("openpyxl is required to read Excel uploads") from exc

        workbook = load_workbook(path, read_only=True, data_only=True)
        candidates: list[tuple[int, int, list[str], Any]] = []
        fallback: tuple[int, list[str], Any] | None = None
        try:
            for sheet_index, worksheet in enumerate(workbook.worksheets):
                best: tuple[int, int, list[str]] | None = None
                for row_number, values in enumerate(
                    worksheet.iter_rows(min_row=1, max_row=25, values_only=True), start=1
                ):
                    headers = [str(value or "").strip() for value in values]
                    nonempty = [header for header in headers if header]
                    if not nonempty:
                        continue
                    unique_count = len({normalize(header) for header in nonempty})
                    if fallback is None and unique_count >= 2:
                        fallback = (row_number, headers, worksheet)
                    score = self._header_signature_score(headers)
                    candidate = (score, unique_count, headers)
                    if best is None or candidate[:2] > best[:2]:
                        best = candidate
                        best_row = row_number
                if best and best[0] > 0:
                    candidates.append((best[0], best_row, best[2], worksheet))

            if candidates:
                candidates.sort(key=lambda value: (-value[0], workbook.worksheets.index(value[3])))
                best_score, _, best_headers, _ = candidates[0]
                signature = tuple(normalize(header) for header in best_headers)
                selected = [
                    candidate for candidate in candidates
                    if candidate[0] == best_score
                    and tuple(normalize(header) for header in candidate[2]) == signature
                ]
            elif fallback:
                row_number, headers, worksheet = fallback
                selected = [(0, row_number, headers, worksheet)]
            else:
                return [], []

            headers = selected[0][2]
            rows: list[dict[str, Any]] = []
            for _, header_row, sheet_headers, worksheet in selected:
                for source_row, values in enumerate(
                    worksheet.iter_rows(min_row=header_row + 1, values_only=True),
                    start=header_row + 1,
                ):
                    if not any(value not in (None, "") for value in values):
                        continue
                    row_data = {
                        headers[index]: values[index] if index < len(values) else ""
                        for index in range(len(headers))
                        if headers[index]
                    }
                    row_data["__source_sheet"] = worksheet.title
                    row_data["__source_row"] = source_row
                    rows.append(row_data)
            return [header for header in headers if header], rows
        finally:
            workbook.close()

    def _read_table(self, path: Path) -> tuple[list[str], list[dict[str, Any]]]:
        suffix = path.suffix.lower()
        if suffix in {".csv", ".tsv", ".txt"}:
            with path.open("r", encoding="utf-8-sig", newline="", errors="replace") as handle:
                sample = handle.read(8192)
                handle.seek(0)
                try:
                    dialect = csv.Sniffer().sniff(sample, delimiters=",\t;|")
                except csv.Error:
                    dialect = csv.excel_tab if suffix == ".tsv" else csv.excel
                reader = csv.DictReader(handle, dialect=dialect)
                headers = [str(value or "").strip() for value in (reader.fieldnames or [])]
                rows = []
                for row_number, row in enumerate(reader, 2):
                    row_data = dict(row)
                    row_data["__source_sheet"] = path.name
                    row_data["__source_row"] = row_number
                    rows.append(row_data)
                return headers, rows
        if suffix in {".xlsx", ".xlsm"}:
            try:
                return self._read_excel_table(path)
            except Exception as exc:
                raise AutoUploadError(f"Unable to read Excel file: {exc}") from exc
        raise AutoUploadError(f"Unsupported table type: {suffix}")

    def _materialize_table(self, source: Path, headers: list[str], rows: list[dict[str, Any]]) -> tuple[Path, bool]:
        if source.suffix.lower() == ".csv":
            return source, False
        temp_dir = self.workspace.folders["logs"] / "Auto Upload Temp"
        temp_dir.mkdir(parents=True, exist_ok=True)
        temp_path = temp_dir / f"{safe_filename(source.stem)}.csv"
        with temp_path.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=headers, extrasaction="ignore")
            writer.writeheader()
            for row in rows:
                writer.writerow({
                    key: (
                        value.date().isoformat()
                        if isinstance(value, datetime)
                        else value.isoformat()
                        if isinstance(value, date)
                        else value
                    )
                    for key, value in row.items()
                })
        return temp_path, True

    def classify(self, path: Path) -> Classification:
        suffix = path.suffix.lower()
        if suffix in INVOICE_SUFFIXES:
            return Classification("Invoice", 1.0, f"{suffix.upper()} document routed to invoice extraction")
        if suffix == ".ics":
            return Classification("Event Calendar", 1.0, "iCalendar file")
        if suffix == ".zip":
            return Classification("Archive", 1.0, "ZIP archive containing upload files")
        if suffix == ".json":
            try:
                payload = json.loads(path.read_text(encoding="utf-8-sig"))
            except Exception:
                return Classification("Unclassified", 0.0, "JSON could not be parsed")
            if isinstance(payload, dict):
                keys = {normalize(key) for key in payload.keys()}
                if {"vendor", "invoice number"}.issubset(keys) or (
                    "items" in keys and ("total" in keys or "subtotal" in keys)
                ):
                    return Classification("Invoice", 0.98, "Canonical invoice fields found in JSON")
            return Classification("Unclassified", 0.2, "JSON is not a recognized invoice payload")
        if suffix not in TABLE_SUFFIXES:
            return Classification("Unsupported", 0.0, f"Unsupported file extension {suffix or '(none)'}")

        try:
            headers, rows = self._read_table(path)
        except Exception as exc:
            return Classification("Unclassified", 0.0, f"Table could not be read: {exc}")
        normalized = {normalize(header) for header in headers if str(header).strip()}
        filename = normalize(path.stem)

        def has(*names: str) -> bool:
            return any(normalize(name) in normalized for name in names)

        def contains(text: str) -> bool:
            target = normalize(text)
            return any(target in header for header in normalized)

        result: tuple[str, float, str] | None = None
        structured_invoice = (
            has("vendor", "vendor name")
            and has("invoice number", "invoice no", "invoice #")
            and has("invoice date")
            and has("item description", "product description", "description")
            and has("quantity", "qty")
            and has("unit price", "price")
            and has("line total", "extended amount", "extension")
        )
        likely_invoice = (
            has("invoice number", "invoice no", "invoice #")
            and has("vendor", "vendor name")
            and (
                has("invoice date")
                or has("subtotal")
                or has("total", "invoice total")
                or contains("line total")
            )
        )
        if structured_invoice:
            result = ("Invoice", 1.0, "Structured invoice header and line-item columns detected")
        elif has("mapping key") and has("debit account") and has("credit account"):
            result = ("Accounting Mappings", 1.0, "Accounting mapping columns detected")
        elif has("receiving date") and has("vendor") and has("invoice number") and has("received qty", "received quantity"):
            result = (
                "Receiving Log",
                1.0,
                "Receiving record columns detected; controlled receiving records require manager review",
            )
        elif has("menu item name") and has("quantity count units") and (
            has("inventory item id") or has("vendor sku") or has("inventory item name")
        ):
            result = ("Recipes", 1.0, "Menu recipe and ingredient columns detected")
        elif has("menu item name") and any(
            has(alias) for alias in ["menu price", "category", "active", "pos item key"]
        ) and not any(
            has(neg) for neg in ["receipt_id", "business_date", "payment_type", "gross_sales", "net_sales"]
        ):
            result = ("Menu Items", 1.0, "Menu item master columns detected")
        elif (has("counted quantity") or has("quantity on hand") or has("ending quantity")) and (
            has("item id") or has("inventory item id") or has("vendor sku") or has("item name")
        ):
            result = ("Inventory Count", 1.0, "Physical inventory count columns detected")
        elif has("inventory item id") and has("count unit") and has("units per purchase unit") and (
            has("lead time days") or has("order cycle days") or has("safety stock days")
        ):
            result = ("Item Planning", 1.0, "Product conversion and planning columns detected")
        elif (has("event date") or has("waste date") or has("date")) and has("item name") and has("quantity count units") and has("count unit") and has("reason"):
            result = ("Waste Log", 1.0, "Waste event columns detected")
        elif (has("date") or has("cost date")) and has("description") and has("amount"):
            result = ("Operating Costs", 1.0, "Operating-cost columns detected")
        elif has("event name") and (has("start date") or has("event date")):
            result = ("Event Calendar", 0.98, "Event name and date columns detected")
        elif has("period start") and has("net sales"):
            result = ("Sales Summary", 1.0, "Period sales summary columns detected")
        elif has("net sales") and has("date") and not (has("quantity") or has("units sold")):
            result = ("Sales Summary", 0.95, "Daily sales summary columns detected")
        elif (has("date") or has("cost date")) and has("category") and has("description") and has("amount"):
            result = ("Operating Costs", 1.0, "Operating-cost columns detected")
        elif has("sku") and has("description") and (has("price") or has("unit price")) and not (
            has("date") and (has("quantity") or has("units sold"))
        ):
            result = ("Distributor Catalog", 0.94, "Distributor SKU, description, and price columns detected")
        else:
            mapping = self.pipeline.phase2.suggest_mapping(headers)
            if likely_invoice:
                result = (
                    "Invoice",
                    0.55,
                    "Invoice-like columns were found, but the structured invoice signature is incomplete; file was not treated as POS sales",
                )
            elif all(mapping.get(field) for field in ("business_date", "menu_item_name", "quantity")):
                result = ("POS Sales", 0.96, "Item-level sales date, product, and quantity columns detected")

        if not result:
            # Filename hints only raise a file to review. They never silently choose a parser.
            hints = {
                "recipe": "Recipes", "inventory count": "Inventory Count", "waste": "Waste Log",
                "operating cost": "Operating Costs", "sales": "Sales Summary",
                "catalog": "Distributor Catalog", "confirmation": "Distributor Confirmation",
            }
            hinted = next((kind for token, kind in hints.items() if token in filename), "Unclassified")
            confidence = 0.45 if hinted != "Unclassified" else 0.0
            reason = f"Filename suggests {hinted}, but required columns were not found" if confidence else "No supported data signature was found"
            return Classification(hinted, confidence, reason, headers, len(rows))
        return Classification(result[0], result[1], result[2], headers, len(rows))

    def _prior_success(self, digest: str) -> Any:
        with self.workspace.connect() as conn:
            return conn.execute(
                """SELECT * FROM auto_upload_events
                   WHERE source_sha256=? AND status IN ('Processed','Duplicate')
                   ORDER BY event_id DESC LIMIT 1""",
                (digest,),
            ).fetchone()

    def _record_event(
        self,
        *,
        digest: str,
        original_name: str,
        original_path: Path,
        classification: Classification,
        outcome: RoutingOutcome,
        archived_path: Path,
    ) -> None:
        completed = now_iso()
        with self.workspace.connect() as conn:
            conn.execute(
                """INSERT INTO auto_upload_events(
                       source_sha256,original_name,original_path,detected_type,
                       classification_confidence,status,summary,details_json,
                       archived_path,created_at,completed_at)
                   VALUES(?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    digest, original_name, str(original_path), outcome.detected_type,
                    classification.confidence, outcome.status, outcome.summary,
                    json.dumps({"classification": asdict(classification), "outcome": outcome.as_dict()}, default=str),
                    str(archived_path), completed, completed,
                ),
            )
        self._write_activity_csv()
        try:
            self.pipeline.controls.audit(
                "auto_upload.process",
                "file",
                digest[:16],
                f"{outcome.status}: {original_name} routed as {outcome.detected_type}",
                details={"summary": outcome.summary, "archived_path": str(archived_path)},
            )
        except Exception:
            pass

    def _write_activity_csv(self) -> None:
        with self.workspace.connect() as conn:
            rows = conn.execute(
                """SELECT completed_at,original_name,detected_type,status,summary,archived_path
                   FROM auto_upload_events ORDER BY event_id DESC LIMIT 250"""
            ).fetchall()
        destination = self.inbox / "AUTO_UPLOAD_ACTIVITY.csv"
        with destination.open("w", encoding="utf-8-sig", newline="") as handle:
            writer = csv.writer(handle)
            writer.writerow(["Completed At", "Original File", "Detected Type", "Status", "Summary", "Archived Path"])
            for row in rows:
                writer.writerow([row[key] for key in row.keys()])

    def _archive(self, claimed: Path, outcome: RoutingOutcome, original_name: str) -> Path:
        top = {
            "Processed": "_Processed",
            "Duplicate": "_Processed",
            "Needs Review": "_Needs Review",
            "Failed": "_Failed",
        }.get(outcome.status, "_Failed")
        subtype = "Duplicates" if outcome.status == "Duplicate" else safe_filename(outcome.detected_type)
        month = datetime.now().strftime("%Y-%m")
        destination_dir = self.inbox / top / subtype / month
        destination_dir.mkdir(parents=True, exist_ok=True)
        destination = self._unique_path(destination_dir / original_name)
        try:
            shutil.move(str(claimed), str(destination))
        except OSError:
            shutil.copy2(claimed, destination)
            claimed.unlink(missing_ok=True)
        result_path = destination.with_name(destination.name + ".result.json")
        result_path.write_text(json.dumps(outcome.as_dict(), indent=2, default=str), encoding="utf-8")
        return destination

    def _match_item(self, row: dict[str, Any]) -> str:
        item_id = str(row.get("Inventory Item ID") or row.get("Item ID") or row.get("item_id") or "").strip()
        sku = str(row.get("Vendor SKU") or row.get("SKU") or row.get("vendor_sku") or "").strip()
        name = str(row.get("Inventory Item Name") or row.get("Item Name") or row.get("item_name") or "").strip()
        with self.workspace.connect() as conn:
            if item_id:
                found = conn.execute("SELECT item_id FROM items WHERE item_id=?", (item_id,)).fetchone()
                if found:
                    return found["item_id"]
            if sku:
                found = conn.execute(
                    "SELECT item_id FROM items WHERE vendor_sku=? COLLATE NOCASE ORDER BY last_purchase_date DESC LIMIT 1",
                    (sku,),
                ).fetchone()
                if found:
                    return found["item_id"]
            if name:
                found = conn.execute(
                    "SELECT item_id FROM items WHERE item_name=? COLLATE NOCASE OR item_name LIKE ? ORDER BY last_purchase_date DESC LIMIT 1",
                    (name, f"%{name}%"),
                ).fetchone()
                if found:
                    return found["item_id"]
        raise AutoUploadError(f"Inventory item could not be matched: {item_id or sku or name or 'blank item'}")

    @staticmethod
    def _row_value(row: dict[str, Any], *names: str, default: Any = "") -> Any:
        normalized = {normalize(key): value for key, value in row.items()}
        for name in names:
            key = normalize(name)
            if key in normalized and normalized[key] not in (None, ""):
                return normalized[key]
        return default

    def _process_waste(self, rows: list[dict[str, Any]], digest: str) -> RoutingOutcome:
        imported = 0
        errors: list[str] = []
        for index, row in enumerate(rows, 2):
            try:
                item_id = self._match_item(row)
                event_date = str(self._row_value(row, "Event Date", "Waste Date", "Date")).strip()
                quantity = self._row_value(row, "Quantity Count Units", "Quantity", "Amount")
                reason = str(self._row_value(row, "Reason", "Waste Reason")).strip()
                if not event_date or not reason:
                    raise AutoUploadError("Event Date and Reason are required")
                marker = f"AUTOUPLOAD:{digest[:16]}:ROW:{index}"
                with self.workspace.connect() as conn:
                    exists = conn.execute("SELECT waste_id FROM waste_events WHERE notes LIKE ?", (f"%{marker}%",)).fetchone()
                if exists:
                    continue
                notes = str(self._row_value(row, "Notes", default="")).strip()
                notes = f"{notes} {marker}".strip()
                self.pipeline.log_waste(
                    item_id,
                    quantity,
                    reason,
                    event_date=event_date,
                    shift=str(self._row_value(row, "Shift", default="")),
                    notes=notes,
                    created_by="auto_upload",
                )
                imported += 1
            except Exception as exc:
                errors.append(f"Row {index}: {exc}")
        status = "Processed" if imported and not errors else "Needs Review" if imported or errors else "Failed"
        return RoutingOutcome(status, "Waste Log", f"Logged {imported} waste event(s); {len(errors)} row(s) need review.", imported, len(errors), {"errors": errors[:100]})

    def _process_item_planning(self, rows: list[dict[str, Any]]) -> RoutingOutcome:
        imported = 0
        errors: list[str] = []
        approved_item_ids: list[str] = []
        for index, row in enumerate(rows, 2):
            try:
                item_id = self._match_item(row)
                count_unit = str(self._row_value(row, "Count Unit", default="") or "").strip()
                units_per = self._row_value(row, "Units Per Purchase Unit", default=None)
                if not count_unit:
                    raise AutoUploadError("Count Unit is required")
                try:
                    if Decimal(str(units_per)) <= 0:
                        raise ValueError
                except (TypeError, ValueError, ArithmeticError):
                    raise AutoUploadError("Units Per Purchase Unit must be greater than zero")
                values = {
                    "count_unit": count_unit,
                    "units_per_purchase_unit": units_per,
                    "lead_time_days": self._row_value(row, "Lead Time Days", default=None),
                    "order_cycle_days": self._row_value(row, "Order Cycle Days", default=None),
                    "safety_stock_days": self._row_value(row, "Safety Stock Days", default=None),
                    "order_multiple": self._row_value(row, "Order Multiple", default=None),
                    "minimum_order_qty": self._row_value(row, "Minimum Order Qty", default=None),
                    "par_override_count_units": self._row_value(row, "Par Override Count Units", default=None),
                    "planning_confirmed": True,
                }
                values = {key: value for key, value in values.items() if value not in (None, "")}
                active = self._row_value(row, "Active", default="")
                if active not in (None, ""):
                    values["active"] = str(active).strip().lower() not in {"0", "false", "no"}
                self.pipeline.update_item_planning(item_id, **values)
                self.pipeline.approve_item_configuration(
                    item_id,
                    "Product planning approved by structured Item Planning import",
                )
                approved_item_ids.append(item_id)
                imported += 1
            except Exception as exc:
                errors.append(f"Row {index}: {exc}")
        status = "Processed" if imported and not errors else "Needs Review" if imported or errors else "Failed"
        return RoutingOutcome(
            status,
            "Item Planning",
            f"Updated and approved {imported} product planning record(s); {len(errors)} row(s) need review.",
            imported,
            len(errors),
            {"errors": errors[:100], "approved_item_ids": approved_item_ids},
        )

    def _process_menu_items(self, rows: list[dict[str, Any]], digest: str) -> RoutingOutcome:
        imported = 0
        errors: list[str] = []
        marker = f"AUTOUPLOAD:{digest[:16]}"
        try:
            with self.workspace.connect() as conn:
                for row in rows:
                    menu_item_id = str(row.get("Menu Item ID") or row.get("menu_item_id") or "").strip()
                    pos_item_key = str(row.get("POS Item Key") or row.get("pos_item_key") or "").strip()
                    menu_item_name = str(row.get("Menu Item Name") or row.get("menu_item_name") or "").strip()
                    menu_price = row.get("Menu Price") or row.get("menu_price") or ""
                    category = canonical_inventory_category(
        row.get("Menu Category") or row.get("category"),
        row.get("Menu Item Name") or row.get("menu item name") or row.get("Item Name") or row.get("item_name") or "",
        row.get("Unit") or row.get("unit") or "",
    )
                    active_raw = str(row.get("Active") or row.get("active") or "1").strip().lower()
                    active = 0 if active_raw in {"0", "false", "no", "inactive"} else 1
                    if not menu_item_id or not pos_item_key or not menu_item_name:
                        continue
                    try:
                        price_val = float(str(menu_price).replace("$", "").replace(",", "")) if menu_price != "" else None
                    except Exception:
                        price_val = None
                    conn.execute(
                        """INSERT INTO menu_items(menu_item_id,pos_item_key,menu_item_name,category,menu_price,active,created_at,updated_at)
                           VALUES(?,?,?,?,?,?,?,?) ON CONFLICT(menu_item_id) DO UPDATE SET
                           pos_item_key=excluded.pos_item_key,
                           menu_item_name=excluded.menu_item_name,
                           category=excluded.category,
                           menu_price=excluded.menu_price,
                           active=excluded.active,
                           updated_at=excluded.updated_at""",
                        (menu_item_id, pos_item_key, menu_item_name, category, f"{price_val:.2f}" if isinstance(price_val, float) else "0.00", active, now_iso(), now_iso()),
                    )
                    imported += 1
        except Exception as exc:
            errors.append(str(exc))
        status = "Processed" if imported and not errors else "Needs Review" if imported or errors else "Failed"
        return RoutingOutcome(status, "Menu Items", f"Imported {imported} menu record(s); {len(errors)} row(s) need review.", imported, len(errors), {"errors": errors[:100]})

    def _process_accounting_mappings(self, rows: list[dict[str, Any]]) -> RoutingOutcome:
        imported = 0
        errors: list[str] = []
        for index, row in enumerate(rows, 2):
            try:
                key = str(self._row_value(row, "Mapping Key")).strip()
                debit = str(self._row_value(row, "Debit Account")).strip()
                credit = str(self._row_value(row, "Credit Account")).strip()
                if not all((key, debit, credit)):
                    raise AutoUploadError("Mapping Key, Debit Account, and Credit Account are required")
                self.pipeline.phase2.set_accounting_mapping(key, debit, credit)
                imported += 1
            except Exception as exc:
                errors.append(f"Row {index}: {exc}")
        status = "Processed" if imported and not errors else "Needs Review" if imported or errors else "Failed"
        return RoutingOutcome(status, "Accounting Mappings", f"Applied {imported} accounting mapping(s); {len(errors)} row(s) need review.", imported, len(errors), {"errors": errors[:100]})

    def _process_event_rows(self, rows: list[dict[str, Any]]) -> RoutingOutcome:
        imported = 0
        errors: list[str] = []
        for index, row in enumerate(rows, 2):
            try:
                name = str(self._row_value(row, "Event Name", "Name")).strip()
                start = str(self._row_value(row, "Start Date", "Event Date", "Date")).strip()
                end = str(self._row_value(row, "End Date", default=start)).strip() or start
                if not name or not start:
                    raise AutoUploadError("Event Name and Start Date are required")
                self.pipeline.add_local_event(
                    name,
                    start,
                    end_date=end,
                    category=str(self._row_value(row, "Category", default="Imported")),
                    impact_percent=self._row_value(row, "Expected Sales Impact Percent", "Impact Percent", default="0"),
                    notes=str(self._row_value(row, "Notes", default="")),
                    source="Auto Upload",
                    external_uid=str(self._row_value(row, "Event UID", "UID", default="")) or None,
                )
                imported += 1
            except Exception as exc:
                errors.append(f"Row {index}: {exc}")
        status = "Processed" if imported and not errors else "Needs Review" if imported or errors else "Failed"
        return RoutingOutcome(status, "Event Calendar", f"Configured {imported} event(s); {len(errors)} row(s) need review.", imported, len(errors), {"errors": errors[:100]})

    def _process_receiving_rows(self, rows: list[dict[str, Any]], digest: str) -> RoutingOutcome:
        """Load receiving sheets through the existing controlled workflow.

        A delivery is imported only when it maps unambiguously to one approved
        invoice and covers every invoice line. Existing finalized receiving
        records are never overwritten by a later file.
        """
        grouped: dict[tuple[str, str], list[dict[str, Any]]] = {}
        for row in rows:
            vendor = str(self._row_value(row, "Vendor")).strip()
            invoice_number = str(self._row_value(row, "Invoice Number")).strip()
            grouped.setdefault((normalize(vendor), invoice_number.casefold()), []).append(row)

        imported_rows = 0
        imported_deliveries = 0
        discrepancy_deliveries = 0
        errors: list[str] = []

        def quantity(value: Any) -> Decimal:
            text = str(value if value not in (None, "") else "0").replace(",", "").strip()
            return Decimal(text)

        for (_, invoice_number_key), delivery_rows in grouped.items():
            display_number = str(self._row_value(delivery_rows[0], "Invoice Number")).strip()
            display_vendor = str(self._row_value(delivery_rows[0], "Vendor")).strip()
            try:
                if not display_number or not display_vendor:
                    raise AutoUploadError("Vendor and Invoice Number are required")
                with self.workspace.connect() as conn:
                    invoice_candidates = conn.execute(
                        """SELECT * FROM invoices
                           WHERE LOWER(invoice_number)=? AND status='Approved'
                           ORDER BY invoice_date DESC""",
                        (invoice_number_key,),
                    ).fetchall()
                vendor_key = normalize(display_vendor)
                matching = [
                    row for row in invoice_candidates
                    if vendor_key == normalize(row["vendor"])
                    or vendor_key in normalize(row["vendor"])
                    or normalize(row["vendor"]) in vendor_key
                ]
                if len(matching) != 1:
                    if not matching:
                        raise AutoUploadError(
                            f"no approved invoice matched {display_vendor} {display_number}"
                        )
                    raise AutoUploadError(
                        f"more than one approved invoice matched {display_vendor} {display_number}"
                    )
                invoice = matching[0]
                with self.workspace.connect() as conn:
                    existing = conn.execute(
                        "SELECT status FROM receiving_sessions WHERE invoice_id=?",
                        (invoice["invoice_id"],),
                    ).fetchone()
                    invoice_lines = conn.execute(
                        "SELECT * FROM invoice_lines WHERE invoice_id=? ORDER BY line_number",
                        (invoice["invoice_id"],),
                    ).fetchall()
                if existing and existing["status"] in {"Verified", "Needs Review"}:
                    raise AutoUploadError(
                        f"receiving is already finalized as {existing['status']}; open it in Receiving to make changes"
                    )
                if not invoice_lines:
                    raise AutoUploadError("the approved invoice has no line items")

                unmatched = list(invoice_lines)
                matched_rows: list[tuple[dict[str, Any], Any]] = []
                for supplied in delivery_rows:
                    sku = str(self._row_value(supplied, "Vendor SKU", "SKU")).strip().casefold()
                    item_name = normalize(self._row_value(supplied, "Item", "Item Name", "Description"))
                    choices = [
                        line for line in unmatched
                        if sku and str(line["vendor_sku"] or "").strip().casefold() == sku
                    ]
                    if not choices and item_name:
                        choices = [
                            line for line in unmatched
                            if item_name == normalize(line["description"])
                        ]
                    if len(choices) != 1:
                        raise AutoUploadError(
                            f"line could not be matched uniquely: {sku or item_name or 'blank item'}"
                        )
                    line = choices[0]
                    expected_sheet = quantity(self._row_value(supplied, "Ordered Qty", "Expected Quantity"))
                    expected_invoice = quantity(line["quantity"])
                    if expected_sheet != expected_invoice:
                        raise AutoUploadError(
                            f"ordered quantity for {line['description']} is {expected_sheet} in the receiving file "
                            f"but {expected_invoice} on the invoice"
                        )
                    unmatched.remove(line)
                    matched_rows.append((supplied, line))
                if unmatched:
                    raise AutoUploadError(
                        f"{len(unmatched)} invoice line(s) are missing from the receiving file"
                    )

                session_id = self.pipeline.start_receiving(invoice["invoice_id"])
                _, receiving_lines = self.pipeline.get_receiving(session_id)
                receiving_by_invoice_line = {
                    int(line["invoice_line_id"]): line for line in receiving_lines
                }
                payload: list[dict[str, Any]] = []
                received_dates: set[str] = set()
                marker = f"AUTOUPLOAD:{digest[:16]}"
                for supplied, invoice_line in matched_rows:
                    receiving_line = receiving_by_invoice_line[int(invoice_line["line_id"])]
                    raw_date = self._row_value(supplied, "Receiving Date", "Received Date", "Date")
                    if isinstance(raw_date, datetime):
                        received_dates.add(raw_date.date().isoformat())
                    elif isinstance(raw_date, date):
                        received_dates.add(raw_date.isoformat())
                    else:
                        received_dates.add(parse_date(raw_date))
                    received = quantity(self._row_value(supplied, "Received Qty", "Received Quantity"))
                    raw_status = normalize(self._row_value(supplied, "Status"))
                    if "damage" in raw_status:
                        line_status = "Damaged"
                    elif "reject" in raw_status and received == 0:
                        line_status = "Rejected"
                    elif "not received" in raw_status:
                        line_status = "Not Received"
                    elif "substitut" in raw_status:
                        line_status = "Substituted"
                    elif "short" in raw_status or received < quantity(invoice_line["quantity"]):
                        line_status = "Short"
                    else:
                        line_status = "Received"
                    notes = str(self._row_value(supplied, "Notes")).strip()
                    payload.append({
                        "receiving_line_id": receiving_line["receiving_line_id"],
                        "received_quantity": str(received),
                        "line_status": line_status,
                        "credit_expected": "0.00",
                        "substitution_description": "",
                        "notes": f"{notes} {marker}".strip(),
                    })
                if len(received_dates) != 1:
                    raise AutoUploadError("all lines for one invoice must use the same Receiving Date")
                result = self.pipeline.save_receiving(
                    session_id,
                    payload,
                    received_date=next(iter(received_dates)),
                    notes=f"Imported from Auto Upload file. {marker}",
                    finalize=True,
                )
                imported_rows += len(payload)
                imported_deliveries += 1
                if result.get("status") == "Needs Review":
                    discrepancy_deliveries += 1
            except Exception as exc:
                errors.append(f"{display_vendor or 'Unknown vendor'} {display_number or 'Unknown invoice'}: {exc}")

        rejected = len(errors)
        status = (
            "Processed"
            if imported_deliveries and not errors and not discrepancy_deliveries
            else "Needs Review"
            if imported_deliveries or errors or discrepancy_deliveries
            else "Failed"
        )
        summary = (
            f"Imported {imported_rows} receiving line(s) across {imported_deliveries} delivery record(s); "
            f"{discrepancy_deliveries} delivery record(s) and {rejected} unmatched group(s) need review."
        )
        return RoutingOutcome(
            status,
            "Receiving Log",
            summary,
            imported_rows,
            rejected,
            {
                "deliveries_imported": imported_deliveries,
                "discrepancy_deliveries": discrepancy_deliveries,
                "errors": errors[:100],
            },
        )

    def _resolve_distributor(self, path: Path, rows: list[dict[str, Any]], confirmation: bool = False) -> Any:
        profiles = list(self.pipeline.phase3.list_distributors())
        if not profiles:
            raise AutoUploadError("No distributor profiles are configured. Open Integrations and configure the distributor first.")
        if confirmation:
            po_number = ""
            for row in rows:
                po_number = str(self._row_value(row, "PO Number", "PO", "po_id")).strip()
                if po_number:
                    break
            if po_number:
                with self.workspace.connect() as conn:
                    po = conn.execute("SELECT vendor_name FROM purchase_orders WHERE po_id=?", (po_number,)).fetchone()
                if po:
                    vendor = normalize(po["vendor_name"])
                    matched = [profile for profile in profiles if normalize(profile["vendor_name_match"] or profile["distributor_name"]) in vendor]
                    if len(matched) == 1:
                        return matched[0]
        file_key = normalize(path.stem)
        matched = []
        for profile in profiles:
            distributor_key = normalize(profile["distributor_name"])
            vendor_key = normalize(profile["vendor_name_match"])
            if (distributor_key and distributor_key in file_key) or (vendor_key and vendor_key in file_key):
                matched.append(profile)
        if len(matched) == 1:
            return matched[0]
        skus = {
            str(self._row_value(row, "SKU", "Vendor SKU", "Distributor SKU")).strip().lower()
            for row in rows[:200]
            if str(self._row_value(row, "SKU", "Vendor SKU", "Distributor SKU")).strip()
        }
        scores: list[tuple[int, Any]] = []
        with self.workspace.connect() as conn:
            for profile in profiles:
                vendor_match = str(profile["vendor_name_match"] or profile["distributor_name"])
                vendor_skus = {
                    str(row[0]).strip().lower()
                    for row in conn.execute(
                        "SELECT vendor_sku FROM items WHERE vendor_name LIKE ? AND vendor_sku IS NOT NULL",
                        (f"%{vendor_match}%",),
                    ).fetchall()
                }
                scores.append((len(skus & vendor_skus), profile))
        scores.sort(key=lambda item: item[0], reverse=True)
        if scores and scores[0][0] > 0 and (len(scores) == 1 or scores[0][0] > scores[1][0]):
            return scores[0][1]
        raise AutoUploadError("The distributor could not be identified safely from the filename, PO, or product SKUs.")

    def _extract_zip(self, path: Path) -> RoutingOutcome:
        extracted = 0
        skipped = 0
        total_size = 0
        with zipfile.ZipFile(path, "r") as archive:
            members = [member for member in archive.infolist() if not member.is_dir()]
            if len(members) > 250:
                raise AutoUploadError("ZIP archive contains more than 250 files.")
            for member in members:
                total_size += int(member.file_size)
                if total_size > 500 * 1024 * 1024:
                    raise AutoUploadError("ZIP archive expands beyond the 500 MB safety limit.")
                name = Path(member.filename).name
                suffix = Path(name).suffix.lower()
                if not name or suffix not in SUPPORTED_SUFFIXES or suffix == ".zip":
                    skipped += 1
                    continue
                destination = self._unique_path(self.inbox / f"{safe_filename(path.stem)}__{safe_filename(name)}")
                with archive.open(member) as source, destination.open("wb") as target:
                    shutil.copyfileobj(source, target)
                extracted += 1
        status = "Processed" if extracted else "Needs Review"
        return RoutingOutcome(status, "Archive", f"Extracted {extracted} supported file(s); skipped {skipped} unsupported member(s).", extracted, skipped, {"extracted": extracted, "skipped": skipped})

    @staticmethod
    def _invoice_cell_decimal(value: Any, field_name: str) -> Decimal:
        text = str(value if value not in (None, "") else "").strip()
        if not text:
            raise AutoUploadError(f"{field_name} is required")
        text = text.replace("$", "").replace(",", "").replace("(", "-").replace(")", "")
        try:
            return Decimal(text)
        except Exception as exc:
            raise AutoUploadError(f"{field_name} is not a valid number: {value!r}") from exc

    @staticmethod
    def _source_row_label(row: dict[str, Any]) -> str:
        sheet = str(row.get("__source_sheet") or "worksheet")
        row_number = row.get("__source_row") or "?"
        return f"{sheet} row {row_number}"

    def _source_aware_errors(
        self, rows: list[dict[str, Any]], errors: Iterable[Any]
    ) -> list[str]:
        """Translate temporary CSV row numbers back to workbook sheet/row."""
        translated: list[str] = []
        for value in errors:
            message = str(value)
            match = re.match(r"^\s*Row\s+(\d+)\s*:\s*(.*)$", message, flags=re.I)
            if match:
                index = int(match.group(1)) - 2
                if 0 <= index < len(rows):
                    message = f"{self._source_row_label(rows[index])}: {match.group(2)}"
            translated.append(message)
        return translated

    def _process_structured_invoices(
        self,
        claimed: Path,
        rows: list[dict[str, Any]],
        digest: str,
    ) -> RoutingOutcome:
        """Group invoice rows and run each invoice through normal validation."""
        groups: dict[tuple[str, str, str], list[dict[str, Any]]] = {}
        row_errors: list[str] = []
        for row in rows:
            label = self._source_row_label(row)
            vendor = str(self._row_value(row, "Vendor", "Vendor Name")).strip()
            invoice_number = str(
                self._row_value(row, "Invoice Number", "Invoice No", "Invoice #")
            ).strip()
            raw_date = self._row_value(row, "Invoice Date")
            try:
                invoice_date = (
                    raw_date.date().isoformat()
                    if isinstance(raw_date, datetime)
                    else raw_date.isoformat()
                    if isinstance(raw_date, date)
                    else parse_date(raw_date)
                )
            except Exception:
                invoice_date = ""
            missing = [
                name for name, value in (
                    ("Vendor", vendor),
                    ("Invoice Number", invoice_number),
                    ("Invoice Date", invoice_date),
                )
                if not value
            ]
            if missing:
                row_errors.append(f"{label}: missing {', '.join(missing)}")
                continue
            groups.setdefault((vendor, invoice_number, invoice_date), []).append(row)

        archive_path = self.workspace.archive_original(claimed, digest)
        results: list[dict[str, Any]] = []
        approved = 0
        rejected = len(row_errors)
        for (vendor, invoice_number, invoice_date), invoice_rows in groups.items():
            line_items: list[dict[str, Any]] = []
            group_errors: list[str] = []
            header_values: dict[str, Decimal] = {}
            for row in invoice_rows:
                label = self._source_row_label(row)
                try:
                    quantity = self._invoice_cell_decimal(
                        self._row_value(row, "Quantity", "Qty"), "Quantity"
                    )
                    unit_price = self._invoice_cell_decimal(
                        self._row_value(row, "Unit Price", "Price"), "Unit Price"
                    )
                    line_total = self._invoice_cell_decimal(
                        self._row_value(row, "Line Total", "Extended Amount", "Extension"),
                        "Line Total",
                    )
                    description = str(
                        self._row_value(
                            row, "Item Description", "Product Description", "Description"
                        )
                    ).strip()
                    if not description:
                        raise AutoUploadError("Item Description is required")
                    line_items.append({
                        "sku": str(
                            self._row_value(row, "Vendor SKU", "SKU", "Item Number")
                        ).strip(),
                        "description": description,
                        "category": str(
                            self._row_value(row, "Category", default="Unclassified")
                        ).strip() or "Unclassified",
                        "quantity": str(quantity),
                        "unit": str(
                            self._row_value(row, "Unit", "Purchase Unit", default="each")
                        ).strip() or "each",
                        "unit_price": str(unit_price),
                        "line_total": str(line_total),
                        "confidence": 0.99,
                    })
                    for canonical_name, aliases in {
                        "subtotal": ("Subtotal",),
                        "fees": ("Fees", "Fee"),
                        "tax": ("Tax",),
                        "credits": ("Credits", "Credit"),
                        "total": ("Total", "Invoice Total"),
                    }.items():
                        raw_value = self._row_value(row, *aliases, default="")
                        if raw_value in (None, ""):
                            continue
                        value = self._invoice_cell_decimal(raw_value, canonical_name.title())
                        prior = header_values.get(canonical_name)
                        if prior is not None and value != prior:
                            raise AutoUploadError(
                                f"inconsistent {canonical_name} within invoice "
                                f"({prior} versus {value})"
                            )
                        header_values[canonical_name] = value
                except Exception as exc:
                    group_errors.append(f"{label}: {exc}")

            if group_errors:
                row_errors.extend(group_errors)
                rejected += len(group_errors)
                results.append({
                    "vendor": vendor,
                    "invoice_number": invoice_number,
                    "invoice_date": invoice_date,
                    "status": "Needs Review",
                    "errors": group_errors,
                })
                continue
            subtotal = header_values.get(
                "subtotal", sum((Decimal(item["line_total"]) for item in line_items), Decimal("0"))
            )
            fees = header_values.get("fees", Decimal("0"))
            tax = header_values.get("tax", Decimal("0"))
            credits = header_values.get("credits", Decimal("0"))
            total = header_values.get("total", subtotal + fees + tax - credits)
            canonical = {
                "vendor": vendor,
                "invoice_number": invoice_number,
                "invoice_date": invoice_date,
                "subtotal": str(subtotal),
                "fees": str(fees),
                "tax": str(tax),
                "credits": str(credits),
                "total": str(total),
                "currency": self.pipeline.settings.get("currency", "USD"),
                "source_file": claimed.name,
                "source_link": str(archive_path),
                "items": line_items,
                "extraction_notes": ["Imported from structured Excel invoice rows."],
            }
            group_key = f"{normalize(vendor)}|{normalize(invoice_number)}|{invoice_date}"
            result = self.pipeline.process_structured_invoice(
                claimed,
                canonical,
                workbook_hash=digest,
                archive_path=archive_path,
                group_key=group_key,
            )
            payload = result.as_dict()
            results.append(payload)
            if result.status == "Approved":
                approved += 1
            elif result.status != "Duplicate":
                rejected += max(1, len(result.errors))
                for error in result.errors:
                    row_errors.append(
                        f"Invoice {invoice_number} ({vendor}): {error}"
                    )

        statuses = {str(result.get("status") or "") for result in results}
        if groups and statuses and statuses.issubset({"Approved", "Duplicate"}) and not row_errors:
            status = "Processed" if "Approved" in statuses else "Duplicate"
        else:
            status = "Needs Review"
        summary = (
            f"Validated {len(groups)} structured invoice(s); "
            f"{approved} approved and {rejected} row or invoice error(s) need review."
        )
        return RoutingOutcome(
            status,
            "Invoice",
            summary,
            approved,
            rejected,
            {
                "workbook": claimed.name,
                "invoice_count": len(groups),
                "results": results,
                "errors": row_errors[:250],
            },
        )

    def _route_claimed(self, claimed: Path, classification: Classification, digest: str) -> RoutingOutcome:
        detected = classification.detected_type
        if classification.confidence < 0.50 or detected in {"Unclassified", "Unsupported"}:
            return RoutingOutcome(
                "Needs Review",
                detected,
                f"The file was preserved without importing because it could not be identified safely. {classification.reason}",
                details={"headers": classification.headers},
            )
        if detected == "Archive":
            return self._extract_zip(claimed)
        if detected == "Invoice":
            if claimed.suffix.lower() in TABLE_SUFFIXES:
                _, rows = self._read_table(claimed)
                return self._process_structured_invoices(claimed, rows, digest)
            result = self.pipeline.process_file(claimed)
            status = "Duplicate" if result.status == "Duplicate" else result.status
            if status not in {"Approved", "Needs Review", "Duplicate"}:
                status = "Failed"
            if status == "Approved":
                status = "Processed"
            return RoutingOutcome(
                status,
                "Invoice",
                result.message or f"Invoice finished with status {result.status}.",
                1 if result.status == "Approved" else 0,
                len(result.errors),
                result.as_dict(),
            )
        if detected == "Event Calendar" and claimed.suffix.lower() == ".ics":
            result = self.pipeline.import_event_calendar(claimed)
            imported = int(result.get("imported", 0))
            rejected = int(result.get("skipped", 0))
            status = "Processed" if imported and not rejected else "Needs Review" if imported or rejected else "Failed"
            return RoutingOutcome(status, detected, f"Imported {imported} calendar event(s); {rejected} event(s) need review.", imported, rejected, result)

        headers, rows = self._read_table(claimed)
        import_path, temporary = self._materialize_table(claimed, headers, rows)
        # Route the exact worksheet rows used for classification. This matters
        # for workbooks whose first tab is a cover/summary or whose import data
        # spans matching beginning/ending sheets.
        target = import_path
        try:
            if detected == "POS Sales":
                result = self.pipeline.import_pos_report(target, profile_name=f"Auto Upload - {safe_filename(claimed.stem)[:60]}")
                status = "Processed" if result.imported and not result.rejected else "Needs Review" if result.imported or result.rejected else "Failed"
                errors = self._source_aware_errors(rows, result.errors)
                return RoutingOutcome(status, detected, f"Imported {result.imported} POS row(s); {result.rejected} row(s) need review.", result.imported, result.rejected, {"run_id": result.run_id, "errors": errors[:100], "mapping": result.mapping})
            if detected == "Sales Summary":
                imported = (self.pipeline.import_sales_workbook(target) if target.suffix.lower() in {".xlsx", ".xlsm"} else self.pipeline.import_sales_csv(target))
                return RoutingOutcome("Processed" if imported else "Needs Review", detected, f"Imported {imported} sales period(s).", imported, 0)
            if detected == "Operating Costs":
                imported = (self.pipeline.import_operating_costs_workbook(target) if target.suffix.lower() in {".xlsx", ".xlsm"} else self.pipeline.import_operating_costs_csv(target))
                return RoutingOutcome("Processed" if imported else "Needs Review", detected, f"Imported {imported} operating-cost record(s).", imported, 0)
            if detected == "Inventory Count":
                result = (self.pipeline.import_inventory_count_workbook(target) if target.suffix.lower() in {".xlsx", ".xlsm"} else self.pipeline.import_inventory_count_csv(target))
                rejected = len(result.errors) + int(result.skipped)
                status = "Processed" if result.imported and not rejected else "Needs Review" if result.imported or rejected else "Failed"
                errors = self._source_aware_errors(rows, result.errors)
                return RoutingOutcome(status, detected, f"Imported {result.imported} inventory count(s); {rejected} row(s) need review.", result.imported, rejected, {"errors": errors[:100], "skipped": result.skipped})
            if detected == "Recipes":
                result = (self.pipeline.import_recipes_workbook(target) if target.suffix.lower() in {".xlsx", ".xlsm"} else self.pipeline.import_recipes_csv(target))
                imported = int(result.get("imported", 0)); rejected = int(result.get("skipped", 0))
                status = "Processed" if imported and not rejected else "Needs Review" if imported or rejected else "Failed"
                if isinstance(result.get("errors"), list):
                    result["errors"] = self._source_aware_errors(rows, result["errors"])
                return RoutingOutcome(status, detected, f"Imported {imported} recipe line(s); {rejected} row(s) need review.", imported, rejected, result)
            if detected == "Menu Items":
                outcome = self._process_menu_items(rows, digest)
                if isinstance(outcome.details.get("errors"), list):
                    outcome.details["errors"] = self._source_aware_errors(rows, outcome.details["errors"])
                return outcome
            if detected == "Waste Log":
                outcome = self._process_waste(rows, digest)
                if isinstance(outcome.details.get("errors"), list):
                    outcome.details["errors"] = self._source_aware_errors(rows, outcome.details["errors"])
                return outcome
            if detected == "Item Planning":
                outcome = self._process_item_planning(rows)
                if isinstance(outcome.details.get("errors"), list):
                    outcome.details["errors"] = self._source_aware_errors(rows, outcome.details["errors"])
                return outcome
            if detected == "Accounting Mappings":
                return self._process_accounting_mappings(rows)
            if detected == "Event Calendar":
                return self._process_event_rows(rows)
            if detected == "Receiving Log":
                outcome = self._process_receiving_rows(rows, digest)
                if isinstance(outcome.details.get("errors"), list):
                    outcome.details["errors"] = self._source_aware_errors(rows, outcome.details["errors"])
                return outcome
            if detected == "Distributor Catalog":
                profile = self._resolve_distributor(claimed, rows)
                result = self.pipeline.phase3.import_distributor_catalog(profile["distributor_id"], target)
                imported = int(result.get("imported", 0)); rejected = int(result.get("skipped", 0))
                status = "Processed" if imported and not rejected else "Needs Review" if imported or rejected else "Failed"
                return RoutingOutcome(status, detected, f"Imported {imported} catalog row(s) for {profile['distributor_name']}; {rejected} row(s) need review.", imported, rejected, {**result, "distributor": profile["distributor_name"]})
            if detected == "Distributor Confirmation":
                profile = self._resolve_distributor(claimed, rows, confirmation=True)
                result = self.pipeline.phase3.import_distributor_confirmation(profile["distributor_id"], target)
                imported = int(result.get("updated", 0)); rejected = len(result.get("errors", []))
                status = "Processed" if imported and not rejected else "Needs Review" if imported or rejected else "Needs Review"
                return RoutingOutcome(status, detected, f"Updated {imported} purchase order(s) for {profile['distributor_name']}; {rejected} row(s) need review.", imported, rejected, {**result, "distributor": profile["distributor_name"]})
            return RoutingOutcome("Needs Review", detected, "The detected file type does not yet have an automatic importer.")
        finally:
            if temporary:
                import_path.unlink(missing_ok=True)

    def process_file(
        self,
        source: Path,
        classification: Classification | None = None,
    ) -> dict[str, Any]:
        source = source.expanduser().resolve()
        if not source.exists() or not source.is_file():
            raise AutoUploadRetryLater("File is no longer present.")
        original_name = source.name
        original_path = source
        claimed = self._claim(source)
        try:
            digest = sha256_file(claimed)
            classification = classification or self.classify(claimed)
            prior = self._prior_success(digest)
            if prior:
                outcome = RoutingOutcome(
                    "Duplicate",
                    classification.detected_type,
                    f"Exact duplicate of a file already handled on {prior['completed_at']}.",
                    details={"prior_event_id": prior["event_id"], "prior_status": prior["status"]},
                )
            else:
                try:
                    outcome = self._route_claimed(claimed, classification, digest)
                except Exception as exc:
                    if "database is locked" in str(exc).lower() or "being used by another process" in str(exc).lower():
                        destination = self._unique_path(self.inbox / original_name)
                        shutil.move(str(claimed), str(destination))
                        raise AutoUploadRetryLater(str(exc)) from exc
                    outcome = RoutingOutcome("Failed", classification.detected_type, str(exc), details={"error": repr(exc)})
            archived = self._archive(claimed, outcome, original_name)
            self._record_event(
                digest=digest,
                original_name=original_name,
                original_path=original_path,
                classification=classification,
                outcome=outcome,
                archived_path=archived,
            )
            payload = {
                "restaurant": self.restaurant_name,
                "workspace": str(self.workspace.root),
                "inbox": str(self.inbox),
                "original_name": original_name,
                "archived_path": str(archived),
                "classification": asdict(classification),
                "outcome": outcome.as_dict(),
            }
            return payload
        except AutoUploadRetryLater:
            raise
        except Exception:
            # Preserve any claimed file even if the router itself crashes unexpectedly.
            if claimed.exists():
                try:
                    destination = self._unique_path(self.inbox / "_Failed" / "Internal_Error" / datetime.now().strftime("%Y-%m") / original_name)
                    destination.parent.mkdir(parents=True, exist_ok=True)
                    shutil.move(str(claimed), str(destination))
                except Exception:
                    pass
            raise

    def list_events(self, limit: int = 100) -> list[dict[str, Any]]:
        with self.workspace.connect() as conn:
            rows = conn.execute(
                "SELECT * FROM auto_upload_events ORDER BY event_id DESC LIMIT ?", (int(limit),)
            ).fetchall()
        return [dict(row) for row in rows]

    def list_unresolved_events(self, limit: int = 100) -> list[dict[str, Any]]:
        """Return only the latest unresolved attempt for each exact source file."""
        with self.workspace.connect() as conn:
            rows = conn.execute(
                """SELECT event.*
                   FROM auto_upload_events AS event
                   WHERE event.status IN ('Needs Review','Failed')
                     AND event.event_id = (
                         SELECT MAX(newer.event_id)
                         FROM auto_upload_events AS newer
                         WHERE newer.source_sha256=event.source_sha256
                     )
                   ORDER BY event.completed_at DESC, event.event_id DESC
                   LIMIT ?""",
                (int(limit),),
            ).fetchall()
        return [dict(row) for row in rows]

    def retry_event(self, event_id: int) -> dict[str, Any]:
        """Queue an unresolved archived upload for deterministic reprocessing."""
        with self.workspace.connect() as conn:
            row = conn.execute(
                "SELECT * FROM auto_upload_events WHERE event_id=?", (int(event_id),)
            ).fetchone()
            if row is None:
                raise AutoUploadError(f"Auto Upload event {event_id} was not found.")
            latest = conn.execute(
                """SELECT event_id,status FROM auto_upload_events
                   WHERE source_sha256=? ORDER BY event_id DESC LIMIT 1""",
                (row["source_sha256"],),
            ).fetchone()
        if latest and int(latest["event_id"]) != int(event_id):
            raise AutoUploadError(
                f"A newer attempt already exists for this file (event {latest['event_id']}, "
                f"status {latest['status']})."
            )
        if str(row["status"]) not in {"Needs Review", "Failed"}:
            raise AutoUploadError(
                f"Only unresolved uploads can be retried; event {event_id} is {row['status']}."
            )
        archived = Path(str(row["archived_path"] or "")).expanduser()
        if not archived.exists() or not archived.is_file():
            raise AutoUploadError(
                f"The archived workbook is unavailable and cannot be retried: {archived}"
            )
        destination = self._unique_path(self.inbox / str(row["original_name"]))
        shutil.copy2(archived, destination)
        try:
            self.pipeline.controls.audit(
                "auto_upload.retry",
                "auto_upload_event",
                str(event_id),
                f"Queued {row['original_name']} for Auto Upload reprocessing.",
                details={
                    "source_event_id": int(event_id),
                    "queued_path": str(destination),
                    "prior_status": str(row["status"]),
                },
            )
        except Exception:
            pass
        return {
            "event_id": int(event_id),
            "original_name": str(row["original_name"]),
            "queued_path": str(destination),
            "status": "Queued",
        }


class InitialDocumentDiscovery:
    """Find existing restaurant documents without modifying client originals.

    Discovery is deliberately conservative. Content signatures decide whether a
    workbook is useful; filenames are only used to recognize invoice images,
    scanned PDFs, and archives that cannot be classified safely from a table
    header. Accepted files are staged, then copied into the normal Auto Upload
    inbox as one batch so the existing dependency ordering and review rules are
    reused unchanged.
    """

    INVOICE_PATH_HINTS = {
        "invoice", "invoices", "vendor bill", "vendor bills", "accounts payable",
        "supplier", "food distributor", "purchase invoice", "purchase invoices",
    }
    ARCHIVE_PATH_HINTS = {
        "invoice", "inventory", "sales", "pos", "labor", "payroll", "waste",
        "recipe", "restaurant data", "restaurant export", "vendor", "purchases",
    }

    def __init__(self, workspace: RestaurantWorkspace, restaurant_name: str | None = None):
        self.workspace = workspace
        self.router = AutoUploadRouter(workspace, restaurant_name)
        self.inbox = self.router.inbox.resolve()

    @staticmethod
    def _within(path: Path, parent: Path) -> bool:
        try:
            path.resolve().relative_to(parent.resolve())
            return True
        except (OSError, ValueError):
            return False

    def _excluded_roots(self) -> list[Path]:
        roots = [self.inbox, (self.workspace.root / "Backups").resolve()]
        # These folders contain MarginMise-generated copies, logs, or derived
        # artifacts. Intake-oriented folders such as Recipes, Sales, POS
        # Imports, Inventory Counts, Operating Costs, and Upload Invoices are
        # intentionally searchable because they are also common names in an
        # existing restaurant filing system.
        generated_only = {
            "processed", "review", "originals", "extracted", "exports", "logs",
            "phase3", "owner_reports", "margin_memory",
        }
        roots.extend(
            folder.resolve()
            for key, folder in self.workspace.folders.items()
            if key in generated_only
        )
        return roots

    def _iter_files(self, source_root: Path, max_files: int) -> Iterable[tuple[Path, int]]:
        excluded_roots = self._excluded_roots()
        scanned = 0
        for current, directories, files in os.walk(source_root, topdown=True, followlinks=False):
            current_path = Path(current)
            kept_directories = []
            for name in directories:
                candidate = current_path / name
                if name.casefold() in DISCOVERY_EXCLUDED_DIRECTORY_NAMES:
                    continue
                if candidate.is_symlink():
                    continue
                if any(self._within(candidate, excluded) for excluded in excluded_roots):
                    continue
                kept_directories.append(name)
            directories[:] = kept_directories
            for name in files:
                scanned += 1
                if scanned > max_files:
                    return
                yield current_path / name, scanned

    @staticmethod
    def _path_text(path: Path, source_root: Path) -> str:
        try:
            relative = path.relative_to(source_root)
        except ValueError:
            relative = path
        return normalize(" ".join(relative.parts))

    def _pdf_text_looks_like_invoice(self, path: Path) -> bool:
        try:
            import pymupdf as fitz

            with fitz.open(path) as document:
                text = " ".join(page.get_text("text") for page in list(document)[:2])
        except Exception:
            return False
        normalized = normalize(text[:30000])
        signals = (
            "invoice", "invoice number", "invoice date", "bill to", "ship to",
            "subtotal", "amount due", "unit price", "extended amount",
        )
        return sum(1 for signal in signals if signal in normalized) >= 2

    def _safe_discovery_candidate(
        self,
        path: Path,
        source_root: Path,
        classification: Classification,
    ) -> tuple[bool, str]:
        if classification.detected_type not in DISCOVERY_DOCUMENT_TYPES:
            return False, classification.reason
        # The classifier is deterministic for supported formats. Files with a
        # usable schema are safe to queue; the downstream importer still applies
        # row-level validation and preserves failures for review.
        if classification.confidence < 0.50:
            return False, (
                f"Classification confidence {classification.confidence:.0%} is below the "
                "automatic discovery threshold"
            )
        suffix = path.suffix.casefold()
        path_text = self._path_text(path, source_root)
        if classification.detected_type == "Invoice" and suffix in INVOICE_SUFFIXES:
            hinted = any(hint in path_text for hint in self.INVOICE_PATH_HINTS)
            if not hinted and suffix == ".pdf":
                hinted = self._pdf_text_looks_like_invoice(path)
            if not hinted:
                return False, (
                    "Image or PDF did not contain readable invoice signals and its path did not "
                    "identify it as an invoice"
                )
        if classification.detected_type == "Archive" and not any(
            hint in path_text for hint in self.ARCHIVE_PATH_HINTS
        ):
            return False, "Archive name or folder did not identify restaurant operating data"
        return True, classification.reason

    @staticmethod
    def _copy_name(path: Path) -> str:
        stem = safe_filename(path.stem)[:110] or "restaurant_document"
        suffix = path.suffix.casefold()
        return f"{stem}{suffix}"

    def _known_upload_hashes(self) -> set[str]:
        with self.workspace.connect() as conn:
            conn.executescript(AUTO_UPLOAD_SCHEMA_SQL)
            return {
                str(row[0])
                for row in conn.execute(
                    "SELECT DISTINCT source_sha256 FROM auto_upload_events WHERE source_sha256<>''"
                ).fetchall()
            }

    def _generated_materialization_paths(self) -> set[str]:
        """Identify CSV copies created by prior Auto Upload workbook imports.

        Several existing importers keep a normalized CSV in a workspace folder.
        Its bytes differ from the source XLSX, so hash deduplication alone cannot
        distinguish it from a new client file during a later folder search.
        """
        folder_by_type = {
            "POS Sales": "pos",
            "Sales Summary": "sales",
            "Operating Costs": "costs",
            "Inventory Count": "inventory_counts",
            "Recipes": "recipes",
        }
        output: set[str] = set()
        with self.workspace.connect() as conn:
            conn.executescript(AUTO_UPLOAD_SCHEMA_SQL)
            rows = conn.execute(
                """SELECT original_name,detected_type FROM auto_upload_events
                   WHERE status IN ('Processed','Duplicate')"""
            ).fetchall()
        for row in rows:
            original = Path(str(row["original_name"] or ""))
            folder_key = folder_by_type.get(str(row["detected_type"] or ""))
            if not folder_key or original.suffix.casefold() not in {".xlsx", ".xlsm"}:
                continue
            folder = self.workspace.folders.get(folder_key)
            if folder:
                output.add(str((folder / f"{safe_filename(original.stem)}.csv").resolve()).casefold())
        return output

    def discover(
        self,
        source_root: Path,
        *,
        progress_callback: Callable[[dict[str, Any]], None] | None = None,
    ) -> DocumentDiscoveryReport:
        source_root = Path(source_root).expanduser().resolve()
        if not source_root.exists() or not source_root.is_dir():
            raise AutoUploadError(f"Document discovery folder does not exist: {source_root}")
        if source_root == Path(source_root.anchor):
            raise AutoUploadError(
                "Choose the restaurant's folder, not an entire drive, for automatic document discovery."
            )
        if self._within(source_root, self.inbox) or self._within(self.inbox, source_root) and source_root == self.inbox:
            raise AutoUploadError("Choose the restaurant records folder, not its Auto Upload folder.")

        settings = self.workspace.load_settings()
        max_files = max(100, min(50000, int(settings.get("document_discovery_max_files") or 5000)))
        max_bytes = max(1, int(settings.get("document_discovery_max_file_mb") or 100)) * 1024 * 1024
        started = now_iso()
        run_id = f"DISC-{datetime.now().strftime('%Y%m%d-%H%M%S')}-{uuid.uuid4().hex[:6].upper()}"
        report = DocumentDiscoveryReport(run_id, str(source_root), started)
        staging = self.inbox / "_Discovery Staging" / run_id
        staging.mkdir(parents=True, exist_ok=True)
        seen_hashes = self._known_upload_hashes()
        generated_materializations = self._generated_materialization_paths()
        staged: list[tuple[Path, dict[str, Any]]] = []

        def emit(kind: str, **extra: Any) -> None:
            if progress_callback:
                try:
                    progress_callback({
                        "kind": kind,
                        "run_id": run_id,
                        "source_root": str(source_root),
                        "scanned_files": report.scanned_files,
                        "queued_files": report.queued_files,
                        **extra,
                    })
                except Exception:
                    pass

        emit("started")
        try:
            for path, scanned in self._iter_files(source_root, max_files):
                report.scanned_files = scanned
                if scanned % 25 == 0:
                    emit("progress")
                if path.name.startswith("~$") or path.name in SYSTEM_FILES:
                    continue
                try:
                    if str(path.resolve()).casefold() in generated_materializations:
                        continue
                except OSError:
                    pass
                suffix = path.suffix.casefold()
                if suffix not in SUPPORTED_SUFFIXES or suffix in TEMP_SUFFIXES:
                    continue
                try:
                    size = path.stat().st_size
                except OSError as exc:
                    report.error_files += 1
                    report.errors.append(f"{path}: {exc}")
                    continue
                if size <= 0:
                    report.unclassified_files += 1
                    report.skipped.append({"path": str(path), "reason": "Empty file"})
                    continue
                if size > max_bytes:
                    report.oversized_files += 1
                    report.skipped.append({
                        "path": str(path),
                        "reason": f"File exceeds the {max_bytes // (1024 * 1024)} MB discovery limit",
                    })
                    continue
                try:
                    digest = sha256_file(path)
                    if digest in seen_hashes:
                        report.duplicate_files += 1
                        report.skipped.append({"path": str(path), "reason": "Duplicate content"})
                        continue
                    classification = self.router.classify(path)
                    accepted, reason = self._safe_discovery_candidate(path, source_root, classification)
                    if not accepted:
                        report.unclassified_files += 1
                        report.skipped.append({
                            "path": str(path),
                            "detected_type": classification.detected_type,
                            "confidence": classification.confidence,
                            "reason": reason,
                        })
                        continue
                    report.supported_files += 1
                    destination = self.router._unique_path(staging / self._copy_name(path))
                    shutil.copy2(path, destination)
                    seen_hashes.add(digest)
                    detail = {
                        "source_path": str(path),
                        "queued_name": destination.name,
                        "source_sha256": digest,
                        "detected_type": classification.detected_type,
                        "classification_confidence": classification.confidence,
                        "classification_reason": classification.reason,
                        "size_bytes": size,
                    }
                    staged.append((destination, detail))
                except Exception as exc:
                    report.error_files += 1
                    report.errors.append(f"{path}: {exc}")

            report.stopped_at_limit = report.scanned_files >= max_files
            for staged_path, detail in staged:
                try:
                    destination = self.router._unique_path(self.inbox / staged_path.name)
                    os.replace(staged_path, destination)
                    detail["queued_path"] = str(destination)
                    report.queued.append(detail)
                    report.queued_files += 1
                    emit("queued", file=detail)
                except OSError as exc:
                    report.error_files += 1
                    report.errors.append(f"Could not queue {detail['source_path']}: {exc}")

            report.status = "Completed with Errors" if report.error_files else "Completed"
            report.completed_at = now_iso()
            report_dir = self.workspace.folders["logs"] / "Document Discovery"
            report_dir.mkdir(parents=True, exist_ok=True)
            report_path = report_dir / f"{run_id}.json"
            report.report_path = str(report_path)
            report_path.write_text(json.dumps(report.as_dict(), indent=2), encoding="utf-8")

            settings = self.workspace.load_settings()
            settings.update({
                "initial_document_discovery_pending": False,
                "document_discovery_last_status": report.status,
                "document_discovery_last_source": str(source_root),
                "document_discovery_last_completed": report.completed_at,
                "document_discovery_last_summary": report.summary,
                "document_discovery_last_report": str(report_path),
            })
            self.workspace.save_settings(settings)
            try:
                self.router.pipeline.controls.audit(
                    "auto_upload.discovery",
                    "workspace",
                    str(self.workspace.root),
                    report.summary,
                    details={
                        "run_id": report.run_id,
                        "source_root": str(source_root),
                        "report_path": str(report_path),
                        "queued": report.queued,
                    },
                )
            except Exception:
                pass
            emit("completed", report=report.as_dict(), summary=report.summary)
            return report
        finally:
            try:
                if staging.exists():
                    shutil.rmtree(staging)
                staging_parent = staging.parent
                if staging_parent.exists() and not any(staging_parent.iterdir()):
                    staging_parent.rmdir()
            except OSError:
                pass


class AutoUploadCoordinator:
    """Poll and process all registered restaurant inboxes in one daemon thread."""

    def __init__(
        self,
        restaurant_provider: Callable[[], Iterable[dict[str, Any]]],
        event_callback: Callable[[dict[str, Any]], None] | None = None,
        *,
        scan_interval: float = 2.0,
        max_files_per_cycle: int = 2,
    ):
        self.restaurant_provider = restaurant_provider
        self.event_callback = event_callback
        self.scan_interval = max(0.75, float(scan_interval))
        self.max_files_per_cycle = max(1, min(10, int(max_files_per_cycle)))
        self.stop_event = threading.Event()
        self.wake_event = threading.Event()
        self.thread: threading.Thread | None = None
        self._stable: dict[str, tuple[int, int, float]] = {}
        self._routers: dict[str, AutoUploadRouter] = {}

    def start(self) -> None:
        if self.thread and self.thread.is_alive():
            return
        self.stop_event.clear()
        self.thread = threading.Thread(target=self._run, name="RestaurantAutoUpload", daemon=True)
        self.thread.start()

    def stop(self, timeout: float = 4.0) -> None:
        self.stop_event.set()
        self.wake_event.set()
        if self.thread and self.thread.is_alive():
            self.thread.join(timeout=timeout)

    def scan_now(self) -> None:
        self.wake_event.set()

    def _emit(self, payload: dict[str, Any]) -> None:
        if not self.event_callback:
            return
        try:
            self.event_callback(payload)
        except Exception:
            pass

    def _router_for(self, row: dict[str, Any]) -> AutoUploadRouter | None:
        raw_path = str(row.get("path") or "").strip()
        if not raw_path:
            return None
        try:
            workspace = RestaurantWorkspace(Path(raw_path))
        except Exception as exc:
            self._emit({"kind": "error", "restaurant": row.get("name", "Restaurant"), "message": str(exc)})
            return None
        settings = workspace.load_settings()
        if not bool(settings.get("auto_upload_enabled", True)):
            return None
        key = str(workspace.root)
        router = self._routers.get(key)
        if not router:
            try:
                router = AutoUploadRouter(workspace, str(row.get("name") or settings.get("restaurant_name") or workspace.root.name))
                self._routers[key] = router
                self._emit({"kind": "watching", "restaurant": router.restaurant_name, "workspace": key, "inbox": str(router.inbox)})
            except Exception as exc:
                self._emit({"kind": "error", "restaurant": row.get("name", "Restaurant"), "workspace": key, "message": str(exc)})
                return None
        return router

    @staticmethod
    def _candidate(path: Path) -> bool:
        if not path.is_file() or path.name in SYSTEM_FILES or path.name.startswith("."):
            return False
        suffix = path.suffix.lower()
        if suffix in TEMP_SUFFIXES or path.name.endswith(".result.json"):
            return False
        return True

    def _ready(self, path: Path, stability_seconds: float) -> bool:
        try:
            stat = path.stat()
        except OSError:
            return False
        key = str(path)
        current = (int(stat.st_size), int(stat.st_mtime_ns))
        previous = self._stable.get(key)
        now = time.monotonic()
        if not previous or previous[:2] != current:
            self._stable[key] = (current[0], current[1], now)
            return False
        return now - previous[2] >= stability_seconds

    def _scan_router(self, router: AutoUploadRouter) -> None:
        settings = router.workspace.load_settings()
        stability = max(1.0, float(settings.get("auto_upload_stability_seconds", 2.0)))
        candidates = sorted((path for path in router.inbox.iterdir() if self._candidate(path)), key=lambda path: path.name.lower())
        ready: list[tuple[int, str, Path, Classification]] = []
        priority = {
            "Archive": 0,
            "Invoice": 10,
            "Distributor Catalog": 20,
            "Menu Items": 20,
            "Item Planning": 25,
            "Recipes": 30,
            "Inventory Count": 40,
            "Waste Log": 45,
            "POS Sales": 50,
            "Sales Summary": 50,
            "Operating Costs": 60,
            "Accounting Mappings": 60,
            "Event Calendar": 70,
            "Receiving Log": 80,
        }
        # Do not wait for every file in the inbox to become stable. A single
        # workbook that is still being copied, scanned by antivirus, or held
        # open by Excel must never block unrelated invoices forever.
        for path in candidates:
            if self.stop_event.is_set():
                return
            if not self._ready(path, stability):
                continue
            self._stable.pop(str(path), None)
            classification = router.classify(path)
            ready.append((
                priority.get(classification.detected_type, 90),
                path.name.lower(),
                path,
                classification,
            ))
        if not ready:
            return
        max_per_cycle = max(1, min(10, int(settings.get("auto_upload_max_files_per_cycle", self.max_files_per_cycle))))
        for _, _, path, classification in sorted(ready)[:max_per_cycle]:
            if self.stop_event.is_set():
                return
            try:
                result = router.process_file(path, classification)
                result["kind"] = "processed"
                self._emit(result)
            except AutoUploadRetryLater as exc:
                self._emit({"kind": "retry", "restaurant": router.restaurant_name, "workspace": str(router.workspace.root), "inbox": str(router.inbox), "original_name": path.name, "message": str(exc)})
            except Exception as exc:
                self._emit({"kind": "error", "restaurant": router.restaurant_name, "workspace": str(router.workspace.root), "inbox": str(router.inbox), "original_name": path.name, "message": str(exc)})

    def scan_once(self) -> None:
        active_keys: set[str] = set()
        for row in list(self.restaurant_provider() or []):
            router = self._router_for(dict(row))
            if not router:
                continue
            active_keys.add(str(router.workspace.root))
            self._scan_router(router)
        for key in list(self._routers):
            if key not in active_keys:
                self._routers.pop(key, None)

    def _run(self) -> None:
        while not self.stop_event.is_set():
            try:
                self.scan_once()
            except Exception as exc:
                self._emit({"kind": "error", "restaurant": "Auto Upload", "message": str(exc)})
            self.wake_event.wait(self.scan_interval)
            self.wake_event.clear()


def main() -> int:
    """Optional headless watcher using the normal GUI restaurant registry."""
    import argparse

    parser = argparse.ArgumentParser(description="Watch every registered restaurant Desktop upload folder.")
    parser.add_argument("--registry", default=str(Path.home() / ".restaurant_cost_controller_gui.json"))
    parser.add_argument("--once", action="store_true")
    args = parser.parse_args()

    registry_path = Path(args.registry).expanduser()

    def provider() -> list[dict[str, Any]]:
        if not registry_path.exists():
            return []
        try:
            payload = json.loads(registry_path.read_text(encoding="utf-8"))
            return list(payload.get("restaurants") or [])
        except Exception:
            return []

    def report(payload: dict[str, Any]) -> None:
        print(json.dumps(payload, default=str), flush=True)

    coordinator = AutoUploadCoordinator(provider, report)
    if args.once:
        coordinator.scan_once()
        time.sleep(2.5)
        coordinator.scan_once()
        return 0
    coordinator.start()
    print("Automatic upload service is running. Press Ctrl+C to stop.", flush=True)
    try:
        while True:
            time.sleep(1)
    except KeyboardInterrupt:
        coordinator.stop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
