#!/usr/bin/env python3
"""Restaurant invoice intake, OCR, validation, review, and persistence pipeline.

The pipeline deliberately separates extraction from validation:

    source PDF/image/JSON
        -> existing text extraction or on-demand local OCR
        -> canonical invoice JSON
        -> deterministic validation
        -> auto-commit or human review queue

RapidOCR is the primary scan engine. Tesseract is a fully-local fallback. All
extraction, OCR, and structuring runs entirely on the local machine — no
external AI provider or cloud service is required.

It supports multiple restaurant workspaces. Each workspace owns its own settings,
SQLite database, source archive, review queue, and exports.
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import os
import re
import shutil
import sqlite3
import subprocess
import sys
import tempfile
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime, date, timedelta
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Iterable, Sequence

try:
    import fitz  # PyMuPDF: text extraction and PDF page rendering only, never OCR.
except ImportError:  # pragma: no cover - installer supplies it.
    fitz = None

from excel_io import write_table_as

# All OCR and extraction use RapidOCR (ONNX) and Tesseract, both fully local engines.
from inventory_planning import (
    InventoryPlanningService,
    infer_count_conversion,
    preferred_sales_rows,
)
from operational_controls import OperationalControlsService
from phase2_features import Phase2Service
from phase3_features import Phase3Service
from recipe_costing import RecipeCostingService
from margin_memory import MarginMemoryService
from review_copilot import ReviewCopilotService
from local_ocr import _line_texts


def _run_rapidocr_in_process(image_paths: Sequence[Path], timeout: int = 120) -> dict[str, Any]:
    """Run RapidOCR in the current process with a memoized engine.

    Loading the ONNX detection + recognition models is the expensive part.
    Memoizing the engine means a 35-invoice folder loads the model once instead
    of 35 times (the old per-file subprocess design). The engine is released
    between calls only under memory pressure; on a normal workstation keeping it
    warm is far faster than reloading per document.
    """
    from rapidocr import RapidOCR

    engine = _RAPIDOCR_ENGINE.engine
    if engine is None:
        engine = RapidOCR(
            params={
                "Global.log_level": "error",
                "EngineConfig.onnxruntime.intra_op_num_threads": 2,
                "EngineConfig.onnxruntime.inter_op_num_threads": 1,
                "EngineConfig.onnxruntime.enable_cpu_mem_arena": False,
            }
        )
        _RAPIDOCR_ENGINE.engine = engine

    pages: list[dict[str, Any]] = []
    all_scores: list[float] = []
    for index, image_path in enumerate(image_paths, 1):
        output = engine(image_path)
        raw_texts = getattr(output, "txts", None)
        raw_boxes = getattr(output, "boxes", None)
        raw_scores = getattr(output, "scores", None)
        texts = list(raw_texts) if raw_texts is not None else []
        boxes = list(raw_boxes) if raw_boxes is not None else []
        scores = [float(value) for value in raw_scores] if raw_scores is not None else []
        all_scores.extend(scores)
        lines = _line_texts(boxes, texts) if boxes else [" ".join(str(text).split()) for text in texts]
        pages.append(
            {
                "page": index,
                "path": str(image_path),
                "text": "\n".join(lines).strip(),
                "average_confidence": (sum(scores) / len(scores)) if scores else 0.0,
            }
        )
    return {
        "engine": "rapidocr-onnx",
        "pages": pages,
        "text": "\n\n".join(
            f"--- PAGE {page['page']} ---\n{page['text']}" for page in pages if page["text"]
        ).strip(),
        "average_confidence": (sum(all_scores) / len(all_scores)) if all_scores else 0.0,
    }


# Module-level memoized RapidOCR engine (one load per process).
class _RapidOCREngineHolder:
    engine = None


_RAPIDOCR_ENGINE = _RapidOCREngineHolder()


MONEY = Decimal("0.01")
SUPPORTED_SOURCE_SUFFIXES = {".pdf", ".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp", ".json"}

DEFAULT_SETTINGS: dict[str, Any] = {
    "restaurant_name": "New Restaurant",
    "currency": "USD",
    "timezone": "America/Chicago",
    "minimum_extraction_confidence": 0.40,
    "auto_approve_confidence": 0.70,
    "invoice_math_tolerance": 0.05,
    "price_alert_percent": 5.0,
    "require_review_for_unrecognized_vendors": False,
    "auto_learn_validated_vendors": True,
    "extraction_mode": "local_first",
    "auto_install_scan_engine": True,
    "local_ocr_enabled": True,
    "rapidocr_enabled": True,
    "local_ocr_timeout_seconds": 120,
    "pdf_render_dpi": 200,
    "max_pdf_pages": 30,
    "max_source_file_mb": 50,
    "known_vendors": [],
    "forecast_history_months": 3,
    "default_lead_time_days": 2.0,
    "default_order_cycle_days": 7.0,
    "default_safety_stock_days": 2.0,
    "default_order_multiple": 1.0,
    "sales_adjust_order_predictions": True,
    "include_zero_order_items": True,
    "auto_generate_weekly_order_draft": True,
    "manager_chat_enabled": True,
    "assistant_display_name": "CostPilot",
    "manager_chat_provider": "local",
    "manager_chat_model": "lfm2.5-1.2b-instruct-q4_k_m",
    "manager_chat_free_only": True,
    "manager_chat_timeout_seconds": 240,
    "manager_chat_context_max_items": 120,
    "manager_chat_history_turns": 8,
    "manager_chat_local_fallback": True,
    "manager_chat_cloud_fallback_enabled": False,
    "costpilot_local_migration_version": 1,
    "automatic_backups_enabled": True,
    "automatic_backup_interval_hours": 24,
    "backup_retention_count": 30,
    "require_login": True,
    "data_quality_warning_threshold": 75,
    "receiving_verification_enabled": True,
    "auto_recover_invoice_headers": True,
    "auto_approve_recovered_invoice_headers": True,
    "auto_verify_clean_receiving": True,
    "auto_verify_receiving_date_mode": "invoice_date",
    "costpilot_review_center_enabled": True,
    "costpilot_review_auto_explain": True,
    "costpilot_review_confirm_batch_actions": True,
    "restaurant_group": "My Restaurant Group",
    "address": "",
    "latitude": "",
    "longitude": "",
    "event_weather_forecasting_enabled": True,
    "weather_forecast_days": 16,
    "target_menu_food_cost_percent": 30.0,
    "estimated_manual_invoice_minutes": 8.0,
    "estimated_manager_hourly_cost": 25.0,
    "distributor_exchange_enabled": True,
    "auto_upload_enabled": True,
    "auto_upload_folder": "",
    "auto_upload_scan_seconds": 2.0,
    "auto_upload_stability_seconds": 2.0,
    "auto_upload_max_files_per_cycle": 2,
    "initial_document_discovery_pending": False,
    "document_discovery_max_files": 5000,
    "document_discovery_max_file_mb": 100,
    "margin_memory_enabled": True,
    "margin_memory_materiality_threshold_percent": 10.0,
    "margin_memory_capture_order_overrides": True,
    "margin_memory_capture_transfers": True,
    "margin_memory_capture_receiving": True,
    "margin_memory_capture_invoice_corrections": True,
}

SCHEMA_SQL = """
PRAGMA foreign_keys = ON;
PRAGMA journal_mode = WAL;

CREATE TABLE IF NOT EXISTS vendors (
    vendor_key TEXT PRIMARY KEY,
    vendor_name TEXT NOT NULL,
    recognized INTEGER NOT NULL DEFAULT 0,
    parser_name TEXT NOT NULL DEFAULT 'generic',
    first_seen TEXT NOT NULL,
    last_seen TEXT NOT NULL
);

CREATE TABLE IF NOT EXISTS invoices (
    invoice_id TEXT PRIMARY KEY,
    source_sha256 TEXT NOT NULL,
    duplicate_key TEXT NOT NULL UNIQUE,
    source_name TEXT NOT NULL,
    source_original_path TEXT NOT NULL,
    source_archive_path TEXT,
    vendor TEXT,
    invoice_number TEXT,
    invoice_date TEXT,
    subtotal TEXT,
    fees TEXT,
    tax TEXT,
    credits TEXT,
    total TEXT,
    extraction_method TEXT,
    extraction_confidence REAL,
    recognized_vendor INTEGER NOT NULL DEFAULT 0,
    status TEXT NOT NULL,
    notes TEXT,
    canonical_json TEXT NOT NULL,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
);

CREATE INDEX IF NOT EXISTS idx_invoices_date ON invoices(invoice_date);
CREATE INDEX IF NOT EXISTS idx_invoices_vendor ON invoices(vendor);
CREATE INDEX IF NOT EXISTS idx_invoices_hash ON invoices(source_sha256);

CREATE TABLE IF NOT EXISTS invoice_lines (
    line_id INTEGER PRIMARY KEY AUTOINCREMENT,
    invoice_id TEXT NOT NULL REFERENCES invoices(invoice_id) ON DELETE CASCADE,
    line_number INTEGER NOT NULL,
    vendor_sku TEXT,
    item_id TEXT,
    description TEXT,
    normalized_description TEXT,
    category TEXT,
    quantity TEXT,
    unit TEXT,
    unit_price TEXT,
    line_total TEXT,
    confidence REAL,
    match_status TEXT,
    notes TEXT
);

CREATE TABLE IF NOT EXISTS items (
    item_id TEXT PRIMARY KEY,
    vendor_key TEXT NOT NULL,
    vendor_name TEXT NOT NULL,
    vendor_sku TEXT,
    item_name TEXT NOT NULL,
    normalized_description TEXT NOT NULL,
    category TEXT NOT NULL DEFAULT 'Unclassified',
    unit TEXT,
    first_purchase_date TEXT,
    last_purchase_date TEXT,
    first_price TEXT,
    previous_price TEXT,
    current_price TEXT,
    price_change_percent TEXT,
    average_price TEXT,
    lowest_price TEXT,
    highest_price TEXT,
    total_spent TEXT,
    purchase_count INTEGER NOT NULL DEFAULT 0,
    review_status TEXT NOT NULL DEFAULT 'Approved',
    UNIQUE(vendor_key, vendor_sku),
    UNIQUE(vendor_key, normalized_description)
);

CREATE TABLE IF NOT EXISTS price_history (
    price_id INTEGER PRIMARY KEY AUTOINCREMENT,
    invoice_id TEXT NOT NULL REFERENCES invoices(invoice_id) ON DELETE CASCADE,
    invoice_date TEXT NOT NULL,
    vendor_name TEXT NOT NULL,
    vendor_sku TEXT,
    item_id TEXT NOT NULL REFERENCES items(item_id),
    item_description TEXT NOT NULL,
    category TEXT,
    quantity TEXT,
    unit TEXT,
    unit_price TEXT,
    line_total TEXT,
    previous_price TEXT,
    price_change_percent TEXT,
    price_alert INTEGER NOT NULL DEFAULT 0,
    confidence REAL,
    match_status TEXT,
    notes TEXT
);

CREATE TABLE IF NOT EXISTS reviews (
    review_id INTEGER PRIMARY KEY AUTOINCREMENT,
    invoice_id TEXT NOT NULL REFERENCES invoices(invoice_id) ON DELETE CASCADE,
    item_id TEXT,
    severity TEXT NOT NULL,
    issue_type TEXT NOT NULL,
    issue TEXT NOT NULL,
    payload_json TEXT,
    status TEXT NOT NULL DEFAULT 'Open',
    resolution TEXT,
    created_at TEXT NOT NULL,
    resolved_at TEXT
);

CREATE TABLE IF NOT EXISTS sales (
    sales_id INTEGER PRIMARY KEY AUTOINCREMENT,
    period_start TEXT NOT NULL,
    period_end TEXT NOT NULL,
    gross_sales TEXT NOT NULL DEFAULT '0.00',
    discounts TEXT NOT NULL DEFAULT '0.00',
    refunds TEXT NOT NULL DEFAULT '0.00',
    sales_tax TEXT NOT NULL DEFAULT '0.00',
    net_sales TEXT NOT NULL,
    source_file TEXT,
    UNIQUE(period_start, period_end, source_file)
);

CREATE TABLE IF NOT EXISTS operating_costs (
    cost_id INTEGER PRIMARY KEY AUTOINCREMENT,
    cost_date TEXT NOT NULL,
    category TEXT NOT NULL,
    description TEXT NOT NULL,
    amount TEXT NOT NULL,
    source_file TEXT
);

CREATE TABLE IF NOT EXISTS shift_report_logs (
    log_id INTEGER PRIMARY KEY AUTOINCREMENT,
    source_path TEXT NOT NULL,
    source_name TEXT NOT NULL,
    report_date TEXT,
    shift TEXT,
    labor_cost REAL,
    guests INTEGER,
    net_sales REAL,
    surcharge REAL,
    notes TEXT,
    extracted_at TEXT NOT NULL
);
"""


@dataclass
class Finding:
    severity: str
    issue_type: str
    issue: str
    line_number: int | None = None
    item_id: str = ""


@dataclass
class ExtractionResult:
    data: dict[str, Any]
    method: str
    confidence: float
    vendor_recognized: bool
    parser_name: str
    warnings: list[str] = field(default_factory=list)
    raw_text: str = ""


@dataclass
class ProcessResult:
    source: str
    invoice_id: str = ""
    status: str = "Failed"
    extraction_method: str = ""
    extraction_confidence: float = 0.0
    recognized_vendor: bool = False
    errors: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    message: str = ""

    def as_dict(self) -> dict[str, Any]:
        return {
            "source": self.source,
            "invoice_id": self.invoice_id,
            "status": self.status,
            "extraction_method": self.extraction_method,
            "extraction_confidence": round(self.extraction_confidence, 4),
            "recognized_vendor": self.recognized_vendor,
            "errors": self.errors,
            "warnings": self.warnings,
            "message": self.message,
        }


class PipelineError(RuntimeError):
    pass


class ExtractionFailed(PipelineError):
    pass


def now_iso() -> str:
    return datetime.now().replace(microsecond=0).isoformat()


def normalize_text(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^A-Za-z0-9]+", " ", text).strip().upper()
    return re.sub(r"\s+", " ", text)


def clean_numeric_text(value: Any) -> str:
    text = str(value if value is not None else "").strip()
    text = text.replace("$", "").replace(",", "")
    if text.startswith("(") and text.endswith(")"):
        text = "-" + text[1:-1]
    return text


def decimal_value(value: Any, field_name: str, *, required: bool = True) -> Decimal:
    text = clean_numeric_text(value)
    if not text:
        if required:
            raise ValueError(f"Missing numeric value for {field_name}")
        return Decimal("0")
    try:
        return Decimal(text)
    except (InvalidOperation, ValueError) as exc:
        raise ValueError(f"Invalid numeric value for {field_name}: {value!r}") from exc


def money(value: Any, field_name: str, *, required: bool = True) -> Decimal:
    return decimal_value(value, field_name, required=required).quantize(MONEY, rounding=ROUND_HALF_UP)


def money_string(value: Any, field_name: str = "amount", *, required: bool = True) -> str:
    return f"{money(value, field_name, required=required):.2f}"


def parse_date(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"(?<=\d)(?:st|nd|rd|th)\b", "", text, flags=re.I)
    text = re.sub(r"\s+", " ", text).strip().rstrip(".,")
    formats = (
        "%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d",
        "%m/%d/%Y", "%m-%d-%Y", "%m.%d.%Y",
        "%m/%d/%y", "%m-%d-%y", "%m.%d.%y",
        "%B %d, %Y", "%b %d, %Y", "%B %d %Y", "%b %d %Y",
        "%d %B %Y", "%d %b %Y", "%d-%B-%Y", "%d-%b-%Y",
        "%d-%B-%y", "%d-%b-%y", "%B-%d-%Y", "%b-%d-%Y",
        "%B-%d-%y", "%b-%d-%y",
    )
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f"Unsupported or missing invoice date: {text!r}")


def deterministic_id(prefix: str, value: str, length: int = 12) -> str:
    digest = hashlib.sha256(value.encode("utf-8")).hexdigest()[:length].upper()
    return f"{prefix}-{digest}"


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def safe_filename(value: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]+", "_", value.strip())
    return cleaned.strip("._") or "document"


def json_safe(value: Any) -> Any:
    if isinstance(value, Decimal):
        return f"{value:.2f}"
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, dict):
        return {str(k): json_safe(v) for k, v in value.items()}
    if isinstance(value, list):
        return [json_safe(v) for v in value]
    return value


class ManagedSQLiteConnection(sqlite3.Connection):
    """SQLite connection that is always closed after a ``with`` block.

    ``sqlite3.Connection.__exit__`` commits or rolls back but does not close the
    native handle.  That behavior leaves restaurant databases locked on Windows
    until garbage collection, preventing backup restore and temporary-workspace
    cleanup.  The workspace API is used as a context manager throughout the app,
    so closing here gives those call sites the lifecycle they already imply.
    """

    def __exit__(self, exc_type: Any, exc: BaseException | None, traceback: Any) -> bool:
        try:
            return bool(super().__exit__(exc_type, exc, traceback))
        finally:
            self.close()


class RestaurantWorkspace:
    """A restaurant-specific data boundary and persistent ledger."""

    def __init__(self, root: Path):
        self.root = root.expanduser().resolve()
        self.config_path = self.root / "restaurant_config.json"
        self.db_path = self.root / "restaurant_costs.sqlite3"
        self.folders = {
            "upload": self.root / "Upload Invoices",
            "processed": self.root / "Processed Invoices",
            "review": self.root / "Needs Review",
            "originals": self.root / "Original Documents",
            "extracted": self.root / "Extracted JSON",
            "sales": self.root / "Sales",
            "costs": self.root / "Operating Costs",
            "exports": self.root / "Exports",
            "logs": self.root / "Logs",
        }
        self.initialize()

    def initialize(self) -> None:
        self.root.mkdir(parents=True, exist_ok=True)
        for folder in self.folders.values():
            folder.mkdir(parents=True, exist_ok=True)
        if not self.config_path.exists():
            self.save_settings(dict(DEFAULT_SETTINGS))
        with self.connect() as conn:
            conn.executescript(SCHEMA_SQL)
            columns = {row["name"] for row in conn.execute("PRAGMA table_info(reviews)").fetchall()}
            if "item_id" not in columns:
                conn.execute("ALTER TABLE reviews ADD COLUMN item_id TEXT")
        self._seed_known_vendors()

    def connect(self) -> sqlite3.Connection:
        conn = sqlite3.connect(self.db_path, factory=ManagedSQLiteConnection)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        return conn

    def load_settings(self) -> dict[str, Any]:
        settings = dict(DEFAULT_SETTINGS)
        if self.config_path.exists():
            try:
                loaded = json.loads(self.config_path.read_text(encoding="utf-8"))
                if isinstance(loaded, dict):
                    settings.update(loaded)
            except json.JSONDecodeError:
                pass
        # Migration: OCR-extracted invoices run at ~0.80 confidence, so the
        # legacy thresholds (0.82 min / 0.92 auto-approve) made it impossible
        # for any realistic scan to auto-post. Promote any stale, OCR-hostile
        # threshold on existing workspaces so clean documents post automatically.
        stale_min = float(settings.get("minimum_extraction_confidence", 0.40))
        if stale_min > 0.70:
            settings["minimum_extraction_confidence"] = 0.40
        stale_auto = float(settings.get("auto_approve_confidence", 0.70))
        if stale_auto > 0.85:
            settings["auto_approve_confidence"] = 0.70
        return settings

    def save_settings(self, settings: dict[str, Any]) -> None:
        merged = dict(DEFAULT_SETTINGS)
        merged.update(settings)
        self.config_path.write_text(json.dumps(merged, indent=2), encoding="utf-8")

    def _seed_known_vendors(self) -> None:
        settings = self.load_settings()
        vendors = settings.get("known_vendors") or []
        timestamp = now_iso()
        with self.connect() as conn:
            for vendor_name in vendors:
                key = normalize_text(vendor_name)
                if not key:
                    continue
                conn.execute(
                    """INSERT INTO vendors(vendor_key, vendor_name, recognized, parser_name, first_seen, last_seen)
                       VALUES(?, ?, 1, 'generic', ?, ?)
                       ON CONFLICT(vendor_key) DO UPDATE SET recognized=1, vendor_name=excluded.vendor_name, last_seen=excluded.last_seen""",
                    (key, str(vendor_name), timestamp, timestamp),
                )

    def mark_vendor_recognized(self, vendor_name: str, parser_name: str = "generic") -> None:
        key = normalize_text(vendor_name)
        if not key:
            return
        timestamp = now_iso()
        with self.connect() as conn:
            conn.execute(
                """INSERT INTO vendors(vendor_key, vendor_name, recognized, parser_name, first_seen, last_seen)
                   VALUES(?, ?, 1, ?, ?, ?)
                   ON CONFLICT(vendor_key) DO UPDATE SET recognized=1, vendor_name=excluded.vendor_name,
                       parser_name=excluded.parser_name, last_seen=excluded.last_seen""",
                (key, vendor_name, parser_name, timestamp, timestamp),
            )

    def vendor_recognition(self, vendor_name: str) -> tuple[bool, str]:
        key = normalize_text(vendor_name)
        if not key:
            return False, "generic"
        with self.connect() as conn:
            row = conn.execute(
                "SELECT recognized, parser_name FROM vendors WHERE vendor_key=?", (key,)
            ).fetchone()
        if row:
            return bool(row["recognized"]), row["parser_name"] or "generic"
        return False, "generic"

    def archive_original(self, source: Path, source_hash: str) -> Path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        name = f"{timestamp}_{source_hash[:10]}_{safe_filename(source.name)}"
        destination = self.folders["originals"] / name
        if not destination.exists():
            shutil.copy2(source, destination)
        return destination

    def copy_to_status_folder(self, source: Path, status: str, invoice_id: str) -> Path:
        folder = self.folders["processed"] if status == "Approved" else self.folders["review"]
        destination = folder / f"{invoice_id}_{safe_filename(source.name)}"
        if source.resolve() != destination.resolve():
            shutil.copy2(source, destination)
        return destination


class LocalExtractor:
    """Fully-local invoice extraction.

    Text PDFs are read locally with PyMuPDF, which is deterministic and does not
    perform OCR. Local deterministic parsing structures that text. Image-only or
    scanned documents are processed with RapidOCR (primary) or Tesseract (fallback),
    both running entirely on the local machine with no external service dependency.
    """


    HEADER_FIELDS = (
        "vendor", "invoice_number", "invoice_date", "subtotal", "fees",
        "tax", "credits", "total", "currency",
    )
    MONEY_PATTERN = r"(?:\$\s*)?[-+]?\(?[\d,]+(?:\.\d{2})?\)?"

    def __init__(self, settings: dict[str, Any], logs_dir: Path | None = None):
        self.settings = settings
        self.logs_dir = logs_dir
        if self.logs_dir:
            self.logs_dir.mkdir(parents=True, exist_ok=True)

    @staticmethod
    def _confidence_value(value: Any) -> float:
        try:
            confidence = float(value)
        except (TypeError, ValueError):
            return 0.0
        if 1.0 < confidence <= 100.0:
            confidence /= 100.0
        return max(0.0, min(1.0, confidence))

    @staticmethod
    def _has_meaningful_text(text: str) -> bool:
        compact = re.sub(r"\s+", " ", text or "").strip()
        return len(compact) >= 60 and bool(re.search(r"[A-Za-z]", compact)) and bool(re.search(r"\d", compact))

    @staticmethod
    def _money_from_text(value: str | None) -> str:
        if not value:
            return ""
        cleaned = str(value).strip().replace("$", "").replace(",", "")
        negative = cleaned.startswith("(") and cleaned.endswith(")")
        cleaned = cleaned.strip("() ")
        try:
            amount = Decimal(cleaned).quantize(MONEY, rounding=ROUND_HALF_UP)
        except (InvalidOperation, ValueError):
            return ""
        if negative:
            amount = -amount
        return f"{amount:.2f}"

    @staticmethod
    def _date_from_text(value: str | None) -> str:
        if not value:
            return ""
        cleaned = str(value).strip().rstrip(".,")
        try:
            return parse_date(cleaned)
        except Exception:
            for fmt in ("%b %d, %Y", "%B %d, %Y", "%d %b %Y", "%d %B %Y"):
                try:
                    return datetime.strptime(cleaned, fmt).date().isoformat()
                except ValueError:
                    pass
        return ""

    def _quality_score(self, payload: dict[str, Any]) -> float:
        required_header = ("vendor", "invoice_number", "invoice_date", "subtotal", "total")
        header_score = sum(bool(str(payload.get(k, "")).strip()) for k in required_header) / len(required_header)
        items = payload.get("items") if isinstance(payload.get("items"), list) else []
        if not items:
            return min(0.48, 0.18 + 0.30 * header_score)
        item_scores, arithmetic_scores = [], []
        for item in items:
            complete = [
                bool(str(item.get("description", "")).strip()),
                str(item.get("quantity", "")).strip() not in {"", "None"},
                str(item.get("unit_price", "")).strip() not in {"", "None"},
                str(item.get("line_total", "")).strip() not in {"", "None"},
            ]
            item_scores.append(sum(complete) / len(complete))
            try:
                qty = decimal_value(item.get("quantity"), "quantity")
                price = money(item.get("unit_price"), "unit price")
                total = money(item.get("line_total"), "line total")
                arithmetic_scores.append(1.0 if abs((qty * price).quantize(MONEY) - total) <= Decimal("0.05") else 0.0)
            except Exception:
                arithmetic_scores.append(0.0)
        item_score = sum(item_scores) / len(item_scores)
        arithmetic_score = sum(arithmetic_scores) / len(arithmetic_scores)
        score = 0.25 + 0.25 * header_score + 0.30 * item_score + 0.15 * arithmetic_score
        return max(0.0, min(0.95, score))

    def _normalize_payload_confidence(self, payload: dict[str, Any]) -> float:
        items = payload.get("items") if isinstance(payload.get("items"), list) else []
        global_conf = self._confidence_value(payload.get("extraction_confidence"))
        positive = [self._confidence_value(i.get("confidence")) for i in items if self._confidence_value(i.get("confidence")) > 0]
        if global_conf <= 0 and positive:
            global_conf = sum(positive) / len(positive)
        if global_conf <= 0:
            global_conf = self._quality_score(payload)
            payload.setdefault("extraction_notes", []).append(
                "The extractor omitted confidence; a conservative deterministic score was calculated."
            )
        for item in items:
            item_conf = self._confidence_value(item.get("confidence"))
            item["confidence"] = round(item_conf or global_conf, 4)
        payload["extraction_confidence"] = round(global_conf, 4)
        return global_conf

    def _find_labeled_value(self, raw_text: str, labels: Sequence[str], value_pattern: str) -> str:
        for label in labels:
            match = re.search(rf"(?:^|\b){label}\s*(?:[:#-]\s*)?({value_pattern})", raw_text, re.I | re.M)
            if match:
                return match.group(1).strip()
        return ""

    def _find_amount(self, raw_text: str, labels: Sequence[str]) -> str:
        return self._money_from_text(self._find_labeled_value(raw_text, labels, self.MONEY_PATTERN))

    def _parse_text_locally(self, raw_text: str, source: Path) -> tuple[dict[str, Any], float]:
        lines = [re.sub(r"\s+", " ", line).strip() for line in (raw_text or "").splitlines()]
        lines = [line for line in lines if line and not line.startswith("--- PAGE")]
        vendor = ""
        for line in lines[:12]:
            if re.search(r"\bBILL\s+TO\b", line, re.I):
                break
            if "INVOICE" in line.upper():
                before = re.split(r"\b(?:DELIVERY\s+)?INVOICE\b", line, maxsplit=1, flags=re.I)[0].strip(" :-")
                if (
                    before
                    and not vendor
                    and not re.search(r"\d{5}", before)
                    and before.upper() not in {"SUPPLIER", "VENDOR", "DELIVERY"}
                ):
                    vendor = before
                    break
            if not vendor and not re.search(r"invoice|bill to|terms|phone|\d{3}[- )]", line, re.I):
                vendor = line
        invoice_number = self._find_labeled_value(
            raw_text,
            [r"Invoice\s*#", r"Invoice\s*(?:No\.?|Number)", r"Invoice"],
            r"[A-Za-z0-9][A-Za-z0-9._/-]*",
        )
        date_raw = self._find_labeled_value(raw_text, [r"Invoice\s*Date", r"Date"], r"(?:\d{4}-\d{1,2}-\d{1,2}|\d{1,2}/\d{1,2}/\d{2,4}|[A-Za-z]{3,9}\s+\d{1,2},?\s+\d{4})")
        invoice_date = self._date_from_text(date_raw)
        subtotal = self._find_amount(raw_text, [r"Subtotal"])
        delivery = self._find_amount(raw_text, [r"Delivery\s*Fee", r"Freight"])
        fuel = self._find_amount(raw_text, [r"Fuel\s*Surcharge"])
        tax = self._find_amount(raw_text, [r"Sales\s*Tax", r"Tax"])
        credits = self._find_amount(raw_text, [r"Credits?", r"Credit"])
        total = self._find_amount(
            raw_text,
            [r"Invoice\s*Total", r"Amount\s*Due", r"Grand\s*Total", r"Total"],
        )
        try:
            fees = f"{(Decimal(delivery or '0') + Decimal(fuel or '0')).quantize(MONEY):.2f}"
        except Exception:
            fees = delivery or fuel or "0.00"
        items: list[dict[str, Any]] = []
        row_re = re.compile(
            r"^([A-Z0-9][A-Z0-9._/-]{1,30})\s+(.+)\s+(-?\d+(?:\.\d+)?)\s+([A-Za-z][A-Za-z0-9 /._-]{0,18})\s+\$?([\d,]+\.\d{2})\s+\$?([\d,]+\.\d{2})$",
            re.I,
        )
        for line in lines:
            if re.search(r"^(SKU\b|Item\b.*Description|Subtotal\b|Delivery\b|Fuel\b|Tax\b|Invoice\s*Total|Synthetic test)", line, re.I):
                continue
            match = row_re.match(line)
            if not match:
                continue
            sku, desc, qty, unit, unit_price, line_total = match.groups()
            unit_price_s = self._money_from_text(unit_price)
            line_total_s = self._money_from_text(line_total)
            confidence = 0.985
            try:
                expected = (Decimal(qty) * Decimal(unit_price_s)).quantize(MONEY)
                if abs(expected - Decimal(line_total_s)) > Decimal("0.05"):
                    confidence = 0.84
            except Exception:
                confidence = 0.80
            items.append({
                "sku": sku.strip(), "description": desc.strip(), "category": "Unclassified",
                "quantity": qty, "unit": unit.strip(), "unit_price": unit_price_s,
                "line_total": line_total_s, "confidence": confidence,
            })
        if not subtotal and items:
            subtotal = f"{sum((Decimal(i['line_total']) for i in items), Decimal('0')).quantize(MONEY):.2f}"
        if not total and subtotal:
            try:
                total = f"{(Decimal(subtotal) + Decimal(fees or '0') + Decimal(tax or '0') - Decimal(credits or '0')).quantize(MONEY):.2f}"
            except Exception:
                pass
        payload = {
            "vendor": vendor, "invoice_number": invoice_number, "invoice_date": invoice_date,
            "subtotal": subtotal, "fees": fees or "0.00", "tax": tax or "0.00",
            "credits": credits or "0.00", "total": total,
            "currency": str(self.settings.get("currency", "USD")),
            "source_file": source.name, "source_link": str(source.resolve()),
            "document_type": "text_pdf",
            "layout_recognized": bool(vendor and invoice_number and invoice_date and total and items),
            "extraction_notes": ["Text was extracted deterministically with PyMuPDF and parsed locally."],
            "items": items, "_raw_text": raw_text,
        }
        confidence = self._quality_score(payload)
        if payload["layout_recognized"] and confidence >= 0.90:
            confidence = 0.98
        payload["extraction_confidence"] = confidence
        for item in items:
            item["confidence"] = max(float(item.get("confidence", 0)), min(confidence, 0.985))
        return payload, confidence

    def _find_tesseract(self) -> str:
        """Find a locally installed Tesseract executable."""
        configured = str(self.settings.get("tesseract_executable") or "").strip()
        candidates = [configured, shutil.which("tesseract") or ""]
        if os.name == "nt":
            candidates.extend(
                [
                    str(Path(os.environ.get("PROGRAMFILES", r"C:\Program Files")) / "Tesseract-OCR" / "tesseract.exe"),
                    str(Path(os.environ.get("LOCALAPPDATA", "")) / "Programs" / "Tesseract-OCR" / "tesseract.exe"),
                ]
            )
        for candidate in candidates:
            if candidate and Path(candidate).expanduser().is_file():
                return str(Path(candidate).expanduser().resolve())
        return ""

    def _local_ocr_fallback(self, source: Path) -> tuple[dict[str, Any], float, str]:
        """Extract a scan locally.

        RapidOCR runs in-process with a memoized engine so the ONNX model is
        loaded once per session instead of once per file (the old design
        spawned a subprocess per document, which re-loaded the model 35 times
        for a 35-invoice folder and could take 30+ minutes on a low-end PC).
        Tesseract remains a fully-local subprocess fallback.
        """
        if not bool(self.settings.get("local_ocr_enabled", True)):
            raise ExtractionFailed("Local OCR is disabled.")

        timeout = min(300, max(30, int(self.settings.get("local_ocr_timeout_seconds", 120))))
        with tempfile.TemporaryDirectory(prefix="marginmise-local-ocr-") as temp_name:
            temp = Path(temp_name)
            images: list[Path] = []
            if source.suffix.lower() == ".pdf":
                if fitz is None:
                    raise ExtractionFailed("PyMuPDF is required to render scanned PDF pages.")
                dpi = max(150, min(400, int(self.settings.get("pdf_render_dpi", 200))))
                matrix = fitz.Matrix(dpi / 72.0, dpi / 72.0)
                with fitz.open(source) as document:
                    for index, page in enumerate(document):
                        image = temp / f"page-{index + 1:03d}.png"
                        page.get_pixmap(matrix=matrix, alpha=False).save(image)
                        images.append(image)
            else:
                images.append(source)

            failures: list[str] = []
            if bool(self.settings.get("rapidocr_enabled", True)):
                try:
                    result = _run_rapidocr_in_process(images, timeout=timeout)
                    raw_text = str(result.get("text") or "").strip()
                    if not self._has_meaningful_text(raw_text):
                        raise ExtractionFailed("RapidOCR did not extract usable invoice text.")
                    payload, confidence = self._parse_text_locally(raw_text, source)
                    payload["document_type"] = "scanned_pdf" if source.suffix.lower() == ".pdf" else "image"
                    payload["_raw_text"] = raw_text
                    ocr_confidence = self._confidence_value(result.get("average_confidence"))
                    payload.setdefault("extraction_notes", []).append(
                        f"RapidOCR processed the scan locally (average text confidence {ocr_confidence:.1%})."
                    )
                    payload["extraction_confidence"] = confidence
                    return payload, confidence, "rapidocr-onnx+local-parser"
                except Exception as exc:
                    failures.append(f"RapidOCR: {exc}")

            executable = self._find_tesseract()
            if not executable:
                detail = "; ".join(failures) or "no local engine was available"
                raise ExtractionFailed(f"Local OCR failed: {detail}; Tesseract is not installed.")
            page_texts: list[str] = []
            for index, image in enumerate(images):
                completed = subprocess.run(
                    [executable, str(image), "stdout", "-l", "eng", "--psm", "6"],
                    capture_output=True,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    timeout=timeout,
                )
                if completed.returncode != 0:
                    detail = (completed.stderr or completed.stdout or "no error text").strip()
                    raise ExtractionFailed(
                        f"Tesseract failed on page {index + 1}: {detail[-800:]}"
                    )
                page_texts.append(f"--- PAGE {index + 1} ---\n{completed.stdout.strip()}")

        raw_text = "\n\n".join(page_texts).strip()
        if not self._has_meaningful_text(raw_text):
            raise ExtractionFailed("Tesseract did not extract usable invoice text.")
        payload, confidence = self._parse_text_locally(raw_text, source)
        payload["document_type"] = "scanned_pdf" if source.suffix.lower() == ".pdf" else "image"
        payload["_raw_text"] = raw_text
        payload.setdefault("extraction_notes", []).append(
            "Locally provisioned Tesseract OCR was used after RapidOCR was unavailable."
        )
        payload["extraction_confidence"] = confidence
        if failures:
            payload.setdefault("extraction_notes", []).extend(failures)
        return payload, confidence, "tesseract+local-parser"

    def _tesseract_fallback(self, source: Path) -> tuple[dict[str, Any], float, str]:
        """Backward-compatible alias for older settings and callers."""
        return self._local_ocr_fallback(source)

    def _extract_pdf(self, source: Path) -> tuple[dict[str, Any], float, str]:
        if fitz is None:
            raise ExtractionFailed("PyMuPDF is missing. Run the application installer again.")
        max_pages = int(self.settings.get("max_pdf_pages", 30))
        with fitz.open(source) as document:
            if document.page_count < 1:
                raise ExtractionFailed("The PDF contains no pages.")
            if document.page_count > max_pages:
                raise ExtractionFailed(f"The PDF has {document.page_count} pages; maximum is {max_pages}.")
            page_texts = [page.get_text("text", sort=True) or "" for page in document]
        raw_text = "\n\n".join(f"--- PAGE {i+1} ---\n{page_text}" for i, page_text in enumerate(page_texts)).strip()
        meaningful_pages = sum(self._has_meaningful_text(page_text) for page_text in page_texts)
        if meaningful_pages == len(page_texts):
            payload, confidence = self._parse_text_locally(raw_text, source)
            return payload, confidence, "pymupdf+local-parser"
        try:
            payload, confidence, method = self._local_ocr_fallback(source)
            payload["document_type"] = "scanned_pdf" if meaningful_pages == 0 else "mixed_pdf"
            return payload, confidence, method
        except Exception as local_exc:
            if self._has_meaningful_text(raw_text):
                local, confidence = self._parse_text_locally(raw_text, source)
                local["layout_recognized"] = False
                local.setdefault("extraction_notes", []).append(
                    f"Some pages needed OCR, but local OCR failed: {local_exc}"
                )
                return local, min(confidence, 0.78), "partial-pymupdf-review"
            raise ExtractionFailed(
                f"The PDF has no usable text layer and local OCR failed: {local_exc}"
            ) from local_exc

    def extract(self, source: Path, raw_text: str = "") -> tuple[dict[str, Any], float, str]:
        suffix = source.suffix.lower()
        if suffix == ".pdf":
            return self._extract_pdf(source)
        if suffix in {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp"}:
            return self._local_ocr_fallback(source)
        raise ExtractionFailed(f"Unsupported source type: {source.suffix}")

class InvoiceExtractor:
    """Route documents through local extraction, then attach recognition state."""

    def __init__(self, workspace: RestaurantWorkspace):
        self.workspace = workspace
        self.settings = workspace.load_settings()
        self.local = LocalExtractor(self.settings, workspace.folders["logs"])

    def extract(self, source: Path) -> ExtractionResult:
        max_bytes = max(1, int(self.settings.get("max_source_file_mb", 50))) * 1024 * 1024
        try:
            source_size = source.stat().st_size
        except OSError as exc:
            raise ExtractionFailed(f"Could not inspect source file: {source}") from exc
        if source_size > max_bytes:
            raise ExtractionFailed(
                f"Source file is {source_size / 1024 / 1024:.1f} MiB; "
                f"the configured limit is {max_bytes / 1024 / 1024:.0f} MiB."
            )
        suffix = source.suffix.lower()
        if suffix == ".json":
            data = json.loads(source.read_text(encoding="utf-8"))
            vendor = str(data.get("vendor") or "")
            recognized, parser_name = self.workspace.vendor_recognition(vendor)
            confidence = LocalExtractor._confidence_value(data.get("extraction_confidence")) or 0.99
            return ExtractionResult(data, "json", confidence, recognized, parser_name)
        if suffix not in {".pdf", ".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp"}:
            raise ExtractionFailed(f"Unsupported source type: {source.suffix}")

        data, confidence, method = self.local.extract(source)
        vendor = str(data.get("vendor") or "")
        recognized, parser_name = self.workspace.vendor_recognition(vendor)
        warnings: list[str] = []
        notes = data.get("extraction_notes")
        if isinstance(notes, list):
            # Keep every note in canonical JSON for diagnostics, but only create
            # a review warning when the note describes an actual problem. A
            # successful extraction method is evidence, not paperwork for a
            # manager to dismiss.
            warning_terms = (
                "fail", "error", "missing", "unusable", "unfamiliar",
                "invalid", "timeout", "timed out", "could not", "required",
                "low confidence", "instead of the required artifact",
            )
            for note in notes:
                note_text = str(note).strip()
                if note_text and any(term in note_text.lower() for term in warning_terms):
                    warnings.append(note_text)
        if data.get("layout_recognized") is False:
            warnings.append(
                "The extractor identified an unfamiliar invoice layout; manager review is required before first posting."
            )
        return ExtractionResult(
            data, method, confidence, recognized, parser_name,
            warnings=warnings, raw_text=str(data.get("_raw_text") or ""),
        )


class InvoiceValidator:
    def __init__(self, settings: dict[str, Any]):
        self.settings = settings
        self.tolerance = money(settings.get("invoice_math_tolerance", 0.05), "invoice math tolerance")
        self.min_confidence = Decimal(str(settings.get("minimum_extraction_confidence", 0.82)))

    def canonicalize(self, data: dict[str, Any], source: Path) -> dict[str, Any]:
        canonical = dict(data)
        canonical["vendor"] = str(data.get("vendor") or "").strip()
        canonical["invoice_number"] = str(data.get("invoice_number") or "").strip()
        date_value = str(data.get("invoice_date") or "").strip()
        canonical["invoice_date"] = parse_date(date_value) if date_value else ""
        for field_name in ("subtotal", "fees", "tax", "credits", "total"):
            required = field_name in {"subtotal", "total"}
            canonical[field_name] = money_string(data.get(field_name), field_name, required=required)
        canonical["currency"] = str(data.get("currency") or self.settings.get("currency", "USD"))
        canonical["source_file"] = str(data.get("source_file") or source.name)
        canonical["source_link"] = str(data.get("source_link") or source.resolve())
        canonical_items: list[dict[str, Any]] = []
        items = data.get("items") if isinstance(data.get("items"), list) else []
        for index, item in enumerate(items, 1):
            confidence = float(item.get("confidence", 0.0) or 0.0)
            canonical_items.append({
                "sku": str(item.get("sku") or "").strip(),
                "description": str(item.get("description") or "").strip(),
                "category": str(item.get("category") or "Unclassified").strip() or "Unclassified",
                "quantity": str(decimal_value(item.get("quantity"), f"line {index} quantity")),
                "unit": str(item.get("unit") or "each").strip() or "each",
                "unit_price": money_string(item.get("unit_price"), f"line {index} unit price"),
                "line_total": money_string(item.get("line_total"), f"line {index} line total"),
                "confidence": max(0.0, min(1.0, confidence)),
            })
        canonical["items"] = canonical_items
        return canonical

    def validate(self, data: dict[str, Any], extraction_confidence: float) -> list[Finding]:
        findings: list[Finding] = []
        vendor = str(data.get("vendor") or "").strip()
        invoice_number = str(data.get("invoice_number") or "").strip()
        invoice_date = str(data.get("invoice_date") or "").strip()
        items = data.get("items") if isinstance(data.get("items"), list) else []
        if not vendor:
            findings.append(Finding("ERROR", "Invoice Header", "Vendor is missing"))
        if not invoice_number:
            findings.append(Finding("ERROR", "Invoice Header", "Invoice number is missing"))
        if not invoice_date:
            findings.append(Finding("ERROR", "Invoice Header", "Invoice date is missing"))
        if not items:
            findings.append(Finding("ERROR", "Invoice Header", "No invoice line items were extracted"))
        if Decimal(str(extraction_confidence)) < self.min_confidence:
            findings.append(Finding(
                "ERROR", "Extraction Confidence",
                f"Extraction confidence {extraction_confidence:.2f} is below {self.min_confidence}"
            ))

        line_sum = Decimal("0.00")
        for index, item in enumerate(items, 1):
            description = str(item.get("description") or "").strip()
            sku = str(item.get("sku") or "").strip()
            quantity = decimal_value(item.get("quantity"), f"line {index} quantity")
            unit_price = money(item.get("unit_price"), f"line {index} unit price")
            line_total = money(item.get("line_total"), f"line {index} line total")
            confidence = Decimal(str(item.get("confidence", 0)))
            line_sum += line_total
            if not description:
                findings.append(Finding("ERROR", "Invoice Line", f"Line {index} description is missing", index))
            if not sku:
                findings.append(Finding("WARNING", "Invoice Line", f"Line {index} vendor SKU is missing", index))
            if quantity <= 0:
                findings.append(Finding("ERROR", "Invoice Line", f"Line {index} quantity must be positive", index))
            expected = (quantity * unit_price).quantize(MONEY, rounding=ROUND_HALF_UP)
            if abs(expected - line_total) > self.tolerance:
                findings.append(Finding(
                    "ERROR", "Invoice Line",
                    f"Line {index} arithmetic mismatch: quantity x unit price is {expected:.2f}, line total is {line_total:.2f}",
                    index,
                ))
            if confidence < self.min_confidence:
                findings.append(Finding(
                    "ERROR", "Line Confidence",
                    f"Line {index} confidence {confidence:.2f} is below {self.min_confidence}", index,
                ))

        if items:
            subtotal = money(data.get("subtotal"), "subtotal")
            fees = money(data.get("fees"), "fees", required=False)
            tax = money(data.get("tax"), "tax", required=False)
            credits = money(data.get("credits"), "credits", required=False)
            total = money(data.get("total"), "total")
            if abs(line_sum - subtotal) > self.tolerance:
                findings.append(Finding(
                    "ERROR", "Invoice Header",
                    f"Line total sum {line_sum:.2f} does not match subtotal {subtotal:.2f}",
                ))
            calculated_total = (subtotal + fees + tax - credits).quantize(MONEY)
            if abs(calculated_total - total) > self.tolerance:
                findings.append(Finding(
                    "ERROR", "Invoice Header",
                    f"Subtotal + fees + tax - credits is {calculated_total:.2f}, invoice total is {total:.2f}",
                ))
        return findings


class InvoicePipeline:
    def __init__(self, workspace: RestaurantWorkspace):
        self.workspace = workspace
        self.settings = workspace.load_settings()
        self.extractor = InvoiceExtractor(workspace)
        self.validator = InvoiceValidator(self.settings)
        self.planning = InventoryPlanningService(workspace)
        self.controls = OperationalControlsService(workspace)
        self.phase2 = Phase2Service(workspace, self.planning, self.controls)
        self.phase3 = Phase3Service(workspace, self.planning, self.controls, self.phase2)
        self.recipe_costing = RecipeCostingService(workspace)
        self.margin_memory = MarginMemoryService(workspace, self.planning, self.controls)
        self.review_copilot = ReviewCopilotService(workspace, self, self.controls)

    @staticmethod
    def _normalized_ocr_text(raw_text: str) -> str:
        text = str(raw_text or "").replace("\u00a0", " ")
        # OCR sometimes spaces every character in a heading. Normalize only the
        # labels used for header recovery, leaving invoice values untouched.
        substitutions = {
            r"\bI\s*N\s*V\s*O\s*I\s*C\s*E\b": "INVOICE",
            r"\bN\s*U\s*M\s*B\s*E\s*R\b": "NUMBER",
            r"\bD\s*A\s*T\s*E\b": "DATE",
        }
        for pattern, replacement in substitutions.items():
            text = re.sub(pattern, replacement, text, flags=re.I)
        return text

    @staticmethod
    def _date_token_pattern() -> str:
        months = (
            r"Jan(?:uary)?|Feb(?:ruary)?|Mar(?:ch)?|Apr(?:il)?|May|Jun(?:e)?|"
            r"Jul(?:y)?|Aug(?:ust)?|Sep(?:t(?:ember)?)?|Oct(?:ober)?|"
            r"Nov(?:ember)?|Dec(?:ember)?"
        )
        return (
            rf"(?:\d{{4}}[./-]\d{{1,2}}[./-]\d{{1,2}}|"
            rf"\d{{1,2}}[./-]\d{{1,2}}[./-]\d{{2,4}}|"
            rf"(?:{months})[\s-]+\d{{1,2}}(?:st|nd|rd|th)?[,]?[\s-]+\d{{2,4}}|"
            rf"\d{{1,2}}(?:st|nd|rd|th)?[\s-]+(?:{months})[,]?[\s-]+\d{{2,4}})"
        )

    def _raw_text_for_recovery(self, data: dict[str, Any], source: Path | None = None) -> str:
        raw_text = str(data.get("_raw_text") or "")
        raw_path = str(data.get("_raw_text_path") or "").strip()
        if not raw_text and raw_path:
            try:
                candidate = Path(raw_path)
                if candidate.exists() and candidate.is_file():
                    raw_text = candidate.read_text(encoding="utf-8", errors="replace")
            except OSError:
                pass
        if not raw_text and source and source.exists() and source.suffix.lower() == ".pdf" and fitz is not None:
            try:
                with fitz.open(source) as document:
                    raw_text = "\n\n".join(page.get_text("text", sort=True) or "" for page in document)
            except Exception:
                pass
        return self._normalized_ocr_text(raw_text)

    @staticmethod
    def _clean_invoice_number_candidate(value: str) -> str:
        candidate = str(value or "").strip().strip(":#=,;[](){}")
        candidate = re.sub(r"\s+", "", candidate)
        if not candidate or len(candidate) > 48:
            return ""
        if candidate.upper() in {"DATE", "TOTAL", "NUMBER", "NO", "N/A", "NA", "INVOICE"}:
            return ""
        if re.fullmatch(r"\d{1,2}[./-]\d{1,2}[./-]\d{2,4}", candidate):
            return ""
        if not re.search(r"\d", candidate):
            return ""
        if not re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9._/\-]*", candidate):
            return ""
        return candidate

    def _recover_invoice_number(self, raw_text: str, source_name: str = "") -> tuple[str, str]:
        text = self._normalized_ocr_text(raw_text)
        patterns = (
            r"\b(?:invoice|inv)\s*(?:number|num(?:ber)?|no\.?|#|id)\s*[:#=\-]?\s*([A-Za-z0-9][A-Za-z0-9._/\-]{1,47})",
            r"\b(?:document|bill)\s*(?:number|no\.?|#)\s*[:#=\-]?\s*([A-Za-z0-9][A-Za-z0-9._/\-]{1,47})",
            r"\binvoice\s*[:#=\-]\s*([A-Za-z0-9][A-Za-z0-9._/\-]{1,47})",
            r"\binvoice\s+([A-Za-z0-9][A-Za-z0-9._/\-]{2,47})",
        )
        for pattern in patterns:
            for match in re.finditer(pattern, text, flags=re.I):
                candidate = self._clean_invoice_number_candidate(match.group(1))
                if candidate:
                    return candidate, match.group(0).strip()
        stem = Path(source_name or "").stem
        filename_match = re.search(
            r"(?:^|[_\-\s])(?:invoice|inv)[_\-\s#]*([A-Za-z0-9][A-Za-z0-9._\-]{1,47})",
            stem,
            flags=re.I,
        )
        if filename_match:
            candidate = self._clean_invoice_number_candidate(filename_match.group(1))
            if candidate:
                return candidate, f"filename:{stem}"
        return "", ""

    def _recover_invoice_date(self, raw_text: str, source_name: str = "") -> tuple[str, str]:
        text = self._normalized_ocr_text(raw_text)
        token = self._date_token_pattern()
        labeled_patterns = (
            rf"\b(?:invoice\s*date|inv\.?\s*date|date\s*of\s*invoice|document\s*date|billing\s*date|date\s*issued|issued\s*on)\s*[:#=\-]?\s*({token})",
            rf"(?m)^\s*date\s*[:#=\-]\s*({token})",
        )
        for pattern in labeled_patterns:
            for match in re.finditer(pattern, text, flags=re.I):
                try:
                    return parse_date(match.group(1)), match.group(0).strip()
                except ValueError:
                    continue

        excluded = re.compile(r"\b(?:due|delivery|delivered|ship|shipped|order|purchase\s*order|po|service\s*period|statement)\b", re.I)
        candidates: list[tuple[int, str, str]] = []
        lines = [re.sub(r"\s+", " ", line).strip() for line in text.splitlines() if line.strip()]
        for index, line in enumerate(lines[:60]):
            for match in re.finditer(token, line, flags=re.I):
                context = line[max(0, match.start() - 35):match.end() + 15]
                if excluded.search(context):
                    continue
                try:
                    parsed = parse_date(match.group(0))
                except ValueError:
                    continue
                score = max(0, 45 - index)
                if re.search(r"\binvoice\b", line, re.I):
                    score += 45
                if re.search(r"\bdate\b", line, re.I):
                    score += 25
                candidates.append((score, parsed, line))
        if candidates:
            candidates.sort(key=lambda item: item[0], reverse=True)
            return candidates[0][1], candidates[0][2]

        stem = Path(source_name or "").stem
        for match in re.finditer(token, stem, flags=re.I):
            try:
                return parse_date(match.group(0)), f"filename:{stem}"
            except ValueError:
                continue
        return "", ""

    def recover_missing_invoice_headers(
        self,
        data: dict[str, Any],
        source: Path | None = None,
    ) -> tuple[dict[str, Any], dict[str, Any]]:
        recovered = dict(data)
        source_name = source.name if source else str(data.get("source_file") or "")
        raw_text = self._raw_text_for_recovery(data, source)
        details: dict[str, Any] = {
            "recovered_fields": [], "unresolved_fields": [], "evidence": {},
            "raw_text_available": bool(raw_text),
        }

        invoice_number = str(recovered.get("invoice_number") or "").strip()
        if not invoice_number:
            invoice_number, evidence = self._recover_invoice_number(raw_text, source_name)
            if invoice_number:
                recovered["invoice_number"] = invoice_number
                details["recovered_fields"].append("invoice_number")
                details["evidence"]["invoice_number"] = evidence
            else:
                details["unresolved_fields"].append("invoice_number")

        current_date = str(recovered.get("invoice_date") or "").strip()
        try:
            normalized_date = parse_date(current_date) if current_date else ""
        except ValueError:
            normalized_date = ""
        if normalized_date:
            recovered["invoice_date"] = normalized_date
        else:
            invoice_date, evidence = self._recover_invoice_date(raw_text, source_name)
            if invoice_date:
                recovered["invoice_date"] = invoice_date
                details["recovered_fields"].append("invoice_date")
                details["evidence"]["invoice_date"] = evidence
            else:
                recovered["invoice_date"] = ""
                details["unresolved_fields"].append("invoice_date")

        if details["recovered_fields"]:
            notes = recovered.get("extraction_notes")
            if not isinstance(notes, list):
                notes = [str(notes)] if notes else []
            notes.append(
                "Automatic second-pass header recovery found: "
                + ", ".join(details["recovered_fields"])
                + "."
            )
            recovered["extraction_notes"] = notes
        recovered["_header_recovery"] = details
        if raw_text and not recovered.get("_raw_text"):
            recovered["_raw_text"] = raw_text
        return recovered, details

    def _persist_recovered_review(
        self,
        invoice_id: str,
        canonical: dict[str, Any],
        findings: Iterable[Finding],
        confidence: float,
    ) -> None:
        self._replace_review_payload(invoice_id, canonical, findings)
        with self.workspace.connect() as conn:
            conn.execute(
                """UPDATE invoices SET vendor=?,invoice_number=?,invoice_date=?,subtotal=?,fees=?,tax=?,credits=?,total=?,
                   extraction_confidence=?,canonical_json=?,updated_at=? WHERE invoice_id=?""",
                (
                    canonical.get("vendor", ""), canonical.get("invoice_number", ""), canonical.get("invoice_date", ""),
                    canonical.get("subtotal", ""), canonical.get("fees", "0.00"), canonical.get("tax", "0.00"),
                    canonical.get("credits", "0.00"), canonical.get("total", ""), float(confidence),
                    json.dumps(json_safe(canonical)), now_iso(), invoice_id,
                ),
            )

    def recover_review_invoice(
        self,
        invoice_id: str,
        *,
        approve_eligible: bool = True,
        explicit_approval: bool = False,
        recognize_vendor: bool = False,
    ) -> ProcessResult:
        self.controls.require_permission("invoices.review")
        row = self.get_invoice(invoice_id)
        if not row:
            return ProcessResult(source="", invoice_id=invoice_id, status="Failed", errors=["Invoice not found"])
        source = Path(row["source_original_path"] or row["source_archive_path"] or row["source_name"] or "invoice")
        data = self.get_invoice_data(invoice_id)
        recovered, recovery = self.recover_missing_invoice_headers(data, source)
        result = ProcessResult(
            source=str(source), invoice_id=invoice_id, status="Needs Review",
            extraction_method=row["extraction_method"] or "header-recovery",
            extraction_confidence=float(row["extraction_confidence"] or 0.0),
            recognized_vendor=bool(row["recognized_vendor"]),
        )
        try:
            canonical = self.validator.canonicalize(recovered, source)
            recovered_quality = self.extractor.local._quality_score(canonical)
            confidence = max(float(row["extraction_confidence"] or 0.0), float(recovered_quality))
            findings = self.validator.validate(canonical, 1.0 if explicit_approval else confidence)
        except Exception as exc:
            result.errors.append(str(exc))
            result.message = "Automatic header recovery could not create a valid invoice record."
            return result

        errors = [finding for finding in findings if finding.severity == "ERROR"]
        warnings = [finding for finding in findings if finding.severity == "WARNING"]
        if recovery.get("recovered_fields"):
            self.margin_memory.capture_invoice_correction(
                invoice_id, data, canonical, correction_source="Automatic Header Recovery"
            )
        if approve_eligible and not errors:
            approved = self.approve_review(
                invoice_id,
                canonical,
                recognize_vendor=recognize_vendor,
                allow_warning_override=True,
            )
            if approved.status == "Approved":
                approved.message = (
                    "Recovered missing invoice header data and approved the invoice automatically."
                    if recovery["recovered_fields"] else
                    "Invoice passed batch validation and was approved."
                )
            return approved

        self._persist_recovered_review(invoice_id, canonical, findings, confidence)
        result.errors.extend(f.issue for f in errors)
        result.warnings.extend(f.issue for f in warnings)
        if recovery["unresolved_fields"]:
            result.message = "Still needs review: " + ", ".join(recovery["unresolved_fields"])
        elif errors:
            result.message = "Header recovery completed, but other validation errors still require review."
        else:
            result.message = "Invoice is ready for approval."
        return result

    def batch_process_reviews(
        self,
        invoice_ids: Iterable[str] | None = None,
        *,
        approve_eligible: bool = True,
        explicit_approval: bool = False,
        recognize_vendors: bool = False,
    ) -> dict[str, Any]:
        self.controls.require_permission("invoices.review")
        ids = (
            [row["invoice_id"] for row in self.list_open_reviews()]
            if invoice_ids is None
            else [str(value) for value in invoice_ids]
        )
        summary: dict[str, Any] = {
            "requested": len(ids), "approved": 0, "needs_review": 0,
            "duplicates": 0, "failed": 0, "results": [],
        }
        if not ids:
            return summary
        for invoice_id in ids:
            try:
                result = self.recover_review_invoice(
                    invoice_id,
                    approve_eligible=approve_eligible,
                    explicit_approval=explicit_approval,
                    recognize_vendor=recognize_vendors,
                )
            except Exception as exc:
                result = ProcessResult(source="", invoice_id=invoice_id, status="Failed", errors=[str(exc)])
            key = {
                "Approved": "approved", "Needs Review": "needs_review",
                "Duplicate": "duplicates", "Failed": "failed",
            }.get(result.status, "failed")
            summary[key] += 1
            summary["results"].append(result.as_dict())
        self.controls.audit(
            "invoice.batch_review", "invoice_batch", None,
            f"Batch review processed {len(ids)} invoice(s): {summary['approved']} approved, {summary['needs_review']} still need review.",
            details={key: value for key, value in summary.items() if key != "results"},
        )
        return summary

    def _auto_verify_invoice_if_enabled(self, invoice_id: str) -> dict[str, Any] | None:
        settings = self.workspace.load_settings()
        if not bool(settings.get("receiving_verification_enabled", True)):
            return None
        if not bool(settings.get("auto_verify_clean_receiving", True)):
            return None
        try:
            return self.controls.auto_verify_receiving(
                [invoice_id],
                date_mode=str(settings.get("auto_verify_receiving_date_mode") or "invoice_date"),
            )
        except Exception:
            return None

    def reload_settings(self) -> dict[str, Any]:
        """Reload settings without replacing database-backed service objects."""
        self.settings = self.workspace.load_settings()
        self.extractor = InvoiceExtractor(self.workspace)
        self.validator = InvoiceValidator(self.settings)
        return dict(self.settings)

    def process_file(self, source: Path) -> ProcessResult:
        source = source.expanduser().resolve()
        result = ProcessResult(source=str(source))
        if not source.exists() or not source.is_file():
            result.errors.append("Source file does not exist")
            return result
        if source.suffix.lower() not in SUPPORTED_SOURCE_SUFFIXES:
            result.errors.append(f"Unsupported source type: {source.suffix}")
            return result

        source_hash = sha256_file(source)
        archive_path = self.workspace.archive_original(source, source_hash)
        try:
            extraction = self.extractor.extract(source)
            result.extraction_method = extraction.method
            result.extraction_confidence = extraction.confidence
            result.recognized_vendor = extraction.vendor_recognized
            result.warnings.extend(extraction.warnings)
            if bool(self.settings.get("auto_recover_invoice_headers", True)):
                recovered_data, recovery = self.recover_missing_invoice_headers(extraction.data, source)
                if recovery["recovered_fields"]:
                    extraction.confidence = max(extraction.confidence, self.extractor.local._quality_score(recovered_data))
                    result.extraction_confidence = extraction.confidence
                    result.warnings.append("Recovered missing header field(s): " + ", ".join(recovery["recovered_fields"]))
            else:
                recovered_data = extraction.data
                recovery = {"recovered_fields": [], "unresolved_fields": [], "evidence": {}}
            canonical = self.validator.canonicalize(recovered_data, source)
        except Exception as exc:
            # Create a persistent manual-review record even when extraction completely fails.
            partial_text = ""
            if source.suffix.lower() == ".pdf" and fitz is not None:
                try:
                    with fitz.open(source) as failed_doc:
                        partial_text = "\n\n".join(page.get_text("text", sort=True) or "" for page in failed_doc)
                except Exception:
                    partial_text = ""
            canonical = {
                "vendor": "", "invoice_number": "", "invoice_date": "",
                "subtotal": "", "fees": "0.00", "tax": "0.00", "credits": "0.00", "total": "",
                "currency": self.settings.get("currency", "USD"),
                "source_file": source.name, "source_link": str(source), "items": [],
                "_raw_text": partial_text,
                "_extraction_error": str(exc),
                "extraction_notes": [str(exc)],
            }
            result.extraction_method = "local-extraction-failed"
            result.extraction_confidence = 0.0
            result.errors.append(str(exc))
            extraction = ExtractionResult(canonical, "local-extraction-failed", 0.0, False, "generic")
            if bool(self.settings.get("auto_recover_invoice_headers", True)):
                canonical, recovery = self.recover_missing_invoice_headers(canonical, source)
                if recovery["recovered_fields"]:
                    result.warnings.append("Recovered missing header field(s) from partial raw text: " + ", ".join(recovery["recovered_fields"]))
            else:
                recovery = {"recovered_fields": [], "unresolved_fields": [], "evidence": {}}

        return self._finalize_extraction(
            source=source,
            source_hash=source_hash,
            archive_path=archive_path,
            canonical=canonical,
            extraction=extraction,
            recovery=recovery,
            result=result,
        )

    def process_structured_invoice(
        self,
        source: Path,
        data: dict[str, Any],
        *,
        workbook_hash: str | None = None,
        archive_path: Path | None = None,
        group_key: str = "",
    ) -> ProcessResult:
        """Validate and post a pre-structured invoice through the normal pipeline.

        Excel invoice rows are parsed deterministically by Auto Upload, but they
        must still pass the exact same canonicalization, arithmetic, duplicate,
        review, item-creation, and receiving workflow as OCR invoices.
        """
        source = source.expanduser().resolve()
        result = ProcessResult(source=str(source))
        if not source.exists() or not source.is_file():
            result.errors.append("Source file does not exist")
            return result
        base_hash = workbook_hash or sha256_file(source)
        identity = group_key or json.dumps(json_safe(data), sort_keys=True)
        source_hash = hashlib.sha256(f"{base_hash}|{identity}".encode("utf-8")).hexdigest()
        archive_path = archive_path or self.workspace.archive_original(source, base_hash)
        vendor = str(data.get("vendor") or "").strip()
        recognized, parser_name = self.workspace.vendor_recognition(vendor)
        extraction = ExtractionResult(
            data=dict(data),
            method="structured-excel",
            confidence=0.99,
            vendor_recognized=recognized,
            parser_name=parser_name if recognized else "structured-excel",
        )
        result.extraction_method = extraction.method
        result.extraction_confidence = extraction.confidence
        result.recognized_vendor = recognized
        try:
            canonical = self.validator.canonicalize(extraction.data, source)
        except Exception as exc:
            result.errors.append(str(exc))
            result.message = "Structured invoice could not be canonicalized."
            return result
        return self._finalize_extraction(
            source=source,
            source_hash=source_hash,
            archive_path=archive_path,
            canonical=canonical,
            extraction=extraction,
            recovery={"recovered_fields": [], "unresolved_fields": [], "evidence": {}},
            result=result,
        )

    def _finalize_extraction(
        self,
        *,
        source: Path,
        source_hash: str,
        archive_path: Path,
        canonical: dict[str, Any],
        extraction: ExtractionResult,
        recovery: dict[str, Any],
        result: ProcessResult,
    ) -> ProcessResult:
        """Apply the shared deterministic invoice validation and posting path."""
        invoice_id, duplicate_key = self._invoice_identity(canonical, source_hash)
        result.invoice_id = invoice_id
        duplicate = self._find_duplicate(source_hash, duplicate_key)
        if duplicate:
            result.status = "Duplicate"
            result.message = f"Duplicate of {duplicate}"
            return result

        findings: list[Finding] = []
        try:
            findings.extend(self.validator.validate(canonical, extraction.confidence))
        except Exception as exc:
            findings.append(Finding("ERROR", "Canonical Validation", str(exc)))
        for warning in extraction.warnings:
            findings.append(Finding("WARNING", "Extraction", warning))

        errors = [finding for finding in findings if finding.severity == "ERROR"]
        warnings = [finding for finding in findings if finding.severity == "WARNING"]
        result.errors.extend(f.issue for f in errors if f.issue not in result.errors)
        result.warnings.extend(f.issue for f in warnings if f.issue not in result.warnings)

        if (
            not errors
            and not extraction.vendor_recognized
            and bool(self.settings.get("auto_learn_validated_vendors", True))
            and not bool(self.settings.get("require_review_for_unrecognized_vendors", False))
            and extraction.confidence >= float(self.settings.get("auto_approve_confidence", 0.70))
            and str(canonical.get("vendor") or "").strip()
        ):
            self.workspace.mark_vendor_recognized(str(canonical["vendor"]), extraction.parser_name)
            extraction.vendor_recognized = True
            result.recognized_vendor = True
            result.warnings.append("New vendor was learned automatically after complete arithmetic validation.")

        unrecognized_requires_review = (
            bool(self.settings.get("require_review_for_unrecognized_vendors", False))
            and not extraction.vendor_recognized
        )
        recovered_requires_review = bool(recovery.get("recovered_fields")) and not bool(
            self.settings.get("auto_approve_recovered_invoice_headers", True)
        )
        confidence_requires_review = extraction.confidence < float(
            self.settings.get("auto_approve_confidence", 0.70)
        ) and not (
            extraction.vendor_recognized
            and extraction.method == "structured-excel"
        )
        warning_requires_review = bool(warnings) and extraction.method != "structured-excel" and any(
            term in " ".join(str(getattr(w, "issue", w)).lower() for w in warnings)
            for term in ("missing", "unusable", "unfamiliar", "invalid", "could not", "fail", "error", "required")
        )
        needs_review = (
            bool(errors)
            or unrecognized_requires_review
            or recovered_requires_review
            or confidence_requires_review
            or warning_requires_review
        )
        if recovered_requires_review:
            finding = Finding(
                "WARNING", "Header Recovery",
                "Missing invoice header fields were recovered automatically and are waiting for manager approval."
            )
            findings.append(finding)
            result.warnings.append(finding.issue)
        status = "Needs Review" if needs_review else "Approved"

        if unrecognized_requires_review:
            finding = Finding(
                "WARNING", "Vendor Recognition",
                "This vendor or invoice layout has not been approved for automatic posting."
            )
            findings.append(finding)
            result.warnings.append(finding.issue)

        self._store_invoice_shell(
            invoice_id=invoice_id,
            source_hash=source_hash,
            duplicate_key=duplicate_key,
            source=source,
            archive_path=archive_path,
            canonical=canonical,
            extraction=extraction,
            status=status,
            notes="; ".join(result.errors),
        )
        self._store_reviews(invoice_id, findings, canonical)
        if recovery.get("recovered_fields"):
            self.margin_memory.capture_invoice_correction(
                invoice_id, extraction.data, canonical, correction_source="Automatic Header Recovery"
            )
        extracted_path = self.workspace.folders["extracted"] / f"{invoice_id}.json"
        extracted_path.write_text(json.dumps(json_safe(canonical), indent=2), encoding="utf-8")

        if status == "Approved":
            self._commit_lines_and_items(invoice_id, canonical, extraction)
            self.workspace.copy_to_status_folder(archive_path, "Approved", invoice_id)
            receiving = self._auto_verify_invoice_if_enabled(invoice_id)
            result.message = "Invoice validated and posted."
            if receiving and receiving.get("verified"):
                result.message += " Delivery was automatically verified as received in full."
        else:
            self.workspace.copy_to_status_folder(archive_path, "Needs Review", invoice_id)
            result.message = "Invoice preserved in the review queue."
        # Once archived and routed, remove only the disposable upload copy.
        try:
            if source.parent.resolve() == self.workspace.folders["upload"].resolve():
                source.unlink(missing_ok=True)
        except OSError:
            pass
        result.status = status
        return result

    def process_many(self, sources: Iterable[Path]) -> list[ProcessResult]:
        return [self.process_file(path) for path in sources]

    def approve_review(
        self,
        invoice_id: str,
        edited_data: dict[str, Any],
        *,
        recognize_vendor: bool = False,
        allow_warning_override: bool = True,
    ) -> ProcessResult:
        row = self.get_invoice(invoice_id)
        if not row:
            return ProcessResult(source="", invoice_id=invoice_id, status="Failed", errors=["Invoice not found"])
        source = Path(row["source_original_path"])
        result = ProcessResult(
            source=str(source), invoice_id=invoice_id,
            extraction_method=row["extraction_method"] or "manual-review",
            extraction_confidence=float(row["extraction_confidence"] or 0.0),
            recognized_vendor=bool(row["recognized_vendor"]),
        )
        try:
            canonical = self.validator.canonicalize(edited_data, source)
            # Human review is the confidence override; line confidence becomes 1.0 when omitted.
            for item in canonical.get("items", []):
                if float(item.get("confidence", 0)) <= 0:
                    item["confidence"] = 1.0
            findings = self.validator.validate(canonical, 1.0)
        except Exception as exc:
            result.errors.append(str(exc))
            return result
        errors = [f for f in findings if f.severity == "ERROR"]
        warnings = [f for f in findings if f.severity == "WARNING"]
        if errors:
            result.errors.extend(f.issue for f in errors)
            result.warnings.extend(f.issue for f in warnings)
            result.status = "Needs Review"
            result.message = "Correct the remaining validation errors before approval."
            self._replace_review_payload(invoice_id, canonical, findings)
            return result
        if warnings and not allow_warning_override:
            result.warnings.extend(f.issue for f in warnings)
            result.status = "Needs Review"
            return result

        extraction = ExtractionResult(
            canonical,
            row["extraction_method"] or "manual-review",
            1.0,
            recognize_vendor or bool(row["recognized_vendor"]),
            "generic",
        )
        _new_id, corrected_duplicate_key = self._invoice_identity(canonical, row["source_sha256"])
        with self.workspace.connect() as conn:
            collision = conn.execute(
                "SELECT invoice_id FROM invoices WHERE duplicate_key=? AND invoice_id<>?",
                (corrected_duplicate_key, invoice_id),
            ).fetchone()
            if collision:
                result.status = "Duplicate"
                result.errors.append(f"Corrected invoice duplicates {collision['invoice_id']}")
                return result
            conn.execute("DELETE FROM invoice_lines WHERE invoice_id=?", (invoice_id,))
            conn.execute("DELETE FROM price_history WHERE invoice_id=?", (invoice_id,))
            conn.execute(
                """UPDATE invoices SET vendor=?, invoice_number=?, invoice_date=?, subtotal=?, fees=?, tax=?, credits=?, total=?,
                   extraction_confidence=1.0, recognized_vendor=?, status='Approved', notes='', canonical_json=?,
                   duplicate_key=?, updated_at=? WHERE invoice_id=?""",
                (
                    canonical["vendor"], canonical["invoice_number"], canonical["invoice_date"],
                    canonical["subtotal"], canonical["fees"], canonical["tax"], canonical["credits"], canonical["total"],
                    1 if extraction.vendor_recognized else 0,
                    json.dumps(json_safe(canonical)), corrected_duplicate_key, now_iso(), invoice_id,
                ),
            )
            conn.execute(
                "UPDATE reviews SET status='Resolved', resolution='Approved by user', resolved_at=? WHERE invoice_id=? AND status='Open'",
                (now_iso(), invoice_id),
            )
        if recognize_vendor and canonical.get("vendor"):
            self.workspace.mark_vendor_recognized(canonical["vendor"])
            extraction.vendor_recognized = True
        self._commit_lines_and_items(invoice_id, canonical, extraction)
        extracted_path = self.workspace.folders["extracted"] / f"{invoice_id}.json"
        extracted_path.write_text(json.dumps(json_safe(canonical), indent=2), encoding="utf-8")
        original_source = Path(row["source_original_path"])
        if original_source.exists():
            self.workspace.copy_to_status_folder(original_source, "Approved", invoice_id)
        result.status = "Approved"
        result.recognized_vendor = extraction.vendor_recognized
        result.warnings.extend(f.issue for f in warnings)
        self.margin_memory.capture_invoice_correction(
            invoice_id, dict(row), canonical, correction_source="Manual Review"
        )
        result.message = "Reviewed invoice approved and posted."
        receiving = self._auto_verify_invoice_if_enabled(invoice_id)
        if receiving and receiving.get("verified"):
            result.message += " Delivery was automatically verified as received in full."
        return result

    def reject_review(self, invoice_id: str, reason: str) -> None:
        with self.workspace.connect() as conn:
            conn.execute(
                "UPDATE invoices SET status='Rejected', notes=?, updated_at=? WHERE invoice_id=?",
                (reason, now_iso(), invoice_id),
            )
            conn.execute(
                "UPDATE reviews SET status='Resolved', resolution=?, resolved_at=? WHERE invoice_id=? AND status='Open'",
                (f"Rejected: {reason}", now_iso(), invoice_id),
            )

    def _invoice_identity(self, data: dict[str, Any], source_hash: str) -> tuple[str, str]:
        vendor = normalize_text(data.get("vendor"))
        invoice_number = normalize_text(data.get("invoice_number"))
        invoice_date = str(data.get("invoice_date") or "")
        total = str(data.get("total") or "")
        if vendor and invoice_number and invoice_date and total:
            duplicate_key = f"{vendor}|{invoice_number}|{invoice_date}|{total}"
        else:
            duplicate_key = f"HASH|{source_hash}"
        return deterministic_id("INV", duplicate_key), duplicate_key

    def _find_duplicate(self, source_hash: str, duplicate_key: str) -> str | None:
        with self.workspace.connect() as conn:
            row = conn.execute(
                "SELECT invoice_id FROM invoices WHERE source_sha256=? OR duplicate_key=? LIMIT 1",
                (source_hash, duplicate_key),
            ).fetchone()
        return row["invoice_id"] if row else None

    def _store_invoice_shell(
        self,
        *,
        invoice_id: str,
        source_hash: str,
        duplicate_key: str,
        source: Path,
        archive_path: Path,
        canonical: dict[str, Any],
        extraction: ExtractionResult,
        status: str,
        notes: str,
    ) -> None:
        timestamp = now_iso()
        with self.workspace.connect() as conn:
            conn.execute(
                """INSERT INTO invoices(
                    invoice_id, source_sha256, duplicate_key, source_name, source_original_path, source_archive_path,
                    vendor, invoice_number, invoice_date, subtotal, fees, tax, credits, total,
                    extraction_method, extraction_confidence, recognized_vendor, status, notes, canonical_json,
                    created_at, updated_at
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (
                    invoice_id, source_hash, duplicate_key, source.name, str(archive_path), str(archive_path),
                    canonical.get("vendor", ""), canonical.get("invoice_number", ""), canonical.get("invoice_date", ""),
                    canonical.get("subtotal", ""), canonical.get("fees", "0.00"), canonical.get("tax", "0.00"),
                    canonical.get("credits", "0.00"), canonical.get("total", ""),
                    extraction.method, extraction.confidence, 1 if extraction.vendor_recognized else 0,
                    status, notes, json.dumps(json_safe(canonical)), timestamp, timestamp,
                ),
            )
            vendor_name = str(canonical.get("vendor") or "").strip()
            if vendor_name:
                vendor_key = normalize_text(vendor_name)
                conn.execute(
                    """INSERT INTO vendors(vendor_key, vendor_name, recognized, parser_name, first_seen, last_seen)
                       VALUES(?,?,?,?,?,?)
                       ON CONFLICT(vendor_key) DO UPDATE SET vendor_name=excluded.vendor_name, last_seen=excluded.last_seen""",
                    (
                        vendor_key, vendor_name, 1 if extraction.vendor_recognized else 0,
                        extraction.parser_name, timestamp, timestamp,
                    ),
                )

    def _store_reviews(self, invoice_id: str, findings: Iterable[Finding], canonical: dict[str, Any]) -> None:
        payload = json.dumps(json_safe(canonical))
        timestamp = now_iso()
        with self.workspace.connect() as conn:
            for finding in findings:
                conn.execute(
                    """INSERT INTO reviews(invoice_id, item_id, severity, issue_type, issue, payload_json, status, created_at)
                       VALUES(?,?,?,?,?,?,'Open',?)""",
                    (invoice_id, finding.item_id or None, finding.severity, finding.issue_type, finding.issue, payload, timestamp),
                )

    def _replace_review_payload(self, invoice_id: str, canonical: dict[str, Any], findings: Iterable[Finding]) -> None:
        payload = json.dumps(json_safe(canonical))
        timestamp = now_iso()
        with self.workspace.connect() as conn:
            conn.execute("DELETE FROM reviews WHERE invoice_id=? AND status='Open'", (invoice_id,))
            for finding in findings:
                conn.execute(
                    "INSERT INTO reviews(invoice_id,item_id,severity,issue_type,issue,payload_json,status,created_at) VALUES(?,?,?,?,?,?,'Open',?)",
                    (invoice_id, finding.item_id or None, finding.severity, finding.issue_type, finding.issue, payload, timestamp),
                )
            conn.execute(
                "UPDATE invoices SET canonical_json=?, updated_at=? WHERE invoice_id=?",
                (payload, timestamp, invoice_id),
            )

    def _commit_lines_and_items(
        self,
        invoice_id: str,
        canonical: dict[str, Any],
        extraction: ExtractionResult,
    ) -> None:
        vendor_name = canonical["vendor"]
        vendor_key = normalize_text(vendor_name)
        invoice_date = canonical["invoice_date"]
        price_alert_threshold = Decimal(str(self.settings.get("price_alert_percent", 5.0)))
        with self.workspace.connect() as conn:
            existing_lines = conn.execute(
                "SELECT COUNT(*) AS count FROM invoice_lines WHERE invoice_id=?", (invoice_id,)
            ).fetchone()["count"]
            if existing_lines:
                return
            for index, item in enumerate(canonical.get("items", []), 1):
                sku = str(item.get("sku") or "").strip()
                description = str(item.get("description") or "").strip()
                normalized_description = normalize_text(description)
                category = str(item.get("category") or "Unclassified")
                unit = str(item.get("unit") or "each")
                quantity = decimal_value(item.get("quantity"), f"line {index} quantity")
                unit_price = money(item.get("unit_price"), f"line {index} unit price")
                line_total = money(item.get("line_total"), f"line {index} line total")
                confidence = float(item.get("confidence", extraction.confidence) or extraction.confidence)

                item_row = None
                if sku:
                    item_row = conn.execute(
                        "SELECT * FROM items WHERE vendor_key=? AND vendor_sku=?", (vendor_key, sku)
                    ).fetchone()
                if item_row is None and normalized_description:
                    item_row = conn.execute(
                        "SELECT * FROM items WHERE vendor_key=? AND normalized_description=?",
                        (vendor_key, normalized_description),
                    ).fetchone()
                new_item = item_row is None
                if new_item:
                    identity = f"{vendor_key}|{sku or normalized_description or index}"
                    item_id = deterministic_id("ITM", identity)
                    previous_price = None
                    match_status = "New Item"
                    review_status = "New Item - Review Required"
                    purchase_count = 0
                    average_price = Decimal("0.00")
                    total_spent = Decimal("0.00")
                    lowest_price = unit_price
                    highest_price = unit_price
                    first_price = unit_price
                    first_date = invoice_date
                else:
                    item_id = item_row["item_id"]
                    previous_price = money(item_row["current_price"], "previous price", required=False) if item_row["current_price"] else None
                    match_status = "Matched" if sku and item_row["vendor_sku"] == sku else "Description Match - Review"
                    review_status = item_row["review_status"]
                    purchase_count = int(item_row["purchase_count"] or 0)
                    average_price = money(item_row["average_price"], "average price", required=False)
                    total_spent = money(item_row["total_spent"], "total spent", required=False)
                    lowest_price = money(item_row["lowest_price"], "lowest price", required=False) if item_row["lowest_price"] else unit_price
                    highest_price = money(item_row["highest_price"], "highest price", required=False) if item_row["highest_price"] else unit_price
                    first_price = money(item_row["first_price"], "first price", required=False) if item_row["first_price"] else unit_price
                    first_date = item_row["first_purchase_date"] or invoice_date

                price_change = None
                alert = 0
                if previous_price is not None and previous_price != 0:
                    price_change = ((unit_price - previous_price) / previous_price * Decimal("100")).quantize(MONEY)
                    alert = int(abs(price_change) >= price_alert_threshold)
                new_count = purchase_count + 1
                new_average = (
                    ((average_price * purchase_count) + unit_price) / Decimal(new_count)
                ).quantize(MONEY) if purchase_count else unit_price
                inferred_count_unit, inferred_units_per = infer_count_conversion(description, unit)
                conn.execute(
                    """INSERT INTO items(
                        item_id,vendor_key,vendor_name,vendor_sku,item_name,normalized_description,category,unit,
                        first_purchase_date,last_purchase_date,first_price,previous_price,current_price,price_change_percent,
                        average_price,lowest_price,highest_price,total_spent,purchase_count,review_status,
                        count_unit,units_per_purchase_unit
                    ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                    ON CONFLICT(item_id) DO UPDATE SET
                        vendor_name=excluded.vendor_name, vendor_sku=excluded.vendor_sku,
                        item_name=excluded.item_name, normalized_description=excluded.normalized_description,
                        category=excluded.category, unit=excluded.unit, last_purchase_date=excluded.last_purchase_date,
                        previous_price=excluded.previous_price, current_price=excluded.current_price,
                        price_change_percent=excluded.price_change_percent, average_price=excluded.average_price,
                        lowest_price=excluded.lowest_price, highest_price=excluded.highest_price,
                        total_spent=excluded.total_spent, purchase_count=excluded.purchase_count,
                        review_status=excluded.review_status,
                        count_unit=CASE WHEN items.count_unit IS NULL OR items.count_unit='' THEN excluded.count_unit ELSE items.count_unit END,
                        units_per_purchase_unit=CASE WHEN items.units_per_purchase_unit IS NULL OR items.units_per_purchase_unit='' OR items.units_per_purchase_unit='1.0000' THEN excluded.units_per_purchase_unit ELSE items.units_per_purchase_unit END""",
                    (
                        item_id, vendor_key, vendor_name, sku, description, normalized_description, category, unit,
                        first_date, invoice_date, f"{first_price:.2f}",
                        f"{previous_price:.2f}" if previous_price is not None else "",
                        f"{unit_price:.2f}", f"{price_change:.2f}" if price_change is not None else "",
                        f"{new_average:.2f}", f"{min(lowest_price, unit_price):.2f}",
                        f"{max(highest_price, unit_price):.2f}", f"{total_spent + line_total:.2f}",
                        new_count, review_status, inferred_count_unit, f"{inferred_units_per:.4f}",
                    ),
                )
                notes: list[str] = []
                if new_item:
                    notes.append("New item added to Item Master")
                if match_status == "Description Match - Review":
                    notes.append("Matched by description")
                if alert:
                    notes.append(f"Price change alert: {price_change:.2f}%")
                conn.execute(
                    """INSERT INTO invoice_lines(
                        invoice_id,line_number,vendor_sku,item_id,description,normalized_description,category,
                        quantity,unit,unit_price,line_total,confidence,match_status,notes
                    ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (
                        invoice_id, index, sku, item_id, description, normalized_description, category,
                        str(quantity), unit, f"{unit_price:.2f}", f"{line_total:.2f}", confidence,
                        match_status, "; ".join(notes),
                    ),
                )
                conn.execute(
                    """INSERT INTO price_history(
                        invoice_id,invoice_date,vendor_name,vendor_sku,item_id,item_description,category,
                        quantity,unit,unit_price,line_total,previous_price,price_change_percent,price_alert,
                        confidence,match_status,notes
                    ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                    (
                        invoice_id, invoice_date, vendor_name, sku, item_id, description, category,
                        str(quantity), unit, f"{unit_price:.2f}", f"{line_total:.2f}",
                        f"{previous_price:.2f}" if previous_price is not None else "",
                        f"{price_change:.2f}" if price_change is not None else "",
                        alert, confidence, match_status, "; ".join(notes),
                    ),
                )
                if new_item:
                    conn.execute(
                        """INSERT INTO reviews(invoice_id,item_id,severity,issue_type,issue,payload_json,status,created_at)
                           VALUES(?, ?, 'WARNING', 'New Item', ?, ?, 'Open', ?)""",
                        (
                            invoice_id, item_id,
                            f"New product requires category/name review: {description}",
                            json.dumps({"item_id": item_id, "vendor": vendor_name, "sku": sku, "description": description}),
                            now_iso(),
                        ),
                    )

    def get_invoice(self, invoice_id: str) -> sqlite3.Row | None:
        with self.workspace.connect() as conn:
            return conn.execute("SELECT * FROM invoices WHERE invoice_id=?", (invoice_id,)).fetchone()

    def get_invoice_data(self, invoice_id: str) -> dict[str, Any]:
        row = self.get_invoice(invoice_id)
        if not row:
            raise KeyError(invoice_id)
        return json.loads(row["canonical_json"])

    def list_invoices(self, limit: int = 500) -> list[sqlite3.Row]:
        with self.workspace.connect() as conn:
            return conn.execute(
                "SELECT * FROM invoices ORDER BY COALESCE(invoice_date, created_at) DESC, created_at DESC LIMIT ?",
                (limit,),
            ).fetchall()

    def list_open_reviews(self) -> list[sqlite3.Row]:
        with self.workspace.connect() as conn:
            return conn.execute(
                """SELECT i.invoice_id, i.vendor, i.invoice_number, i.invoice_date, i.total, i.status,
                          i.source_original_path, i.extraction_method, i.extraction_confidence,
                          COUNT(r.review_id) AS issue_count,
                          GROUP_CONCAT(DISTINCT r.issue_type) AS issue_types
                   FROM invoices i
                   JOIN reviews r ON r.invoice_id=i.invoice_id AND r.status='Open'
                   WHERE i.status='Needs Review'
                   GROUP BY i.invoice_id
                   ORDER BY i.created_at DESC"""
            ).fetchall()

    def list_items(self, limit: int = 1000) -> list[sqlite3.Row]:
        with self.workspace.connect() as conn:
            return conn.execute(
                "SELECT * FROM items ORDER BY vendor_name, item_name LIMIT ?", (limit,)
            ).fetchall()

    def update_item(
        self, item_id: str, *, item_name: str, category: str, unit: str,
        vendor_sku: str, review_status: str = "Approved"
    ) -> None:
        normalized = normalize_text(item_name)
        if not item_name.strip():
            raise ValueError("Item name is required")
        with self.workspace.connect() as conn:
            cursor = conn.execute(
                """UPDATE items SET item_name=?, normalized_description=?, category=?, unit=?, vendor_sku=?, review_status=?
                   WHERE item_id=?""",
                (item_name.strip(), normalized, category.strip() or "Unclassified", unit.strip(), vendor_sku.strip(), review_status, item_id),
            )
            if cursor.rowcount == 0:
                raise KeyError(item_id)
            if review_status == "Approved":
                conn.execute(
                    """UPDATE reviews SET status='Resolved', resolution='Item Master approved', resolved_at=?
                       WHERE item_id=? AND issue_type='New Item' AND status='Open'""",
                    (now_iso(), item_id),
                )

    def approve_item_configuration(
        self,
        item_id: str,
        resolution: str = "Product planning configuration confirmed",
    ) -> None:
        """Approve a product after an explicit manager/imported planning setup."""
        with self.workspace.connect() as conn:
            item = conn.execute(
                """SELECT count_unit,units_per_purchase_unit,planning_confirmed
                   FROM items WHERE item_id=?""",
                (item_id,),
            ).fetchone()
            if not item:
                raise KeyError(item_id)
            if not str(item["count_unit"] or "").strip():
                raise ValueError("Count Unit is required before approving a product")
            if Decimal(str(item["units_per_purchase_unit"] or "0")) <= 0:
                raise ValueError("Units Per Purchase Unit must be greater than zero")
            if not int(item["planning_confirmed"] or 0):
                raise ValueError("Product planning must be explicitly confirmed")
            conn.execute(
                "UPDATE items SET review_status='Approved' WHERE item_id=?",
                (item_id,),
            )
            conn.execute(
                """UPDATE reviews SET status='Resolved',resolution=?,resolved_at=?
                   WHERE item_id=? AND issue_type='New Item' AND status='Open'""",
                (resolution, now_iso(), item_id),
            )

    def dashboard_summary(self) -> dict[str, Any]:
        with self.workspace.connect() as conn:
            invoice = conn.execute(
                """SELECT COUNT(*) AS count,
                          COALESCE(SUM(CASE WHEN status='Approved' THEN CAST(total AS REAL) ELSE 0 END), 0) AS purchases,
                          SUM(CASE WHEN status='Needs Review' THEN 1 ELSE 0 END) AS needs_review
                   FROM invoices"""
            ).fetchone()
            items = conn.execute("SELECT COUNT(*) AS count FROM items").fetchone()
            item_reviews = conn.execute(
                "SELECT COUNT(*) FROM items WHERE review_status<>'Approved'"
            ).fetchone()[0]
            alerts = conn.execute("SELECT COUNT(*) AS count FROM price_history WHERE price_alert=1").fetchone()
            sales_total = sum(
                (Decimal(str(row["net_sales"] or 0)) for row in preferred_sales_rows(conn)),
                Decimal("0"),
            )
            costs = conn.execute("SELECT COALESCE(SUM(CAST(amount AS REAL)),0) AS costs FROM operating_costs").fetchone()
        purchases = Decimal(str(invoice["purchases"] or 0)).quantize(MONEY)
        net_sales = sales_total.quantize(MONEY)
        other_costs = Decimal(str(costs["costs"] or 0)).quantize(MONEY)
        summary = {
            "invoice_count": int(invoice["count"] or 0),
            "approved_purchases": purchases,
            "needs_review": int(invoice["needs_review"] or 0),
            "item_count": int(items["count"] or 0),
            "item_reviews": int(item_reviews or 0),
            "price_alerts": int(alerts["count"] or 0),
            "net_sales": net_sales,
            "other_costs": other_costs,
            "sales_minus_purchases": net_sales - purchases,
            "estimated_operating_profit": net_sales - purchases - other_costs,
            "restaurant_name": self.settings.get("restaurant_name", self.workspace.root.name),
        }
        try:
            annual = self.planning.planning_dashboard()
            summary.update({
                "year_sales": annual.get("net_sales", Decimal("0")),
                "year_purchases": annual.get("invoice_purchases", Decimal("0")),
                "year_estimated_contribution": annual.get("estimated_contribution", Decimal("0")),
                "estimated_inventory_value": annual.get("estimated_inventory_value", Decimal("0")),
                "items_to_order": int(annual.get("items_to_order", 0)),
                "closed_months": int(annual.get("closed_months", 0)),
                "ready_to_close_months": int(annual.get("ready_to_close_months", 0)),
            })
        except Exception:
            summary.update({
                "year_sales": Decimal("0"), "year_purchases": Decimal("0"),
                "year_estimated_contribution": Decimal("0"),
                "estimated_inventory_value": Decimal("0"), "items_to_order": 0, "closed_months": 0,
                "ready_to_close_months": 0,
            })
        try:
            quality = self.controls.data_quality_report(save_snapshot=False)
            exceptions = self.controls.list_exceptions()
            receiving = self.controls.list_receiving_invoices()
            summary.update({
                "data_quality_score": int(quality.get("overall_score", 0)),
                "data_quality_grade": quality.get("grade", "Unknown"),
                "open_exceptions": len(exceptions),
                "critical_exceptions": sum(1 for row in exceptions if row["severity"] == "Critical"),
                "deliveries_unverified": sum(1 for row in receiving if row["receiving_status"] != "Verified"),
            })
        except Exception:
            summary.update({
                "data_quality_score": 0, "data_quality_grade": "Unavailable",
                "open_exceptions": 0, "critical_exceptions": 0, "deliveries_unverified": 0,
            })
        try:
            summary.update(self.phase2.dashboard_summary())
            summary.update(self.phase3.dashboard_summary())
        except Exception:
            summary.update({
                "pos_import_runs": 0, "menu_items": 0, "recipes_configured": 0,
                "month_waste_cost": Decimal("0"), "open_mobile_counts": 0,
                "draft_purchase_orders": 0, "last_pos_import": "",
            })
        return summary

    def import_sales_csv(self, path: Path) -> int:
        count = 0
        with path.open("r", encoding="utf-8-sig", newline="") as handle, self.workspace.connect() as conn:
            reader = csv.DictReader(handle)
            for row in reader:
                period_start = row.get("Period Start") or row.get("period_start") or row.get("date") or row.get("Date")
                period_end = row.get("Period End") or row.get("period_end") or period_start
                net_sales = row.get("Net Sales") or row.get("net_sales")
                if not period_start or not period_end or net_sales in (None, ""):
                    continue
                cursor = conn.execute(
                    """INSERT OR IGNORE INTO sales(period_start,period_end,gross_sales,discounts,refunds,sales_tax,net_sales,source_file)
                       VALUES(?,?,?,?,?,?,?,?)""",
                    (
                        parse_date(period_start), parse_date(period_end),
                        money_string(row.get("Gross Sales") or row.get("gross_sales"), "gross sales", required=False),
                        money_string(row.get("Discounts") or row.get("discounts"), "discounts", required=False),
                        money_string(row.get("Refunds") or row.get("refunds"), "refunds", required=False),
                        money_string(
                            row.get("Sales Tax Collected")
                            or row.get("Sales Tax")
                            or row.get("sales_tax_collected")
                            or row.get("sales_tax"),
                            "sales tax",
                            required=False,
                        ),
                        money_string(net_sales, "net sales"), path.name,
                    ),
                )
                count += int(cursor.rowcount > 0)
        shutil.copy2(path, self.workspace.folders["sales"] / safe_filename(path.name))
        return int(count)

    def import_sales_workbook(self, path: Path) -> int:
        from excel_io import read_table
        records = read_table(path)
        if not records:
            return 0
        count = 0
        with self.workspace.connect() as conn:
            for row in records:
                period_start = row.get("Period Start") or row.get("period_start") or row.get("date") or row.get("Date")
                period_end = row.get("Period End") or row.get("period_end") or period_start
                net_sales = row.get("Net Sales") or row.get("net_sales")
                if not period_start or not period_end or net_sales in (None, ""):
                    continue
                cursor = conn.execute(
                    """INSERT OR IGNORE INTO sales(period_start,period_end,gross_sales,discounts,refunds,sales_tax,net_sales,source_file)
                       VALUES(?,?,?,?,?,?,?,?)""",
                    (
                        parse_date(period_start), parse_date(period_end),
                        money_string(row.get("Gross Sales") or row.get("gross_sales"), "gross sales", required=False),
                        money_string(row.get("Discounts") or row.get("discounts"), "discounts", required=False),
                        money_string(row.get("Refunds") or row.get("refunds"), "refunds", required=False),
                        money_string(
                            row.get("Sales Tax Collected")
                            or row.get("Sales Tax")
                            or row.get("sales_tax_collected")
                            or row.get("sales_tax"),
                            "sales tax",
                            required=False,
                        ),
                        money_string(net_sales, "net sales"), path.name,
                    ),
                )
                count += int(cursor.rowcount > 0)
        shutil.copy2(path, self.workspace.folders["sales"] / safe_filename(path.name))
        return int(count)

    def import_operating_costs_csv(self, path: Path) -> int:
        count = 0
        with path.open("r", encoding="utf-8-sig", newline="") as handle, self.workspace.connect() as conn:
            reader = csv.DictReader(handle)
            for row in reader:
                cost_date = (
                    row.get("date") or row.get("Date") or row.get("cost_date")
                    or row.get("Cost Date")
                )
                category = row.get("category") or row.get("Category")
                description = row.get("description") or row.get("Description")
                amount = row.get("amount") or row.get("Amount")
                if not all((cost_date, category, description, amount)):
                    continue
                conn.execute(
                    "INSERT INTO operating_costs(cost_date,category,description,amount,source_file) VALUES(?,?,?,?,?)",
                    (parse_date(cost_date), category, description, money_string(amount, "operating cost"), path.name),
                )
                count += 1
        shutil.copy2(path, self.workspace.folders["costs"] / safe_filename(path.name))
        return count

    def import_operating_costs_workbook(self, path: Path) -> int:
        from excel_io import read_table
        records = read_table(path)
        if not records:
            return 0
        count = 0
        with self.workspace.connect() as conn:
            for row in records:
                cost_date = (
                    row.get("date") or row.get("Date") or row.get("cost_date")
                    or row.get("Cost Date")
                )
                category = row.get("category") or row.get("Category")
                description = row.get("description") or row.get("Description")
                amount = row.get("amount") or row.get("Amount")
                if not all((cost_date, category, description, amount)):
                    continue
                conn.execute(
                    "INSERT INTO operating_costs(cost_date,category,description,amount,source_file) VALUES(?,?,?,?,?)",
                    (parse_date(cost_date), category, description, money_string(amount, "operating cost"), path.name),
                )
                count += 1
        shutil.copy2(path, self.workspace.folders["costs"] / safe_filename(path.name))
        return count

    def export_csvs(self) -> list[Path]:
        exports: list[Path] = []
        table_queries = {
            "invoice_log.csv": "SELECT * FROM invoices ORDER BY invoice_date, vendor, invoice_number",
            "invoice_line_items.csv": "SELECT * FROM invoice_lines ORDER BY invoice_id, line_number",
            "item_master.csv": "SELECT * FROM items ORDER BY vendor_name, item_name",
            "item_price_history.csv": "SELECT * FROM price_history ORDER BY invoice_date, vendor_name, item_description",
            "review_queue.csv": "SELECT * FROM reviews ORDER BY status, created_at DESC",
            "sales.csv": "SELECT * FROM sales ORDER BY period_start",
            "operating_costs.csv": "SELECT * FROM operating_costs ORDER BY cost_date",
            "inventory_counts.csv": "SELECT c.*,i.item_name,i.vendor_name,i.vendor_sku,i.category FROM inventory_counts c JOIN items i ON i.item_id=c.item_id ORDER BY c.count_date,i.item_name",
            "monthly_item_usage.csv": "SELECT * FROM monthly_item_usage ORDER BY month,vendor_name,item_name",
            "monthly_closes.csv": "SELECT * FROM monthly_closes ORDER BY month",
            "order_batches.csv": "SELECT * FROM order_batches ORDER BY created_at DESC",
            "order_predictions.csv": "SELECT * FROM order_predictions ORDER BY batch_id,vendor_name,item_name",
            "receiving_sessions.csv": "SELECT * FROM receiving_sessions ORDER BY invoice_date DESC",
            "receiving_lines.csv": "SELECT * FROM receiving_lines ORDER BY session_id,receiving_line_id",
            "operational_exceptions.csv": "SELECT * FROM operational_exceptions ORDER BY status,severity,last_detected DESC",
            "audit_log.csv": "SELECT * FROM audit_log ORDER BY audit_id DESC",
            "backup_history.csv": "SELECT * FROM backup_history ORDER BY created_at DESC",
            "data_quality_snapshots.csv": "SELECT * FROM data_quality_snapshots ORDER BY created_at DESC",
            "users.csv": "SELECT user_id,username,display_name,role,active,created_at,updated_at,last_login FROM users ORDER BY username",
            "margin_memory_decisions.csv": "SELECT * FROM margin_memory_decisions ORDER BY decision_time DESC",
            "margin_memory_context.csv": "SELECT * FROM margin_memory_context ORDER BY created_at DESC",
            "margin_memory_outcomes.csv": "SELECT * FROM margin_memory_outcomes ORDER BY evaluated_at DESC",
            "margin_memory_recommendations.csv": "SELECT * FROM margin_memory_recommendations ORDER BY generated_at DESC",
            "costpilot_review_actions.csv": "SELECT * FROM costpilot_review_actions ORDER BY created_at DESC",
            "costpilot_review_resolutions.csv": "SELECT * FROM costpilot_review_resolutions ORDER BY updated_at DESC",
        }
        with self.workspace.connect() as conn:
            for filename, query in table_queries.items():
                path = self.workspace.folders["exports"] / filename
                rows = conn.execute(query).fetchall()
                records = [dict(row) for row in rows]
                write_table_as(path, records, "csv")
                exports.append(path)
        try:
            exports.append(self.planning.export_full_inventory_csv())
        except Exception:
            pass
        try:
            latest = self.planning.latest_order_batch()
            if latest:
                exports.append(self.planning.export_order_sheet_csv(latest["batch_id"]))
        except Exception:
            pass
        try:
            exports.extend(self.phase2.export_csvs())
            exports.extend(self.phase3.export_csvs())
        except Exception:
            pass
        return exports

    def export_workbook(self, destination: Path | None = None) -> Path:
        try:
            from openpyxl import Workbook  # type: ignore
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise PipelineError("openpyxl is required for workbook export") from exc
        destination = destination or (self.workspace.folders["exports"] / "Restaurant_Cost_Report.xlsx")
        try:
            self.planning.estimate_inventory()
        except Exception:
            pass
        wb = Workbook()
        wb.remove(wb.active)
        header_fill = PatternFill("solid", fgColor="17324D")
        header_font = Font(color="FFFFFF", bold=True)
        tables = {
            "Dashboard": None,
            "Annual Summary": None,
            "Full Inventory": """SELECT i.item_id,i.vendor_name,i.vendor_sku,i.item_name,i.category,
                    i.unit AS purchase_unit,i.count_unit,i.units_per_purchase_unit,i.current_price,i.average_price,
                    i.lowest_price,i.highest_price,i.estimated_on_hand,i.estimated_on_hand_as_of,
                    (SELECT c.count_date FROM inventory_counts c WHERE c.item_id=i.item_id AND c.finalized=1 ORDER BY c.count_date DESC LIMIT 1) AS last_physical_count,
                    (SELECT u.average_daily_usage FROM monthly_item_usage u WHERE u.item_id=i.item_id ORDER BY u.month DESC LIMIT 1) AS average_daily_usage,
                    (SELECT u.average_weekly_usage FROM monthly_item_usage u WHERE u.item_id=i.item_id ORDER BY u.month DESC LIMIT 1) AS average_weekly_usage,
                    CASE WHEN CAST(COALESCE(i.units_per_purchase_unit,'1') AS REAL)>0
                         THEN ROUND(CAST(COALESCE(i.estimated_on_hand,'0') AS REAL) * CAST(COALESCE(i.current_price,'0') AS REAL) / CAST(COALESCE(i.units_per_purchase_unit,'1') AS REAL),2)
                         ELSE 0 END AS estimated_inventory_value,
                    i.lead_time_days,i.order_cycle_days,i.safety_stock_days,i.order_multiple,i.par_override_count_units,
                    (SELECT o.par_quantity_count_units FROM order_predictions o WHERE o.item_id=i.item_id ORDER BY o.prediction_id DESC LIMIT 1) AS latest_par_quantity,
                    (SELECT o.suggested_order_quantity FROM order_predictions o WHERE o.item_id=i.item_id ORDER BY o.prediction_id DESC LIMIT 1) AS latest_suggested_order,
                    (SELECT o.manager_order_quantity FROM order_predictions o WHERE o.item_id=i.item_id ORDER BY o.prediction_id DESC LIMIT 1) AS latest_manager_order,
                    i.review_status
                FROM items i WHERE i.active=1 ORDER BY i.category,i.vendor_name,i.item_name""",
            "Invoice Log": "SELECT invoice_date,vendor,invoice_number,subtotal,fees,tax,credits,total,status,extraction_method,extraction_confidence,source_original_path FROM invoices ORDER BY invoice_date",
            "Invoice Lines": "SELECT invoice_id,line_number,vendor_sku,item_id,description,category,quantity,unit,unit_price,line_total,confidence,match_status,notes FROM invoice_lines ORDER BY invoice_id,line_number",
            "Item Master": "SELECT * FROM items ORDER BY vendor_name,item_name",
            "Price History": "SELECT * FROM price_history ORDER BY invoice_date,vendor_name,item_description",
            "Review Queue": "SELECT * FROM reviews ORDER BY status,created_at DESC",
            "Sales": "SELECT * FROM sales ORDER BY period_start",
            "Operating Costs": "SELECT * FROM operating_costs ORDER BY cost_date",
            "Inventory Counts": "SELECT c.*,i.item_name,i.vendor_name,i.vendor_sku,i.category FROM inventory_counts c JOIN items i ON i.item_id=c.item_id ORDER BY c.count_date,i.item_name",
            "Monthly Usage": "SELECT * FROM monthly_item_usage ORDER BY month,vendor_name,item_name",
            "Monthly Closes": "SELECT * FROM monthly_closes ORDER BY month",
            "Order Predictions": "SELECT * FROM order_predictions ORDER BY batch_id,vendor_name,item_name",
            "Receiving Sessions": "SELECT * FROM receiving_sessions ORDER BY invoice_date DESC",
            "Receiving Lines": "SELECT * FROM receiving_lines ORDER BY session_id,receiving_line_id",
            "Exceptions": "SELECT * FROM operational_exceptions ORDER BY status,severity,last_detected DESC",
            "Audit Log": "SELECT created_at,username,role,action,entity_type,entity_id,summary FROM audit_log ORDER BY audit_id DESC",
            "Backup History": "SELECT created_at,created_by,backup_type,file_path,size_bytes,sha256,status,notes FROM backup_history ORDER BY created_at DESC",
            "Data Quality": "SELECT created_at,overall_score,completeness_score,freshness_score,integrity_score,operational_score,grade FROM data_quality_snapshots ORDER BY created_at DESC",
            "MarginMemory Decisions": "SELECT decision_time,decision_type,subject_name,reason_code,manager_note,decision_maker,decision_maker_role,override_amount,override_percent,status,evaluation_start_date,evaluation_end_date FROM margin_memory_decisions ORDER BY decision_time DESC",
            "MarginMemory Context": "SELECT * FROM margin_memory_context ORDER BY created_at DESC",
            "MarginMemory Outcomes": "SELECT * FROM margin_memory_outcomes ORDER BY evaluated_at DESC",
            "CostPilot Review Actions": "SELECT * FROM costpilot_review_actions ORDER BY created_at DESC",
            "CostPilot Review Resolutions": "SELECT * FROM costpilot_review_resolutions ORDER BY updated_at DESC",
            "POS Import Runs": "SELECT * FROM pos_import_runs ORDER BY imported_at DESC",
            "POS Sales Lines": "SELECT business_date,order_id,location,pos_item_key,menu_item_name,quantity,unit_price,gross_sales,discounts,refunds,net_sales,sales_tax,channel,modifiers FROM pos_sales_lines ORDER BY business_date,sale_line_id",
            "Menu Items": "SELECT * FROM menu_items ORDER BY category,menu_item_name",
            "Recipes": "SELECT m.menu_item_name,i.item_name,r.quantity_count_units,r.yield_percent,r.notes FROM recipe_ingredients r JOIN menu_items m ON m.menu_item_id=r.menu_item_id JOIN items i ON i.item_id=r.item_id ORDER BY m.menu_item_name,i.item_name",
            "Waste Log": "SELECT w.event_date,i.item_name,i.vendor_name,w.quantity_count_units,i.count_unit,w.reason,w.shift,w.estimated_cost,w.notes,w.created_by FROM waste_events w JOIN items i ON i.item_id=w.item_id ORDER BY w.event_date DESC",
            "Mobile Counts": "SELECT s.session_id,s.count_date,s.status,s.created_by,s.created_at,s.submitted_at,s.finalized_at,COUNT(e.entry_id) AS entry_count FROM mobile_count_sessions s LEFT JOIN mobile_count_entries e ON e.session_id=s.session_id GROUP BY s.session_id ORDER BY s.created_at DESC",
            "Purchase Orders": "SELECT * FROM purchase_orders ORDER BY po_date DESC,vendor_name",
            "Purchase Order Lines": "SELECT * FROM purchase_order_lines ORDER BY po_id,item_name",
            "Accounting Exports": "SELECT * FROM accounting_export_history ORDER BY created_at DESC",
        }
        with self.workspace.connect() as conn:
            for sheet_name, query in tables.items():
                ws = wb.create_sheet(sheet_name)
                if sheet_name == "Dashboard":
                    summary = self.dashboard_summary()
                    ws.append([self.settings.get("restaurant_name", "Restaurant Cost Controller")])
                    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
                    ws["A1"].fill = header_fill
                    ws.merge_cells("A1:B1")
                    ws.append(["Metric", "Value"])
                    for cell in ws[2]:
                        cell.fill = header_fill
                        cell.font = header_font
                    for key, value in summary.items():
                        label = key.replace("_", " ").title()
                        ws.append([label, float(value) if isinstance(value, Decimal) else value])
                    ws.append([])
                    ws.append(["Manager Note", "Yearly contribution and product margin are estimates. Labor is included only when imported; waste is absorbed into inventory depletion but not separately identified."])
                    ws.column_dimensions["A"].width = 32
                    ws.column_dimensions["B"].width = 20
                    continue
                if sheet_name == "Annual Summary":
                    year = datetime.now().year
                    annual_rows = self.planning.year_summary(year)
                    ws.append([f"{year} Annual Summary"])
                    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
                    ws["A1"].fill = header_fill
                    headers = ["month","net_sales","invoice_purchases","product_purchases","opening_inventory_value","ending_inventory_value","estimated_cogs","estimated_product_margin","estimated_product_margin_percent","imported_operating_costs","estimated_contribution","count_status"]
                    ws.append([h.replace("_"," ").title() for h in headers])
                    for cell in ws[2]:
                        cell.fill = header_fill; cell.font = header_font
                    for annual_row in annual_rows:
                        ws.append([annual_row.get(h, "") for h in headers])
                    for col_index, column in enumerate(ws.columns, 1):
                        max_len = max(len(str(cell.value or "")) for cell in column)
                        ws.column_dimensions[get_column_letter(col_index)].width = min(max(max_len + 2, 12), 32)
                    ws.freeze_panes = "A3"
                    continue
                rows = conn.execute(query).fetchall()
                if not rows:
                    continue
                ws.append(list(rows[0].keys()))
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                for row in rows:
                    ws.append(list(row))
                for col_index, column in enumerate(ws.columns, 1):
                    max_len = max(len(str(cell.value or "")) for cell in column)
                    ws.column_dimensions[get_column_letter(col_index)].width = min(max(max_len + 2, 12), 45)
                ws.freeze_panes = "A2"
                ws.auto_filter.ref = ws.dimensions
        destination.parent.mkdir(parents=True, exist_ok=True)
        wb.save(destination)
        return destination


    # Annual inventory, month-close, and order-planning delegates.
    def export_inventory_count_sheet(self, month: str, destination: Path | None = None) -> Path:
        return self.planning.export_count_sheet_csv(month, destination)

    def import_inventory_count_csv(self, path: Path, count_date: Any | None = None):
        return self.planning.import_count_csv(path, count_date)

    def import_inventory_count_workbook(self, path: Path, count_date: Any | None = None):
        from excel_io import read_table
        records = read_table(path)
        if not records:
            return self.planning.import_count_csv(path.with_suffix('.csv'), count_date)
        temp_csv = path.with_suffix('.temp.csv')
        try:
            headers = list(records[0].keys())
            with temp_csv.open('w', encoding='utf-8-sig', newline='') as fh:
                writer = csv.DictWriter(fh, fieldnames=headers, extrasaction='ignore')
                writer.writeheader()
                writer.writerows(records)
            return self.planning.import_count_csv(temp_csv, count_date)
        finally:
            temp_csv.unlink(missing_ok=True)

    def list_inventory_counts(self, month: str | None = None):
        return self.planning.list_counts(month)

    def close_inventory_month(self, month: str) -> dict[str, Any]:
        return self.planning.close_month(month)

    def monthly_usage(self, month: str):
        return self.planning.list_month_usage(month)

    def annual_summary(self, year: int):
        return self.planning.year_summary(year)

    def annual_totals(self, year: int):
        return self.planning.year_totals(year)

    def estimated_inventory(self, as_of: Any | None = None):
        return self.planning.estimate_inventory(as_of)

    def generate_order_sheet(self, as_of: Any | None = None):
        return self.planning.generate_order_predictions(as_of)

    def ensure_weekly_order_draft(self, as_of: Any | None = None):
        return self.planning.ensure_weekly_order_draft(as_of)

    def latest_order_batch(self):
        return self.planning.latest_order_batch()

    def list_order_predictions(self, batch_id: str | None = None):
        return self.planning.list_order_predictions(batch_id)

    def update_order_prediction(
        self, prediction_id: int, manager_qty: Any, status: str = "Reviewed",
        *, reason_code: str | None = None, manager_note: str | None = None,
    ) -> str | None:
        self.planning.update_order_prediction(prediction_id, manager_qty, status, manager_note)
        return self.margin_memory.capture_order_override(
            prediction_id, manager_qty, reason_code=reason_code, manager_note=manager_note,
            status="Pending Approval",
        )

    def approve_order_batch(self, batch_id: str) -> dict[str, Any]:
        self.planning.approve_order_batch(batch_id)
        return self.margin_memory.finalize_order_batch(batch_id)

    def export_order_sheet(self, batch_id: str | None = None, destination: Path | None = None) -> Path:
        return self.planning.export_order_sheet_csv(batch_id, destination)

    def export_full_inventory(self, destination: Path | None = None) -> Path:
        return self.planning.export_full_inventory_csv(destination)

    def update_item_planning(self, item_id: str, **values: Any) -> None:
        self.planning.update_item_planning(item_id, **values)

    # Operational controls delegates.
    def automatic_backup_if_due(self):
        return self.controls.automatic_backup_if_due()

    def create_backup(self, backup_type: str = "Manual", notes: str = ""):
        return self.controls.create_backup(backup_type, notes)

    def restore_backup(self, path: Path) -> None:
        self.controls.restore_backup(path)

    def list_backups(self, limit: int = 100):
        return self.controls.list_backups(limit)

    def data_quality_report(self, save_snapshot: bool = True):
        return self.controls.data_quality_report(save_snapshot=save_snapshot)

    def refresh_exceptions(self):
        return self.controls.refresh_exceptions()

    def list_exceptions(self, include_resolved: bool = False, limit: int = 500):
        return self.controls.list_exceptions(include_resolved=include_resolved, limit=limit)

    def set_exception_status(self, exception_id: int, status: str, resolution: str = "") -> None:
        self.controls.set_exception_status(exception_id, status, resolution)

    def list_receiving_invoices(self, limit: int = 500):
        return self.controls.list_receiving_invoices(limit)

    def start_receiving(self, invoice_id: str) -> str:
        return self.controls.start_receiving(invoice_id)

    def get_receiving(self, session_id: str):
        return self.controls.get_receiving(session_id)

    def save_receiving(self, session_id: str, lines, **kwargs):
        result = self.controls.save_receiving(session_id, lines, **kwargs)
        self.margin_memory.capture_receiving_discrepancies(session_id)
        return result

    def auto_verify_receiving(self, invoice_ids: Iterable[str] | None = None, **kwargs):
        settings = self.workspace.load_settings()
        kwargs.setdefault("date_mode", str(settings.get("auto_verify_receiving_date_mode") or "invoice_date"))
        return self.controls.auto_verify_receiving(invoice_ids, **kwargs)

    def list_audit(self, limit: int = 1000, entity_type: str | None = None):
        return self.controls.list_audit(limit, entity_type=entity_type)

    # Phase 2 delegates.
    def import_pos_report(self, path: Path, **kwargs):
        return self.phase2.import_pos_report(path, **kwargs)

    def list_pos_runs(self, limit: int = 200):
        return self.phase2.list_pos_runs(limit)

    def list_menu_costs(self, start: str | None = None, end: str | None = None):
        return self.phase2.list_menu_costs(start, end)

    def import_recipes_csv(self, path: Path):
        return self.phase2.import_recipes_csv(path)

    def export_recipe_template(self, destination: Path | None = None):
        return self.phase2.export_recipe_template(destination)

    def log_waste(self, item_id: str, quantity: Any, reason: str, **kwargs):
        return self.phase2.log_waste(item_id, quantity, reason, **kwargs)

    def list_waste(self, start: str | None = None, end: str | None = None, limit: int = 1000):
        return self.phase2.list_waste(start, end, limit)

    def generate_purchase_orders(self, batch_id: str | None = None, **kwargs):
        return self.phase2.generate_purchase_orders(batch_id, **kwargs)

    def list_purchase_orders(self, limit: int = 300):
        return self.phase2.list_purchase_orders(limit)

    def export_purchase_orders(self, po_ids=None, destination: Path | None = None):
        return self.phase2.export_purchase_orders(po_ids, destination)

    def export_accounting(self, start: str, end: str, export_type: str = "General Journal CSV"):
        return self.phase2.export_accounting(start, end, export_type)


    # ---------- Phase 3 intelligence wrappers ----------
    def portfolio_summary(self, year: int | None = None):
        return self.phase3.portfolio_summary(year)

    def create_inventory_transfer(self, destination_path: Path, lines: Iterable[dict[str, Any]], **kwargs):
        transfer_id = self.phase3.create_transfer(destination_path, lines, **kwargs)
        self.margin_memory.capture_transfer(transfer_id)
        return transfer_id

    def receive_inventory_transfer(self, transfer_id: str, **kwargs):
        return self.phase3.receive_transfer(transfer_id, **kwargs)

    def list_inventory_transfers(self, limit: int = 300):
        return self.phase3.list_transfers(limit)

    # CostPilot Review Center delegates.
    def list_costpilot_review_cases(self):
        return self.review_copilot.list_cases()

    def get_costpilot_review_case(self, case_id: str):
        return self.review_copilot.get_case(case_id)

    def costpilot_review_summary(self):
        return self.review_copilot.summary()

    def costpilot_review_introduction(self):
        return self.review_copilot.queue_introduction()

    def explain_costpilot_review_case(self, case_id: str):
        return self.review_copilot.explain_case(case_id)

    def parse_costpilot_review_command(self, text: str, selected_case_ids=()):
        return self.review_copilot.parse_command(text, selected_case_ids)

    def preview_costpilot_review_action(self, action: str, case_ids=None):
        return self.review_copilot.preview(action, case_ids)

    def execute_costpilot_review_action(self, action: str, case_ids=None, **kwargs):
        return self.review_copilot.execute(action, case_ids, **kwargs)

    # MarginMemory delegates.
    def margin_memory_summary(self):
        return self.margin_memory.summary()

    def list_margin_memory_decisions(self, **kwargs):
        return self.margin_memory.list_decisions(**kwargs)

    def get_margin_memory_decision(self, decision_id: str):
        return self.margin_memory.get_decision(decision_id)

    def margin_memory_filter_options(self):
        return self.margin_memory.filter_options()

    def export_margin_memory_decisions(self, destination: Path | None = None):
        return self.margin_memory.export_decisions_csv(destination)

    def add_local_event(self, event_name: str, event_date: str, **kwargs):
        return self.phase3.add_event(event_name, event_date, **kwargs)

    def import_event_calendar(self, path: Path, **kwargs):
        return self.phase3.import_ics(path, **kwargs)

    def refresh_weather_forecast(self, forecast_days: int = 16, **kwargs):
        return self.phase3.refresh_weather(forecast_days, **kwargs)

    def generate_demand_forecasts(self, start: str | None = None, days: int = 14):
        return self.phase3.generate_forecast_range(start, days)

    def learn_forecasts(self):
        return self.phase3.learn_from_actuals()

    def menu_profitability(self, start: str | None = None, end: str | None = None):
        return self.phase3.menu_profitability(start, end, self.settings.get("target_menu_food_cost_percent", 30))

    def usage_variance(self, month: str):
        return self.phase3.usage_variance(month)

    def generate_sales_driven_orders(self, as_of: str | None = None, forecast_days: int = 9):
        return self.phase3.generate_sales_driven_order_batch(as_of, forecast_days)

    def savings_dashboard(self, start: str | None = None, end: str | None = None):
        return self.phase3.savings_dashboard(start, end)

    def export_owner_report(self, start: str, end: str, destination: Path | None = None):
        return self.phase3.export_owner_report(start, end, destination)

    def weekly_invoice_log(self, week_start: str | None = None) -> dict:
        """Generate a weekly invoice log PDF for the past 7 days.

        Args:
            week_start: ISO date string for Monday of the target week.
                        Defaults to the most recent Monday.

        Returns:
            Dict with keys: pdf_path, week_start, week_end, invoice_count, total_cost
        """
        from weekly_invoice_log import generate_weekly_invoice_log, WeeklyInvoiceRow

        ws = self.workspace
        reference = date.fromisoformat(week_start) if week_start else None
        pdf_path = generate_weekly_invoice_log(ws, reference)

        if pdf_path is None:
            empty_start, empty_end = _get_week_range(reference)
            return {
                "pdf_path": None,
                "week_start": str(empty_start),
                "week_end": str(empty_end),
                "invoice_count": 0,
                "total_cost": "0.00",
            }

        conn = ws.connect()
        try:
            week_start_dt, week_end_dt = _get_week_range(reference)
            raw_invoices = _get_week_invoices(conn, week_start_dt, week_end_dt)
            invoices = [
                WeeklyInvoiceRow(
                    invoice_id=row["invoice_id"],
                    invoice_number=row["invoice_number"] or "",
                    vendor=row["vendor"] or "",
                    vendor_key="",
                    invoice_date=row["invoice_date"] or "",
                    total=row["total"] or "0.00",
                    source_path=row["source_original_path"],
                    source_archive_path=row["source_archive_path"],
                    source_name=row["source_name"] or "",
                )
                for row in raw_invoices
            ]
            total_cost = sum(
                float(inv.total) for inv in invoices
                if inv.total
            )
        finally:
            conn.close()

        return {
            "pdf_path": str(pdf_path),
            "week_start": str(week_start_dt),
            "week_end": str(week_end_dt),
            "invoice_count": len(invoices),
            "total_cost": f"{total_cost:,.2f}",
        }


def _get_week_range(reference: date | None = None) -> tuple[date, date]:
    monday = (reference or date.today()) - timedelta(days=(reference or date.today()).weekday())
    sunday = monday + timedelta(days=6)
    return monday, sunday


def _get_week_invoices(conn: sqlite3.Connection, week_start: date, week_end: date) -> list:
    return conn.execute(
        """
        SELECT invoice_id, invoice_number, vendor, invoice_date, total,
               source_original_path, source_archive_path, source_name
          FROM invoices
         WHERE status IN ('Processed', 'Approved', 'Reviewed')
           AND invoice_date >= ?
           AND invoice_date <= ?
         ORDER BY invoice_date ASC, vendor ASC
        """,
        (week_start.isoformat(), week_end.isoformat()),
    ).fetchall()


def discover_sources(path: Path) -> list[Path]:
    if path.is_file():
        return [path] if path.suffix.lower() in SUPPORTED_SOURCE_SUFFIXES else []
    if path.is_dir():
        return sorted(p for p in path.rglob("*") if p.is_file() and p.suffix.lower() in SUPPORTED_SOURCE_SUFFIXES)
    return []


def build_cli() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path, nargs="?", help="Invoice PDF/image/JSON or a directory")
    parser.add_argument("--workspace", type=Path, required=True, help="Restaurant workspace directory")
    parser.add_argument("--export", action="store_true", help="Export CSVs and workbook after processing")
    parser.add_argument("--list-review", action="store_true", help="Print open review queue")
    return parser


def main() -> int:
    args = build_cli().parse_args()
    workspace = RestaurantWorkspace(args.workspace)
    pipeline = InvoicePipeline(workspace)
    if args.list_review:
        print(json.dumps([dict(row) for row in pipeline.list_open_reviews()], indent=2))
        return 0
    if not args.input:
        raise SystemExit("input is required unless --list-review is used")
    sources = discover_sources(args.input)
    results = pipeline.process_many(sources)
    if args.export:
        pipeline.export_csvs()
        pipeline.export_workbook()
    summary = {
        "processed": len(results),
        "approved": sum(r.status == "Approved" for r in results),
        "needs_review": sum(r.status == "Needs Review" for r in results),
        "duplicates": sum(r.status == "Duplicate" for r in results),
        "failed": sum(r.status == "Failed" for r in results),
        "results": [r.as_dict() for r in results],
    }
    print(json.dumps(summary, indent=2))
    if summary["failed"]:
        return 3
    if summary["needs_review"]:
        return 2
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
