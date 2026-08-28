from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parent / "bar-grill-month"
EXPECTED_CATEGORIES = {"Dry Goods", "Frozen Goods", "Dairy", "Produce", "Beverage", "Liquor", "Misc"}


def main() -> None:
    inventory = load_workbook(ROOT / "inventory_counts.xlsx", read_only=True, data_only=True)
    assert inventory.sheetnames == ["Beginning Inventory", "Ending Inventory"]
    for sheet_name, expected_date in (("Beginning Inventory", "2026-07-01"), ("Ending Inventory", "2026-07-31")):
        rows = list(inventory[sheet_name].iter_rows(values_only=True))
        assert rows[0][2] == expected_date
        assert len(rows) - 2 == 30
        assert {row[5] for row in rows[2:]} == EXPECTED_CATEGORIES
        assert all(float(row[7]) >= 0 for row in rows[2:])

    recipes = load_workbook(ROOT / "recipe_guide.xlsx", read_only=True, data_only=True)
    recipe_rows = list(recipes.active.iter_rows(values_only=True))
    assert len(recipe_rows) - 1 >= 15
    assert {row[2] for row in recipe_rows[1:]} == {"Bar & Grill"}
    assert len({row[1] for row in recipe_rows[1:]}) == 24

    with (ROOT / "sales_detail_july_2026.csv").open(encoding="utf-8-sig", newline="") as handle:
        sales = list(csv.DictReader(handle))
    assert len(sales) == 684
    assert {row["Business Date"] for row in sales} == {f"2026-07-{day:02d}" for day in range(1, 32)}
    assert {row["Category"] for row in sales} == EXPECTED_CATEGORIES

    pdfs = sorted((ROOT / "invoices").glob("*.pdf"))
    assert len(pdfs) == 35
    for path in pdfs:
        reader = PdfReader(str(path))
        assert len(reader.pages) == 1
        assert not any((page.extract_text() or "").strip() for page in reader.pages)
        assert path.stat().st_size > 10000

    daily_reports = sorted((ROOT / "daily_sales_reports").glob("*.csv"))
    assert len(daily_reports) == 31
    for path in daily_reports:
        with path.open(encoding="utf-8-sig", newline="") as handle:
            rows = list(csv.DictReader(handle))
        assert rows, f"{path.name} is empty"
        assert {row["Business Date"] for row in rows} == {path.stem.replace("daily_sales_", "")}
    expected_csv_rows = {
        "daily_sales_july_2026.csv": 31,
        "sales_detail_july_2026.csv": 684,
        "sales_summary_july_2026.csv": 1,
    }
    for filename, count in expected_csv_rows.items():
        with (ROOT / filename).open(encoding="utf-8-sig", newline="") as handle:
            assert len(list(csv.reader(handle))) - 1 == count
    print("bar-grill dataset verification: OK")


if __name__ == "__main__":
    main()
